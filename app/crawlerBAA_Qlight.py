"""
Crawler BAA Qlight - Cào dữ liệu sản phẩm từ https://baa.vn/vn/qlight/
Hỗ trợ đa luồng, chuyển đổi ảnh WebP và xử lý series
"""

import os
import shutil
import time
import zipfile
from datetime import datetime
import pandas as pd
import re
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup
import requests
import traceback
from queue import Queue
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from PIL import Image
import io
import hashlib
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import openpyxl.styles
import tempfile

# Import các hàm từ module crawler hiện có
from app.crawler import (
    is_category_url, is_product_url, extract_product_urls, extract_product_info,
    download_baa_product_images_fixed, get_html_content, HEADERS
)
from app.webp_converter import WebPConverter
from app import socketio

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class BAAQlightCrawler:
    """Crawler chuyên dụng cho BAA Qlight với hỗ trợ đa luồng và xử lý series"""
    
    def __init__(self, base_url="https://baa.vn/vn/qlight/", max_workers=10):
        self.base_url = base_url
        self.max_workers = max_workers
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        
    def extract_series_info(self, url):
        """
        Trích xuất thông tin series từ trang sản phẩm
        
        Args:
            url (str): URL trang sản phẩm
            
        Returns:
            dict: Thông tin series bao gồm tên và URL
        """
        try:
            html = get_html_content(url)
            if not html:
                return None
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Phương pháp 1: Tìm theo cấu trúc HTML đã cung cấp
            series_header = soup.select_one('div.product__symbol-header.px-2.py-1.m-0.row')
            if not series_header:
                # Thử tìm với selector khác
                series_header = soup.select_one('div.product__symbol-header')
            
            if series_header:
                # Tìm label "Series:"
                series_label = series_header.select_one('span.product__symbol-label')
                if series_label and 'Series:' in series_label.text:
                    # Tìm link series trong span.product__symbol__value
                    series_value = series_header.select_one('span.product__symbol__value')
                    if series_value:
                        series_link = series_value.select_one('a.text-decoration-none.text-link')
                        if series_link:
                            series_name = series_link.get_text(strip=True)
                            series_url = urljoin(self.base_url, series_link.get('href', ''))
                            
                            return {
                                'name': series_name,
                                'url': series_url,
                                'original_url': url
                            }
            
            # Phương pháp 2: Tìm series từ breadcrumb
            breadcrumb = soup.select_one('.breadcrumb, .breadcrumb-item')
            if breadcrumb:
                breadcrumb_links = breadcrumb.select('a[href*="series"]')
                for link in breadcrumb_links:
                    href = link.get('href', '')
                    if 'series' in href.lower():
                        series_name = link.get_text(strip=True)
                        series_url = urljoin(self.base_url, href)
                        
                        return {
                            'name': series_name,
                            'url': series_url,
                            'original_url': url
                        }
            
            # Phương pháp 3: Tìm series từ URL pattern
            if 'series' in url.lower():
                # Trích xuất tên series từ URL
                url_parts = url.split('/')
                for i, part in enumerate(url_parts):
                    if 'series' in part.lower() and i + 1 < len(url_parts):
                        series_name = url_parts[i + 1].replace('_', ' ').title()
                        series_url = url
                        
                        return {
                            'name': series_name,
                            'url': series_url,
                            'original_url': url
                        }
            
            # Phương pháp 4: Tìm series từ meta tags
            meta_series = soup.select_one('meta[name="series"], meta[property="series"]')
            if meta_series:
                series_name = meta_series.get('content', '')
                if series_name:
                    # Tạo URL series từ tên
                    series_url = urljoin(self.base_url, f"/vn/series/{series_name.lower().replace(' ', '-')}/")
                    
                    return {
                        'name': series_name,
                        'url': series_url,
                        'original_url': url
                    }
            
            return None
            
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất thông tin series từ {url}: {str(e)}")
            return None
    
    def get_products_from_series(self, series_url):
        """
        Lấy danh sách sản phẩm từ một series
        
        Args:
            series_url (str): URL của series
            
        Returns:
            list: Danh sách URL sản phẩm trong series
        """
        try:
            html = get_html_content(series_url)
            if not html:
                return []
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Tìm tất cả link sản phẩm trong series
            product_urls = []
            
            # Phương pháp 1: Tìm trong các container sản phẩm
            product_containers = soup.select('.product-item, .product-card, .product, [class*="product"], .item, .card')
            
            for container in product_containers:
                links = container.select('a[href*="/vn/"]')
                for link in links:
                    href = link.get('href')
                    if href and is_product_url(href):
                        full_url = urljoin(self.base_url, href)
                        if full_url not in product_urls:
                            product_urls.append(full_url)
            
            # Phương pháp 2: Tìm tất cả link có pattern sản phẩm
            if not product_urls:
                all_links = soup.select('a[href*="/vn/"]')
                for link in all_links:
                    href = link.get('href')
                    if href and is_product_url(href):
                        full_url = urljoin(self.base_url, href)
                        if full_url not in product_urls:
                            product_urls.append(full_url)
            
            # Phương pháp 3: Tìm theo pattern URL cụ thể của BAA
            if not product_urls:
                # Tìm các link có pattern _số ở cuối (ID sản phẩm)
                pattern_links = soup.select('a[href*="_"]')
                for link in pattern_links:
                    href = link.get('href')
                    if href and re.search(r'_\d+/?$', href):
                        full_url = urljoin(self.base_url, href)
                        if full_url not in product_urls:
                            product_urls.append(full_url)
            
            # Loại bỏ các URL không phải sản phẩm
            filtered_urls = []
            for url in product_urls:
                if is_product_url(url):
                    filtered_urls.append(url)
            
            logger.info(f"Tìm thấy {len(filtered_urls)} sản phẩm trong series: {series_url}")
            return filtered_urls
            
        except Exception as e:
            logger.error(f"Lỗi khi lấy sản phẩm từ series {series_url}: {str(e)}")
            return []
    
    def crawl_product_info(self, url, index=1):
        """
        Cào thông tin một sản phẩm
        
        Args:
            url (str): URL sản phẩm
            index (int): Số thứ tự
            
        Returns:
            dict: Thông tin sản phẩm
        """
        try:
            # Sử dụng hàm extract_product_info có sẵn
            product_info = extract_product_info(url, index=index)
            
            # Thêm thông tin series
            series_info = self.extract_series_info(url)
            if series_info:
                product_info['Series'] = series_info['name']
                product_info['Series_URL'] = series_info['url']
            else:
                product_info['Series'] = ''
                product_info['Series_URL'] = ''
            
            return product_info
            
        except Exception as e:
            logger.error(f"Lỗi khi cào thông tin sản phẩm {url}: {str(e)}")
            return None
    
    def crawl_products_multithread(self, product_urls):
        """
        Cào thông tin sản phẩm sử dụng đa luồng
        
        Args:
            product_urls (list): Danh sách URL sản phẩm
            
        Returns:
            list: Danh sách thông tin sản phẩm
        """
        results = []
        lock = threading.Lock()
        completed_count = 0
        total_count = len(product_urls)
        
        def process_product(args):
            nonlocal completed_count
            url, index = args
            try:
                product_info = self.crawl_product_info(url, index)
                if product_info:
                    with lock:
                        results.append(product_info)
                        completed_count += 1
                        progress = (completed_count / total_count) * 100
                        logger.info(f"✓ [{completed_count}/{total_count}] ({progress:.1f}%) - {product_info.get('Tên sản phẩm', 'Unknown')[:50]}...")
                else:
                    with lock:
                        completed_count += 1
                        logger.warning(f"⚠️ [{completed_count}/{total_count}] Không lấy được thông tin: {url}")
                return product_info
            except Exception as e:
                with lock:
                    completed_count += 1
                    logger.error(f"❌ [{completed_count}/{total_count}] Lỗi xử lý sản phẩm {url}: {str(e)}")
                return None
        
        logger.info(f"🔄 Bắt đầu cào thông tin {total_count} sản phẩm với {self.max_workers} luồng...")
        
        # Tạo danh sách arguments cho ThreadPoolExecutor
        args_list = [(url, i+1) for i, url in enumerate(product_urls)]
        
        # Sử dụng ThreadPoolExecutor để xử lý đa luồng
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = [executor.submit(process_product, args) for args in args_list]
            
            # Chờ tất cả futures hoàn thành
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"❌ Lỗi trong thread: {str(e)}")
        
        logger.info(f"✅ Hoàn thành cào thông tin {len(results)}/{total_count} sản phẩm")
        return results
    
    def convert_images_to_webp(self, product_info_list, output_folder):
        """
        Chuyển đổi tất cả ảnh sản phẩm sang định dạng WebP
        
        Args:
            product_info_list (list): Danh sách thông tin sản phẩm
            output_folder (str): Thư mục lưu ảnh
            
        Returns:
            list: Danh sách thông tin sản phẩm với đường dẫn ảnh WebP
        """
        webp_folder = os.path.join(output_folder, "webp_images")
        os.makedirs(webp_folder, exist_ok=True)
        
        converted_products = []
        total_products = len(product_info_list)
        
        logger.info(f"🖼️ Bắt đầu chuyển đổi {total_products} ảnh sang WebP...")
        
        for i, product_info in enumerate(product_info_list, 1):
            try:
                img_url = product_info.get('Ảnh sản phẩm', '')
                if not img_url:
                    product_info['Ảnh_WebP'] = ''
                    converted_products.append(product_info)
                    logger.info(f"⏭️ Bỏ qua sản phẩm {i}/{total_products}: Không có ảnh")
                    continue
                
                # Tải ảnh từ URL
                response = self.session.get(img_url, timeout=30)
                response.raise_for_status()
                
                # Tạo tên file WebP an toàn
                product_code = product_info.get('Mã sản phẩm', 'unknown')
                # Loại bỏ ký tự đặc biệt khỏi tên file
                safe_product_code = re.sub(r'[<>:"/\\|?*]', '_', product_code)
                webp_filename = f"{safe_product_code}.webp"
                webp_path = os.path.join(webp_folder, webp_filename)
                
                # Kiểm tra xem file đã tồn tại chưa
                if os.path.exists(webp_path):
                    product_info['Ảnh_WebP'] = webp_path
                    converted_products.append(product_info)
                    logger.info(f"✓ Sử dụng ảnh có sẵn {i}/{total_products}: {safe_product_code}")
                    continue
                
                # Chuyển đổi sang WebP sử dụng WebPConverter
                img = Image.open(io.BytesIO(response.content))
                
                # Chuyển đổi sang RGB nếu cần
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                # Lưu ảnh tạm thời để sử dụng WebPConverter
                temp_img_path = os.path.join(webp_folder, f"temp_{safe_product_code}.jpg")
                img.save(temp_img_path, 'JPEG', quality=95)
                
                # Sử dụng WebPConverter để đảm bảo chất lượng
                result = WebPConverter.convert_to_webp(
                    input_path=temp_img_path,
                    output_path=webp_path,
                    quality=90,
                    lossless=False,
                    method=6
                )
                
                # Xóa file tạm
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
                
                if result['success']:
                    # Cập nhật đường dẫn ảnh
                    product_info['Ảnh_WebP'] = webp_path
                    converted_products.append(product_info)
                    logger.info(f"✓ Đã chuyển đổi ảnh {i}/{total_products}: {safe_product_code} ({result['compression_ratio']:.1f}%)")
                else:
                    logger.error(f"❌ Lỗi chuyển đổi WebP cho {safe_product_code}: {result.get('error', 'Unknown error')}")
                    product_info['Ảnh_WebP'] = ''
                    converted_products.append(product_info)
                
            except Exception as e:
                logger.error(f"❌ Lỗi chuyển đổi ảnh cho sản phẩm {product_info.get('Mã sản phẩm', 'unknown')}: {str(e)}")
                product_info['Ảnh_WebP'] = ''
                converted_products.append(product_info)
        
        logger.info(f"✅ Hoàn thành chuyển đổi {len(converted_products)} ảnh sang WebP")
        return converted_products
    
    def create_excel_by_series(self, product_info_list, output_folder):
        """
        Tạo file Excel riêng cho từng series
        
        Args:
            product_info_list (list): Danh sách thông tin sản phẩm
            output_folder (str): Thư mục lưu file Excel
            
        Returns:
            list: Danh sách đường dẫn file Excel đã tạo
        """
        # Nhóm sản phẩm theo series
        series_groups = {}
        
        for product in product_info_list:
            series_name = product.get('Series', 'Unknown_Series')
            if series_name not in series_groups:
                series_groups[series_name] = []
            series_groups[series_name].append(product)
        
        excel_files = []
        total_series = len(series_groups)
        
        logger.info(f"📊 Bắt đầu tạo {total_series} file Excel theo series...")
        
        for i, (series_name, products) in enumerate(series_groups.items(), 1):
            try:
                # Tạo tên file Excel an toàn
                safe_series_name = re.sub(r'[<>:"/\\|?*]', '_', series_name)
                excel_filename = f"BAA_Qlight_{safe_series_name}.xlsx"
                excel_path = os.path.join(output_folder, excel_filename)
                
                # Tạo DataFrame với thứ tự cột hợp lý
                df = pd.DataFrame(products)
                
                # Sắp xếp lại thứ tự cột
                column_order = [
                    'STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Series', 'Giá', 
                    'Tổng quan', 'Ảnh sản phẩm', 'Ảnh_WebP', 'Series_URL', 'URL'
                ]
                
                # Chỉ giữ lại các cột có trong dữ liệu
                available_columns = [col for col in column_order if col in df.columns]
                other_columns = [col for col in df.columns if col not in column_order]
                final_columns = available_columns + other_columns
                
                df = df[final_columns]
                
                # Lưu file Excel
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Products', index=False)
                    
                    # Thêm ảnh vào sheet riêng nếu có
                    workbook = writer.book
                    if 'Ảnh_WebP' in df.columns:
                        self._add_images_to_excel(workbook, products, output_folder)
                    
                    # Thêm thông tin series vào sheet mới
                    self._add_series_info_sheet(workbook, series_name, len(products))
                
                excel_files.append(excel_path)
                logger.info(f"✓ [{i}/{total_series}] Đã tạo file Excel: {excel_filename} ({len(products)} sản phẩm)")
                
            except Exception as e:
                logger.error(f"❌ Lỗi tạo Excel cho series {series_name}: {str(e)}")
        
        logger.info(f"✅ Hoàn thành tạo {len(excel_files)} file Excel")
        return excel_files
    
    def _add_images_to_excel(self, workbook, products, output_folder):
        """
        Thêm ảnh vào file Excel
        
        Args:
            workbook: Workbook object
            products (list): Danh sách sản phẩm
            output_folder (str): Thư mục chứa ảnh
        """
        try:
            # Tạo sheet mới cho ảnh
            if 'Images' in workbook.sheetnames:
                workbook.remove(workbook['Images'])
            
            image_sheet = workbook.create_sheet('Images')
            
            # Thêm tiêu đề
            image_sheet['A1'] = 'Mã sản phẩm'
            image_sheet['B1'] = 'Tên sản phẩm'
            image_sheet['C1'] = 'Ảnh sản phẩm'
            
            # Thêm ảnh cho từng sản phẩm
            for i, product in enumerate(products, start=2):
                image_sheet[f'A{i}'] = product.get('Mã sản phẩm', '')
                image_sheet[f'B{i}'] = product.get('Tên sản phẩm', '')
                
                webp_path = product.get('Ảnh_WebP', '')
                if webp_path and os.path.exists(webp_path):
                    try:
                        # Thêm ảnh vào Excel
                        img = XLImage(webp_path)
                        img.width = 100
                        img.height = 100
                        image_sheet.add_image(img, f'C{i}')
                    except Exception as e:
                        logger.error(f"Lỗi thêm ảnh vào Excel: {str(e)}")
        
        except Exception as e:
            logger.error(f"Lỗi tạo sheet ảnh: {str(e)}")
    
    def _add_series_info_sheet(self, workbook, series_name, product_count):
        """
        Thêm sheet thông tin series vào workbook
        
        Args:
            workbook: Workbook object
            series_name (str): Tên series
            product_count (int): Số lượng sản phẩm trong series
        """
        try:
            # Tạo sheet mới cho thông tin series
            if 'Series_Info' in workbook.sheetnames:
                workbook.remove(workbook['Series_Info'])
            
            info_sheet = workbook.create_sheet('Series_Info')
            
            # Thêm thông tin series
            info_sheet['A1'] = 'Thông tin Series'
            info_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            
            info_sheet['A3'] = 'Tên Series:'
            info_sheet['B3'] = series_name
            info_sheet['A3'].font = openpyxl.styles.Font(bold=True)
            
            info_sheet['A4'] = 'Số sản phẩm:'
            info_sheet['B4'] = product_count
            info_sheet['A4'].font = openpyxl.styles.Font(bold=True)
            
            info_sheet['A5'] = 'Ngày tạo:'
            info_sheet['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            info_sheet['A5'].font = openpyxl.styles.Font(bold=True)
            
            # Điều chỉnh độ rộng cột
            info_sheet.column_dimensions['A'].width = 15
            info_sheet.column_dimensions['B'].width = 30
            
        except Exception as e:
            logger.error(f"Lỗi tạo sheet thông tin series: {str(e)}")
    
    def get_all_series_list(self):
        """
        Lấy danh sách tất cả series từ widget trên trang web
        
        Returns:
            list: Danh sách series với thông tin tên, URL và số lượng sản phẩm
        """
        try:
            logger.info("🔍 Đang lấy danh sách tất cả series...")
            
            # Lấy HTML từ trang chính
            html = get_html_content(self.base_url)
            if not html:
                logger.error("❌ Không thể tải trang chính")
                return []
            
            soup = BeautifulSoup(html, 'html.parser')
            
            # Tìm widget series theo cấu trúc HTML đã cung cấp
            series_widget = soup.select_one('div.widget ul.widget_Items[data-id="Seri"]')
            if not series_widget:
                logger.warning("⚠️ Không tìm thấy widget series")
                return []
            
            series_list = []
            
            # Tìm tất cả các item series
            series_items = series_widget.select('li.widget_Item[data-input="Seri"]')
            
            for item in series_items:
                try:
                    # Lấy link series
                    link = item.select_one('a')
                    if not link:
                        continue
                    
                    href = link.get('href', '')
                    if not href:
                        continue
                    
                    # Lấy tên series
                    label = item.select_one('span.itemLabel')
                    series_name = label.get_text(strip=True) if label else ''
                    
                    # Lấy số lượng sản phẩm
                    count_span = item.select_one('span.count')
                    count_text = count_span.get_text(strip=True) if count_span else '[0]'
                    # Trích xuất số từ text [123]
                    count_match = re.search(r'\[(\d+)\]', count_text)
                    product_count = int(count_match.group(1)) if count_match else 0
                    
                    # Tạo URL đầy đủ
                    full_url = urljoin(self.base_url, href)
                    
                    series_info = {
                        'name': series_name,
                        'url': full_url,
                        'product_count': product_count,
                        'original_href': href
                    }
                    
                    series_list.append(series_info)
                    
                except Exception as e:
                    logger.error(f"❌ Lỗi xử lý series item: {str(e)}")
                    continue
            
            # Sắp xếp theo số lượng sản phẩm giảm dần
            series_list.sort(key=lambda x: x['product_count'], reverse=True)
            
            logger.info(f"✅ Tìm thấy {len(series_list)} series")
            return series_list
            
        except Exception as e:
            logger.error(f"❌ Lỗi khi lấy danh sách series: {str(e)}")
            return []
    
    def crawl_specific_series(self, series_url, series_name, output_folder=None):
        """
        Cào dữ liệu cho một series cụ thể
        
        Args:
            series_url (str): URL của series
            series_name (str): Tên series
            output_folder (str): Thư mục lưu kết quả
            
        Returns:
            dict: Kết quả cào dữ liệu cho series
        """
        if not output_folder:
            output_folder = os.path.join(os.getcwd(), f"output_baa_qlight_{series_name}")
        
        os.makedirs(output_folder, exist_ok=True)
        
        start_time = time.time()
        logger.info(f"🚀 Bắt đầu cào dữ liệu series: {series_name}")
        
        try:
            # Bước 1: Lấy danh sách sản phẩm từ series
            logger.info(f"📋 Đang lấy danh sách sản phẩm từ series {series_name}...")
            product_urls = self.get_products_from_series(series_url)
            logger.info(f"✓ Tìm thấy {len(product_urls)} sản phẩm từ series {series_name}")
            
            if not product_urls:
                return {
                    'success': False,
                    'error': 'Không tìm thấy sản phẩm nào trong series',
                    'message': f"Series {series_name} không có sản phẩm"
                }
            
            # Bước 2: Cào thông tin sản phẩm đa luồng
            logger.info(f"🔄 Đang cào thông tin sản phẩm từ series {series_name} (đa luồng)...")
            product_info_list = self.crawl_products_multithread(product_urls)
            logger.info(f"✓ Đã cào thông tin {len(product_info_list)} sản phẩm")
            
            # Bước 3: Chuyển đổi ảnh sang WebP
            logger.info(f"🖼️ Đang chuyển đổi ảnh sang WebP cho series {series_name}...")
            product_info_list = self.convert_images_to_webp(product_info_list, output_folder)
            logger.info("✓ Hoàn thành chuyển đổi ảnh WebP")
            
            # Bước 4: Tạo file Excel cho series
            logger.info(f"📊 Đang tạo file Excel cho series {series_name}...")
            excel_files = self.create_excel_by_series(product_info_list, output_folder)
            logger.info(f"✓ Đã tạo {len(excel_files)} file Excel")
            
            # Bước 5: Tạo file tổng hợp cho series
            logger.info(f"📋 Đang tạo file tổng hợp cho series {series_name}...")
            summary_df = pd.DataFrame(product_info_list)
            summary_path = os.path.join(output_folder, f"BAA_Qlight_{series_name}_Summary.xlsx")
            summary_df.to_excel(summary_path, index=False)
            
            end_time = time.time()
            duration = end_time - start_time
            
            result = {
                'success': True,
                'series_name': series_name,
                'series_url': series_url,
                'total_products': len(product_info_list),
                'excel_files': excel_files,
                'summary_file': summary_path,
                'output_folder': output_folder,
                'duration': duration,
                'message': f"Hoàn thành cào dữ liệu {len(product_info_list)} sản phẩm từ series {series_name} trong {duration:.2f} giây"
            }
            
            logger.info(f"✅ {result['message']}")
            return result
            
        except Exception as e:
            logger.error(f"❌ Lỗi trong quá trình cào dữ liệu series {series_name}: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'message': f"Lỗi cào series {series_name}: {str(e)}"
            }
    
    def crawl_baa_qlight(self, output_folder=None):
        """
        Hàm chính để cào dữ liệu BAA Qlight (tất cả series)
        
        Args:
            output_folder (str): Thư mục lưu kết quả
            
        Returns:
            dict: Kết quả cào dữ liệu
        """
        if not output_folder:
            output_folder = os.path.join(os.getcwd(), "output_baa_qlight")
        
        os.makedirs(output_folder, exist_ok=True)
        
        start_time = time.time()
        logger.info("🚀 Bắt đầu cào dữ liệu BAA Qlight...")
        
        try:
            # Bước 1: Lấy danh sách sản phẩm từ trang chính
            logger.info("📋 Đang lấy danh sách sản phẩm từ trang chính...")
            product_urls = extract_product_urls(self.base_url)
            logger.info(f"✓ Tìm thấy {len(product_urls)} sản phẩm từ trang chính")
            
            # Bước 2: Thu thập thông tin series từ các sản phẩm
            logger.info("🔍 Đang thu thập thông tin series...")
            series_info = {}
            series_products = {}
            
            # Lấy thông tin series từ một số sản phẩm đầu tiên để tìm các series
            sample_products = product_urls[:min(10, len(product_urls))]
            for url in sample_products:
                series_data = self.extract_series_info(url)
                if series_data and series_data['name'] not in series_info:
                    series_info[series_data['name']] = series_data
                    logger.info(f"✓ Tìm thấy series: {series_data['name']}")
            
            # Bước 3: Lấy sản phẩm từ từng series
            logger.info("📦 Đang lấy sản phẩm từ các series...")
            all_product_urls = set(product_urls)  # Bắt đầu với sản phẩm từ trang chính
            
            for series_name, series_data in series_info.items():
                series_products_list = self.get_products_from_series(series_data['url'])
                series_products[series_name] = series_products_list
                all_product_urls.update(series_products_list)
                logger.info(f"✓ Series '{series_name}': {len(series_products_list)} sản phẩm")
            
            # Chuyển về list và loại bỏ trùng lặp
            final_product_urls = list(all_product_urls)
            logger.info(f"✓ Tổng cộng {len(final_product_urls)} sản phẩm unique")
            
            # Bước 4: Cào thông tin sản phẩm đa luồng
            logger.info("🔄 Đang cào thông tin sản phẩm (đa luồng)...")
            product_info_list = self.crawl_products_multithread(final_product_urls)
            logger.info(f"✓ Đã cào thông tin {len(product_info_list)} sản phẩm")
            
            # Bước 5: Chuyển đổi ảnh sang WebP
            logger.info("🖼️ Đang chuyển đổi ảnh sang WebP...")
            product_info_list = self.convert_images_to_webp(product_info_list, output_folder)
            logger.info("✓ Hoàn thành chuyển đổi ảnh WebP")
            
            # Bước 6: Tạo file Excel theo series
            logger.info("📊 Đang tạo file Excel theo series...")
            excel_files = self.create_excel_by_series(product_info_list, output_folder)
            logger.info(f"✓ Đã tạo {len(excel_files)} file Excel")
            
            # Bước 7: Tạo file tổng hợp
            logger.info("📋 Đang tạo file tổng hợp...")
            summary_df = pd.DataFrame(product_info_list)
            summary_path = os.path.join(output_folder, "BAA_Qlight_Summary.xlsx")
            summary_df.to_excel(summary_path, index=False)
            
            # Bước 8: Tạo báo cáo series
            logger.info("📈 Đang tạo báo cáo series...")
            series_report = []
            for series_name, products in series_products.items():
                series_report.append({
                    'Series': series_name,
                    'Số sản phẩm': len(products),
                    'URL Series': series_info.get(series_name, {}).get('url', '')
                })
            
            series_df = pd.DataFrame(series_report)
            series_report_path = os.path.join(output_folder, "BAA_Qlight_Series_Report.xlsx")
            series_df.to_excel(series_report_path, index=False)
            
            end_time = time.time()
            duration = end_time - start_time
            
            result = {
                'success': True,
                'total_products': len(product_info_list),
                'total_series': len(series_info),
                'excel_files': excel_files,
                'summary_file': summary_path,
                'series_report': series_report_path,
                'output_folder': output_folder,
                'duration': duration,
                'series_info': series_info,
                'message': f"Hoàn thành cào dữ liệu {len(product_info_list)} sản phẩm từ {len(series_info)} series trong {duration:.2f} giây"
            }
            
            logger.info(f"✅ {result['message']}")
            return result
            
        except Exception as e:
            logger.error(f"❌ Lỗi trong quá trình cào dữ liệu: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'message': f"Lỗi: {str(e)}"
            }

# Hàm tiện ích để sử dụng từ bên ngoài
def crawl_baa_qlight(output_folder=None, max_workers=10):
    """
    Hàm tiện ích để cào dữ liệu BAA Qlight (tất cả series)
    
    Args:
        output_folder (str): Thư mục lưu kết quả
        max_workers (int): Số luồng tối đa
        
    Returns:
        dict: Kết quả cào dữ liệu
    """
    crawler = BAAQlightCrawler(max_workers=max_workers)
    return crawler.crawl_baa_qlight(output_folder)

def get_all_series_list(max_workers=5):
    """
    Hàm tiện ích để lấy danh sách tất cả series
    
    Args:
        max_workers (int): Số luồng tối đa
        
    Returns:
        list: Danh sách series với thông tin tên, URL và số lượng sản phẩm
    """
    crawler = BAAQlightCrawler(max_workers=max_workers)
    return crawler.get_all_series_list()

def crawl_specific_series(series_url, series_name, output_folder=None, max_workers=10):
    """
    Hàm tiện ích để cào dữ liệu cho một series cụ thể
    
    Args:
        series_url (str): URL của series
        series_name (str): Tên series
        output_folder (str): Thư mục lưu kết quả
        max_workers (int): Số luồng tối đa
        
    Returns:
        dict: Kết quả cào dữ liệu cho series
    """
    crawler = BAAQlightCrawler(max_workers=max_workers)
    return crawler.crawl_specific_series(series_url, series_name, output_folder)

def test_series_extraction():
    """
    Hàm test để kiểm tra việc trích xuất series
    """
    crawler = BAAQlightCrawler()
    test_url = "https://baa.vn/vn/den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779/"
    
    print("🔍 Testing series extraction...")
    series_info = crawler.extract_series_info(test_url)
    if series_info:
        print(f"✓ Tìm thấy series: {series_info['name']}")
        print(f"  URL: {series_info['url']}")
    else:
        print("❌ Không tìm thấy thông tin series")
    
    return series_info

def test_product_extraction():
    """
    Hàm test để kiểm tra việc trích xuất sản phẩm
    """
    crawler = BAAQlightCrawler()
    test_url = "https://baa.vn/vn/den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779/"
    
    print("🔍 Testing product extraction...")
    product_info = crawler.crawl_product_info(test_url, 1)
    if product_info:
        print(f"✓ Tìm thấy sản phẩm: {product_info.get('Tên sản phẩm', 'Unknown')}")
        print(f"  Mã: {product_info.get('Mã sản phẩm', 'Unknown')}")
        print(f"  Series: {product_info.get('Series', 'Unknown')}")
    else:
        print("❌ Không tìm thấy thông tin sản phẩm")
    
    return product_info

if __name__ == "__main__":
    print("🚀 BAA Qlight Crawler - Test Mode")
    print("=" * 50)
    
    # Test 1: Trích xuất series
    print("\n1. Testing Series Extraction:")
    series_result = test_series_extraction()
    
    # Test 2: Trích xuất sản phẩm
    print("\n2. Testing Product Extraction:")
    product_result = test_product_extraction()
    
    # Test 3: Chạy crawler đầy đủ (tùy chọn)
    print("\n3. Full Crawler Test (tùy chọn):")
    response = input("Bạn có muốn chạy crawler đầy đủ không? (y/n): ")
    if response.lower() == 'y':
        print("🔄 Bắt đầu crawler đầy đủ...")
        result = crawl_baa_qlight(max_workers=5)  # Giảm số luồng để test
        print(f"✅ Kết quả: {result}")
    else:
        print("⏭️ Bỏ qua test crawler đầy đủ")
    
    print("\n✅ Test hoàn thành!")
