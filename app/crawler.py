import requests
from bs4 import BeautifulSoup
import pandas as pd
import tempfile
import re
import time
import random
import os
from urllib.parse import urljoin, urlparse, quote
import logging
from app import socketio
from datetime import datetime
from PIL import Image
import io
import traceback
from werkzeug.utils import secure_filename
import openpyxl
import hashlib
import threading
from urllib.parse import parse_qs
from io import BytesIO
import json
from openpyxl import Workbook
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed

# Headers giả lập trình duyệt để tránh bị chặn
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Connection': 'keep-alive',
    'Referer': 'https://google.com'
}

def get_html_content(url, headers=None):
    """
    Tải nội dung HTML từ URL
    
    Args:
        url (str): URL cần tải nội dung
        headers (dict, optional): Headers cho request
        
    Returns:
        str: Nội dung HTML hoặc None nếu có lỗi
    """
    if headers is None:
        headers = HEADERS
    
    try:
        # Tắt cảnh báo SSL không an toàn
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        response = requests.get(url, headers=headers, timeout=30, verify=False)
        response.raise_for_status()
        return response.text
    except Exception as e:
        print(f"Lỗi khi tải nội dung từ {url}: {e}")
        return None

def is_product_url(url):
    """Kiểm tra xem URL có phải là URL sản phẩm hợp lệ không"""
    # Phân tích URL
    parsed_url = urlparse(url)
    path = parsed_url.path.lower()
    
    # Loại bỏ các URL không phải BAA.vn
    if 'baa.vn' not in parsed_url.netloc.lower():
        return False
    
    # Loại bỏ các URL tin tức, thông tin, v.v. trước tiên
    excluded_paths = ['/tin-tuc/', '/news/', '/thong-tin/', '/information/', '/category/', '/danh-muc/', '/page/', '/search/', '/tim-kiem/', '/about/', '/contact/', '/blog/', '/home/']
    for excluded in excluded_paths:
        if excluded in path:
            return False
        
    # Pattern 1: URL có chứa /san-pham/
    if '/san-pham/' in path:
        # Mẫu URL: baa.vn/vn/san-pham/bo-dieu-khien-nhiet-do-tuong-tu-autonics-tom-f3rj4c_61459
        if '/san-pham/' in path:
            parts = path.split('/san-pham/')
            if len(parts) > 1 and parts[1].strip():
                # Kiểm tra phần sau /san-pham/ có chứa nội dung hợp lệ
                product_part = parts[1].strip('/')
                if product_part and len(product_part) > 3:
                    print(f"  ✓ URL sản phẩm hợp lệ (pattern /san-pham/): {url}")
                    return True
    
    # Pattern 2: URL có chứa /product/
    if '/product/' in path:
        parts = path.split('/product/')
        if len(parts) > 1 and parts[1].strip():
            product_part = parts[1].strip('/')
            if product_part and len(product_part) > 3:
                print(f"  ✓ URL sản phẩm hợp lệ (pattern /product/): {url}")
                return True
    
    # Pattern 3: URL dạng /vn/ten-san-pham_id (pattern đặc biệt của BAA.vn)
    if re.search(r'/vn/[^/]+_\d+/?$', url, re.IGNORECASE):
        # Kiểm tra thêm để đảm bảo không phải danh mục
        if not any(category_keyword in path for category_keyword in ['/category/', '/danh-muc/', '/list/']):
            print(f"  ✓ URL sản phẩm hợp lệ (pattern /vn/name_id): {url}")
            return True
    
    # Pattern 4: URL có chứa mã sản phẩm dạng chữ-số_id ở cuối
    if re.search(r'/[^/]+-[^/]+_\d+/?$', url, re.IGNORECASE):
        print(f"  ✓ URL sản phẩm hợp lệ (pattern name-name_id): {url}")
        return True
    
    # Pattern 5: URL có chứa từ khóa sản phẩm và có ID số ở cuối
    product_keywords = ['san-pham', 'product', 'item', 'detail']
    for keyword in product_keywords:
        if keyword in path and re.search(r'_\d+/?$', url):
            print(f"  ✓ URL sản phẩm hợp lệ (pattern keyword + _id): {url}")
            return True
    
    # Pattern 6: URL BAA.vn với cấu trúc đặc biệt
    # Ví dụ: https://baa.vn/vn/den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779/
    if 'baa.vn' in parsed_url.netloc.lower():
        # Kiểm tra URL có dạng /vn/ten-dai-co-dash_id/
        if re.search(r'/vn/[a-zA-Z0-9\-]{10,}_\d+/?$', url, re.IGNORECASE):
            print(f"  ✓ URL sản phẩm BAA.vn hợp lệ (long name pattern): {url}")
            return True
        
        # Kiểm tra URL có chứa series hoặc model trong tên
        if re.search(r'/vn/[^/]*(series|model|type)[^/]*_\d+/?$', url, re.IGNORECASE):
            print(f"  ✓ URL sản phẩm BAA.vn hợp lệ (series/model pattern): {url}")
            return True
    
    # Debug: Log URL không được nhận diện
    print(f"  ✗ URL không được nhận diện là sản phẩm: {url}")
    return False

def is_category_url(url):
    """
    Kiểm tra xem URL có phải là URL danh mục hay không
    """
    # Phân tích URL
    parsed_url = urlparse(url)
    path = parsed_url.path.lower()
    
    # Loại bỏ các URL không phải BAA.vn
    if 'baa.vn' not in parsed_url.netloc.lower():
        return False
    
    # Nhận diện các URL dạng /vn/ten-danh-muc_xxxx/ (trường hợp đặc biệt)
    if re.search(r'/vn/[^/]+_\d+/?$', url, re.IGNORECASE):
        # Kiểm tra thêm để đảm bảo không phải sản phẩm
        if not any(product_keyword in path for product_keyword in ['/san-pham/', '/product/']):
            print(f"  ✓ URL danh mục hợp lệ (pattern /vn/name_id): {url}")
        return True

    # Xử lý riêng cho URL đèn tháp LED
    led_tower_url = "den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779"
    if led_tower_url in url:
        print(f"  ✓ Đã phát hiện URL đèn tháp LED: {url}")
        return True
    
    # Các mẫu regex để phát hiện URL danh mục
    baa_category_patterns = [
        r'/category/',               # URL có /category/
        r'/danh-muc/',              # URL có /danh-muc/
        r'/Category/',              # URL có /Category/ (viết hoa)
        r'baa\.vn.*\/vn\/[^/]+\/.*-page-\d+',  # URL có phân trang
        r'haiphongtech.+\/danh-muc', # URL haiphongtech với danh-muc
        r'/vn/.*_\d+/?$'            # Pattern chung cho danh mục BAA.vn
    ]
    
    for pattern in baa_category_patterns:
        if re.search(pattern, url, re.IGNORECASE):
            # Kiểm tra thêm nếu là url từ BAA.vn
            if 'baa.vn' in url:
                # Đảm bảo không phải URL sản phẩm
                if not any(product_keyword in path for product_keyword in ['/san-pham/', '/product/']):
                    print(f"  ✓ URL danh mục hợp lệ (pattern match): {url}")
                    return True
            else:
                print(f"  ✓ URL danh mục hợp lệ (other site): {url}")
                return True
    
    # Kiểm tra thêm các URL có thể là danh mục dựa trên query parameters
    if 'page=' in url or '/page/' in url:
        if 'baa.vn' in url and not any(product_keyword in path for product_keyword in ['/san-pham/', '/product/']):
            print(f"  ✓ URL danh mục hợp lệ (có phân trang): {url}")
            return True
    
    return False

def extract_category_links(category_urls):
    """
    Trích xuất liên kết sản phẩm từ URL danh mục với đa luồng
    """
    all_product_urls = []
    processed_urls = set()
    
    # Giới hạn số luồng để tránh quá tải
    max_workers = 10
    
    # Tạo hàm xử lý riêng để sử dụng với ThreadPoolExecutor
    def process_category_url(url):
        local_product_urls = []
        try:
            # Kiểm tra đã xử lý URL này chưa
            if url in processed_urls:
                print(f"Bỏ qua URL đã xử lý: {url}")
                return []
                
            # Đánh dấu đã xử lý
            processed_urls.add(url)
            
            # Kiểm tra URL đèn tháp LED đặc biệt
            is_led_page = False
            led_tower_url = "den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779"
            if led_tower_url in url:
                is_led_page = True
                print(f"Xử lý URL đèn tháp LED đặc biệt: {url}")
                
                # Lấy HTML của trang
                html = get_html_content(url)
                if not html:
                    print(f"Không thể lấy nội dung từ {url}")
                    return []
                
                # Phân tích HTML
                soup = BeautifulSoup(html, 'html.parser')
                
                print(f"Đang xử lý trang đèn tháp LED: {url}")
                
                # Tìm tất cả các card sản phẩm
                product_cards = soup.select('a.card.product__card')
                if product_cards:
                    print(f"Tìm thấy {len(product_cards)} thẻ a.card.product__card")
                    
                    # Lấy href từ các thẻ a
                    for card in product_cards:
                        href = card.get('href')
                        if href:
                            # Đảm bảo URL đầy đủ
                            if not href.startswith('http'):
                                parsed_url = urlparse(url)
                                full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}"
                            else:
                                full_url = href
                                
                            # Thêm vào danh sách nếu chưa có
                            if full_url not in local_product_urls:
                                local_product_urls.append(full_url)
                                print(f"Đã thêm URL sản phẩm: {full_url}")
                else:
                    print("Không tìm thấy thẻ a.card.product__card, thử các selector khác")
                    
                    # Tìm các liên kết sản phẩm khác nếu không tìm thấy a.card.product__card
                    product_links = soup.select('a.product-item-link, a.product__card, a.product_item, a.product-item')
                    if product_links:
                        print(f"Tìm thấy {len(product_links)} liên kết sản phẩm thay thế")
                        
                        for link in product_links:
                            href = link.get('href')
                            if href:
                                # Đảm bảo URL đầy đủ
                                if not href.startswith('http'):
                                    if href.startswith('/'):
                                        parsed_url = urlparse(url)
                                        full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}"
                                    else:
                                        full_url = url.rstrip('/') + '/' + href
                                else:
                                    full_url = href
                                
                                # Thêm vào danh sách nếu chưa có
                                if full_url not in local_product_urls and '/san-pham/' in full_url:
                                    local_product_urls.append(full_url)
                                    print(f"Đã thêm URL sản phẩm thay thế: {full_url}")
                
                # Kiểm tra tìm thấy sản phẩm hoặc thông báo lỗi
                if len(local_product_urls) == 0:
                    print("CẢNH BÁO: Không tìm thấy sản phẩm nào trên trang đèn tháp LED!")
                    
                    # Các thẻ trực tiếp chứa sản phẩm (debug)
                    debug_products = soup.select('.product__card')
                    print(f"DEBUG: Tìm thấy {len(debug_products)} thẻ .product__card")
                    
                    for dp in debug_products:
                        if dp.name == 'a' and dp.has_attr('href'):
                            print(f"DEBUG: Liên kết: {dp.get('href')}")
                
                print(f"Đã tìm thấy {len(local_product_urls)} URL sản phẩm từ trang đèn tháp LED.")
                
                # Tìm liên kết phân trang để xử lý các trang tiếp theo
                pagination_links = soup.select('ul.pagination li a')
                if not pagination_links:
                    pagination_links = soup.select('.pages a, .pagination a, a[href*="page="], a[href*="/page/"]')
                
                new_category_urls = []
                for page_link in pagination_links:
                    page_text = page_link.get_text(strip=True)
                    page_url = page_link.get('href')
                    
                    # Bỏ qua các liên kết không phải số trang
                    if not re.match(r'^\d+$', page_text):
                        continue
                        
                    # Bỏ qua liên kết không có href hoặc liên kết không phải số trang
                    if not page_url or not re.match(r'^\d+$', page_text):
                        continue
                        
                    # Đảm bảo URL đầy đủ
                    if not page_url.startswith('http'):
                        if page_url.startswith('/'):
                            parsed_url = urlparse(url)
                            page_url = f"{parsed_url.scheme}://{parsed_url.netloc}{page_url}"
                        else:
                            page_url = url.rstrip('/') + '/' + page_url
                            
                    # Bỏ qua trang hiện tại
                    if page_url == url:
                        continue
                    
                    # Thêm vào danh sách URL cần xử lý nếu chưa xử lý
                    if page_url not in processed_urls:
                        print(f"Thêm trang phân trang: {page_url}")
                        new_category_urls.append(page_url)
                
                return local_product_urls, new_category_urls
            else:
                # Xử lý các URL danh mục thông thường
                product_urls = extract_product_urls(url)
                return product_urls, []
        
        except Exception as e:
            print(f"Lỗi khi xử lý URL danh mục {url}: {str(e)}")
            print(traceback.format_exc())
            return [], []
    
    # Sử dụng ThreadPoolExecutor để xử lý đa luồng
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Danh sách các URL cần xử lý (bắt đầu với các URL danh mục ban đầu)
        urls_to_process = list(category_urls)
        
        while urls_to_process:
            # Lấy một batch URLs để xử lý
            batch = urls_to_process[:max_workers]
            urls_to_process = urls_to_process[max_workers:]
            
            # Tạo các futures cho batch hiện tại
            futures = {executor.submit(process_category_url, url): url for url in batch}
            
            # Xử lý kết quả khi hoàn thành
            for future in as_completed(futures):
                url = futures[future]
                try:
                    result = future.result()
                    if result:
                        product_urls, new_category_urls = result
                        
                        # Thêm các URL sản phẩm vào danh sách kết quả
                        for product_url in product_urls:
                            if product_url not in all_product_urls:
                                all_product_urls.append(product_url)
                        
                        # Thêm các URL danh mục mới vào danh sách cần xử lý
                        for new_url in new_category_urls:
                            if new_url not in processed_urls and new_url not in urls_to_process:
                                urls_to_process.append(new_url)
                except Exception as e:
                    print(f"Lỗi khi xử lý future cho URL {url}: {str(e)}")
    
    print(f"Tổng cộng tìm thấy {len(all_product_urls)} liên kết sản phẩm độc nhất")
    return all_product_urls

def extract_product_info(url, required_fields=None, index=1):
    """
    Trích xuất thông tin sản phẩm từ URL BAA.vn với selector thực tế
    """
    if not required_fields:
        required_fields = ['STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Giá', 'Tổng quan', 'Ảnh sản phẩm', 'URL']
    print(f"[DEBUG] Đang trích xuất thông tin từ {url}")
    max_retries = 3
    current_retry = 0
    while current_retry < max_retries:
        try:
            response = requests.get(url, headers=HEADERS, timeout=15)
            response.raise_for_status()
            html_content = response.text
            soup = BeautifulSoup(html_content, 'html.parser')
            product_info = {
                'STT': index,
                'URL': url
            }
            # Tên sản phẩm
            name_element = soup.select_one('h1.product__name')
            if name_element:
                product_info['Tên sản phẩm'] = name_element.text.strip()
            else:
                product_info['Tên sản phẩm'] = ''
            # Mã sản phẩm
            code_element = soup.select_one('span.product__symbol__value')
            if code_element:
                product_info['Mã sản phẩm'] = code_element.text.strip()
            else:
                product_info['Mã sản phẩm'] = ''
            # Giá
            product_info['Giá'] = extract_baa_product_price(soup, product_info.get('Tên sản phẩm', ''))
            # Thông số kỹ thuật
            spec_html = ''
            spec_table = soup.select_one('table.feature__metadata--tab.active')
            if spec_table:
                spec_html = str(spec_table)
            product_info['Tổng quan'] = spec_html
            # Ảnh sản phẩm
            img_url = ''
            modal_img = soup.select_one('div.modal-body-image.active img')
            if modal_img and modal_img.get('src'):
                img_url = modal_img['src']
            product_info['Ảnh sản phẩm'] = img_url
            # Sắp xếp lại thứ tự trường
            filtered_info = {}
            for field in required_fields:
                filtered_info[field] = product_info.get(field, '')
            return filtered_info
        except Exception as e:
            current_retry += 1
            print(f"[DEBUG] Lỗi khi xử lý {url} (lần {current_retry}): {str(e)}")
            if current_retry < max_retries:
                time.sleep(2)
            else:
                return {field: '' for field in required_fields}

def extract_full_value(value_cell):
    """
    Trích xuất giá trị đầy đủ từ một ô td, bao gồm cả phần moreellipses và morecontent
    """
    # Kiểm tra cấu trúc HTML ban đầu
    html_content = str(value_cell)
    
    # Danh sách các trường đặc biệt cần xử lý riêng
    special_fields = [
        "Phụ kiện", "Module", "Terminal", "DeviceNet", "CANopen", "Fieldbus", 
        "mounting", "Chế độ điều khiển", "Khả năng bảo vệ", "Ứng dụng", 
        "Kết nối", "Chứng nhận", "Tiêu chuẩn", "Cấp bảo vệ"
    ]
    
    # Kiểm tra xem đây có phải là trường đặc biệt không
    is_special_field = False
    for field in special_fields:
        if field in html_content:
            is_special_field = True
            break
    
    # Kiểm tra nếu có cả phần hiển thị ngắn và phần mở rộng
    has_more = value_cell.select_one('.moreellipses') and value_cell.select_one('.morecontent')
    
    if has_more:
        # Phương pháp 1: Truy cập trực tiếp vào các phần tử DOM
        visible_text = ""
        morecontent_text = ""
        
        # Lấy trực tiếp text từ span đầu tiên (phần hiển thị)
        first_span = value_cell.select_one('div > span')
        if first_span:
            visible_text = first_span.get_text(strip=True)
            # Xóa bỏ nội dung của moreellipses trong phần hiển thị nếu có
            visible_text = re.sub(r'\[\.\.\.\]', '', visible_text)
            visible_text = visible_text.strip()
        
        # Lấy nội dung trong morecontent
        morecontent = value_cell.select_one('.morecontent')
        if morecontent:
            # Lấy nội dung từ span đầu tiên trong morecontent
            morecontent_span = morecontent.select_one('span')
            if morecontent_span:
                morecontent_text = morecontent_span.get_text(strip=True)
        
        # Nếu là trường đặc biệt (ứng dụng, chế độ điều khiển, khả năng bảo vệ, v.v.)
        if is_special_field:
            # Xử lý các trường hợp đặc biệt
            if visible_text and morecontent_text:
                # Phân tách thành các phần riêng biệt, xử lý cả dấu phẩy và cả dấu chấm
                visible_parts = [part.strip() for part in re.split(r',|\.\s+', visible_text) if part.strip()]
                morecontent_parts = [part.strip() for part in re.split(r',|\.\s+', morecontent_text) if part.strip()]
                
                # Tạo danh sách các phần không trùng lặp
                combined_parts = []
                
                # Thêm các phần từ visible_text
                for part in visible_parts:
                    if part and part not in combined_parts:
                        # Kiểm tra xem phần này có chứa chuỗi con "mode" bị lặp không
                        if "mode" in part.lower():
                            # Loại bỏ trường hợp "mal voltage mode" hoặc tương tự
                            if re.search(r'^\w{1,4}\s+\w+\s+mode$', part.lower()):
                                continue
                        combined_parts.append(part)
                
                # Thêm các phần từ morecontent_text mà không trùng lặp
                for part in morecontent_parts:
                    # Kiểm tra xem phần này có phải là một phần của phần nào trong combined_parts không
                    is_duplicate = False
                    
                    # Kiểm tra phần text có bị trùng lặp một phần với các phần khác không
                    for existing_part in combined_parts:
                        # Kiểm tra các trường hợp trùng lặp
                        if (part in existing_part or existing_part in part or 
                            # Kiểm tra xem có chung từ khóa không (với "protection", "mode")
                            (("protection" in part.lower() and "protection" in existing_part.lower()) and 
                             (part.lower().split()[0] == existing_part.lower().split()[0])) or
                            (("mode" in part.lower() and "mode" in existing_part.lower()) and 
                             any(w in existing_part.lower() for w in part.lower().split() if w != "mode"))):
                            is_duplicate = True
                            break
                    
                    # Kiểm tra trường hợp đặc biệt
                    if "mode" in part.lower() and re.search(r'^\w{1,4}\s+\w+\s+mode$', part.lower()):
                        is_duplicate = True
                    
                    # Trường hợp "ply" trong "supply"
                    if len(part) <= 3 and any(part.lower() in existing.lower() for existing in combined_parts):
                        is_duplicate = True
                    
                    if not is_duplicate and part and part not in combined_parts:
                        combined_parts.append(part)
                
                # Nối lại thành một chuỗi hoàn chỉnh
                full_value = ", ".join(combined_parts)
            else:
                # Nếu một trong hai phần trống, lấy phần còn lại
                full_value = visible_text if visible_text else morecontent_text
        else:
            # Các trường hợp thông thường
            if visible_text and morecontent_text:
                # Nếu phần hiển thị đã có dấu phẩy ở cuối
                if visible_text.endswith(','):
                    full_value = visible_text + ' ' + morecontent_text
                else:
                    # Kiểm tra xem phần đầu có nằm trong phần sau không
                    if visible_text in morecontent_text:
                        full_value = morecontent_text
                    else:
                        # Kiểm tra xem có phần trùng lặp không
                        combined_parts = []
                        visible_parts = [part.strip() for part in visible_text.split(',')]
                        morecontent_parts = [part.strip() for part in morecontent_text.split(',')]
                        
                        # Thêm các phần từ visible_text
                        for part in visible_parts:
                            if part and part not in combined_parts:
                                combined_parts.append(part)
                        
                        # Thêm các phần từ morecontent_text
                        for part in morecontent_parts:
                            is_duplicate = False
                            for existing_part in combined_parts:
                                if part in existing_part or existing_part in part:
                                    is_duplicate = True
                                    break
                            
                            if not is_duplicate and part and part not in combined_parts:
                                combined_parts.append(part)
                        
                        full_value = ", ".join(combined_parts)
            else:
                full_value = visible_text if visible_text else morecontent_text
        
        # Làm sạch kết quả
        full_value = re.sub(r'\[\.\.\.\]', '', full_value)  # Loại bỏ [...]
        full_value = re.sub(r'Hiển thị (thêm|bớt)', '', full_value)  # Loại bỏ "Hiển thị thêm/bớt"
        full_value = re.sub(r'\s+', ' ', full_value).strip()  # Chuẩn hóa khoảng trắng
        
        # Loại bỏ trùng lặp một lần nữa - so sánh các phần phân tách bởi dấu phẩy
        parts = [part.strip() for part in full_value.split(',')]
        unique_parts = []
        for part in parts:
            if part and part not in unique_parts and not any(
                part in unique and len(part) < len(unique) for unique in unique_parts
            ):
                # Kiểm tra đặc biệt cho các trường có từ khóa "mode" hoặc "protection"
                if "mode" in part.lower() or "protection" in part.lower():
                    duplicate = False
                    for existing in unique_parts:
                        if (("mode" in existing.lower() and "mode" in part.lower()) and
                            any(word in existing.lower() for word in part.lower().split() if word != "mode")):
                            duplicate = True
                            break
                        if (("protection" in existing.lower() and "protection" in part.lower()) and
                            any(word in existing.lower() for word in part.lower().split() if word != "protection")):
                            duplicate = True
                            break
                    
                    if not duplicate:
                        unique_parts.append(part)
                else:
                    unique_parts.append(part)
        
        full_value = ", ".join(unique_parts)
        
        # Nếu đã tìm được giá trị đầy đủ, trả về kết quả
        if full_value:
            return full_value
    
    # Phương pháp 2: Phân tích trực tiếp HTML nếu phương pháp trên không thành công
    if is_special_field and has_more:
        try:
            # Phân tích HTML bằng cách xác định các phần tử con
            # Regex để tìm nội dung trong span đầu tiên và morecontent
            visible_match = re.search(r'<span[^>]*>(.*?)<span class="moreellipses"', html_content, re.DOTALL)
            morecontent_match = re.search(r'<span class="morecontent"><span[^>]*>(.*?)</span>', html_content, re.DOTALL)
            
            visible_text = ""
            morecontent_text = ""
            
            # Lấy nội dung từ các phần đã tìm thấy
            if visible_match:
                visible_text = BeautifulSoup(visible_match.group(1), 'html.parser').get_text(strip=True)
            
            if morecontent_match:
                morecontent_text = BeautifulSoup(morecontent_match.group(1), 'html.parser').get_text(strip=True)
            
            # Kết hợp nội dung và loại bỏ trùng lặp
            if visible_text and morecontent_text:
                # Phân tách thành các phần
                visible_parts = [part.strip() for part in re.split(r',|\.\s+', visible_text) if part.strip()]
                morecontent_parts = [part.strip() for part in re.split(r',|\.\s+', morecontent_text) if part.strip()]
                
                # Tạo danh sách các phần không trùng lặp
                combined_parts = []
                
                # Thêm các phần từ visible_text
                for part in visible_parts:
                    if part and part not in combined_parts:
                        # Kiểm tra trường hợp đặc biệt
                        if "mode" in part.lower() and re.search(r'^\w{1,4}\s+\w+\s+mode$', part.lower()):
                            continue
                        combined_parts.append(part)
                
                # Thêm các phần từ morecontent_text mà không trùng lặp
                for part in morecontent_parts:
                    is_duplicate = False
                    for existing_part in combined_parts:
                        if (part in existing_part or existing_part in part or
                            (("protection" in part.lower() and "protection" in existing_part.lower()) and
                             any(w in existing_part.lower() for w in part.lower().split() if w != "protection")) or
                            (("mode" in part.lower() and "mode" in existing_part.lower()) and
                             any(w in existing_part.lower() for w in part.lower().split() if w != "mode"))):
                            is_duplicate = True
                            break
                    
                    # Kiểm tra trường hợp đặc biệt
                    if "mode" in part.lower() and re.search(r'^\w{1,4}\s+\w+\s+mode$', part.lower()):
                        is_duplicate = True
                    
                    # Trường hợp "ply" trong "supply"
                    if len(part) <= 3 and any(part.lower() in existing.lower() for existing in combined_parts):
                        is_duplicate = True
                    
                    if not is_duplicate and part and part not in combined_parts:
                        combined_parts.append(part)
                
                # Nối lại thành chuỗi
                full_value = ", ".join(combined_parts)
            else:
                full_value = visible_text if visible_text else morecontent_text
            
            # Làm sạch kết quả
            full_value = re.sub(r'\s+', ' ', full_value).strip()
            
            if full_value:
                return full_value
        except Exception as e:
            print(f"Lỗi khi phân tích HTML: {str(e)}")
    
    # Phương pháp 3: Tìm tất cả các span
    all_spans = []
    for span in value_cell.select('span'):
        # Bỏ qua các phần tử UI
        skip_classes = ['moreellipses', 'morelink']
        if not any(cls in span.get('class', []) for cls in skip_classes) and span.get('role') != 'button':
            span_text = span.get_text(strip=True)
            # Kiểm tra trường hợp đặc biệt
            if is_special_field and "mode" in span_text.lower() and re.search(r'^\w{1,4}\s+\w+\s+mode$', span_text.lower()):
                continue
                
            if span_text and span_text not in all_spans:  # Thêm kiểm tra trùng lặp
                all_spans.append(span_text)
    
    if all_spans:
        # Nối tất cả các span lại với nhau
        full_text = ', '.join(all_spans)
        # Làm sạch text
        full_text = re.sub(r'\s+', ' ', full_text).strip()
        
        # Loại bỏ trùng lặp
        parts = [part.strip() for part in full_text.split(',')]
        unique_parts = []
        for part in parts:
            if part and part not in unique_parts and not any(
                part in unique and len(part) < len(unique) for unique in unique_parts
            ):
                # Kiểm tra đặc biệt cho các trường có từ khóa "mode" hoặc "protection"
                if "mode" in part.lower() or "protection" in part.lower():
                    duplicate = False
                    for existing in unique_parts:
                        if (("mode" in existing.lower() and "mode" in part.lower()) and
                            any(word in existing.lower() for word in part.lower().split() if word != "mode")):
                            duplicate = True
                            break
                        if (("protection" in existing.lower() and "protection" in part.lower()) and
                            any(word in existing.lower() for word in part.lower().split() if word != "protection")):
                            duplicate = True
                            break
                    
                    if not duplicate:
                        unique_parts.append(part)
                else:
                    unique_parts.append(part)
        
        return ", ".join(unique_parts)
    
    # Phương pháp 4: Lấy tất cả text từ ô và làm sạch
    all_text = value_cell.get_text(strip=True)
    all_text = re.sub(r'\[\.\.\.\]', '', all_text)
    all_text = re.sub(r'Hiển thị (thêm|bớt)', '', all_text)
    all_text = re.sub(r'\s+', ' ', all_text).strip()
    
    # Loại bỏ trùng lặp từ all_text
    parts = [part.strip() for part in all_text.split(',')]
    unique_parts = []
    for part in parts:
        if part and part not in unique_parts and not any(
            part in unique and len(part) < len(unique) for unique in unique_parts
        ):
            # Kiểm tra đặc biệt với từ "mode" hoặc "protection"
            if "mode" in part.lower() or "protection" in part.lower():
                duplicate = False
                for existing in unique_parts:
                    if (("mode" in existing.lower() and "mode" in part.lower()) and
                        any(word in existing.lower() for word in part.lower().split() if word != "mode")):
                        duplicate = True
                        break
                    if (("protection" in existing.lower() and "protection" in part.lower()) and
                        any(word in existing.lower() for word in part.lower().split() if word != "protection")):
                        duplicate = True
                        break
                
                if not duplicate:
                    unique_parts.append(part)
            else:
                # Trường hợp "ply" trong "supply"
                if len(part) <= 3 and any(part.lower() in existing.lower() for existing in unique_parts):
                    continue
                unique_parts.append(part)
    
    return ", ".join(unique_parts)

def extract_product_urls(url):
    """
    Trích xuất tất cả URL sản phẩm từ một URL danh mục (với đa luồng)
    """
    product_urls = []
    processed_pages = set()
    pages_to_process = [url]
    
    # Hàm xử lý trang
    def process_page(page_url):
        if page_url in processed_pages:
            return [], []
        
        local_product_urls = []
        next_pages = []
        
        try:
            print(f"Đang xử lý URL danh mục: {page_url}")
            html = get_html_content(page_url)
            if not html:
                print(f"Không thể tải nội dung từ {page_url}")
                return [], []
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Debug: In ra một số thông tin về trang
            print(f"  > Đã tải HTML, kích thước: {len(html)} ký tự")
            
            # Lấy các link sản phẩm với nhiều selector khác nhau
            product_selectors = [
                'a[href*="/san-pham/"]',  # Selector cũ
                'a[href*="/product/"]',   # Có thể dùng product thay vì san-pham
                '.product-item a',        # Link trong product item
                '.product-card a',        # Link trong product card
                '.product-list a',        # Link trong product list
                'a.product-link',         # Class product-link
                'a.product-item-link',    # Class product-item-link
                '.col-product a'          # Link trong column product
            ]
            
            total_links_found = 0
            
            for selector in product_selectors:
                links = soup.select(selector)
                print(f"  > Selector '{selector}' tìm thấy {len(links)} link")
                
                for a in links:
                    href = a.get('href')
                    if href:
                        # Tạo URL đầy đủ
                        if not href.startswith('http'):
                            parsed_url = urlparse(page_url)
                            full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}" if href.startswith('/') else f"{page_url.rstrip('/')}/{href}"
                        else:
                            full_url = href
                                        
                        # Kiểm tra xem có phải URL sản phẩm hợp lệ không
                        if is_product_url(full_url) and full_url not in local_product_urls:
                            local_product_urls.append(full_url)
                            total_links_found += 1
                            print(f"    + Tìm thấy URL sản phẩm: {full_url}")
            
            # Nếu không tìm được sản phẩm với selector cũ, thử tìm với pattern URL đặc biệt của BAA.vn
            if total_links_found == 0:
                print(f"  > Không tìm được sản phẩm với selector cũ, thử pattern BAA.vn...")
                
                # Tìm tất cả link và kiểm tra pattern
                all_links = soup.select('a[href]')
                print(f"  > Tìm thấy {len(all_links)} link để kiểm tra pattern")
                
                for a in all_links:
                    href = a.get('href')
                    if href:
                        # Tạo URL đầy đủ
                        if not href.startswith('http'):
                            parsed_url = urlparse(page_url)
                            full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}" if href.startswith('/') else f"{page_url.rstrip('/')}/{href}"
                        else:
                            full_url = href
                        
                        # Kiểm tra pattern đặc biệt của BAA.vn
                        if _is_baa_product_url(full_url) and full_url not in local_product_urls:
                            local_product_urls.append(full_url)
                            total_links_found += 1
                            print(f"    + Tìm thấy URL sản phẩm BAA pattern: {full_url}")
            
            print(f"  > Tổng cộng tìm thấy {total_links_found} URL sản phẩm hợp lệ")
            
            # Xử lý phân trang với nhiều selector
            pagination_selectors = [
                'a[href*="/page/"]',      # Selector cũ
                'a[href*="page="]',       # Query parameter page
                '.pagination a',          # Link trong pagination
                '.page-list a',           # Link trong page list
                'nav a[href*="page"]'     # Navigation với page
            ]
            
            for selector in pagination_selectors:
                page_links = soup.select(selector)
                print(f"  > Pagination selector '{selector}' tìm thấy {len(page_links)} link")
                
                for page_link in page_links:
                    page_href = page_link.get('href')
                    if page_href:
                        if not page_href.startswith('http'):
                            parsed_url = urlparse(page_url)
                            full_page_url = f"{parsed_url.scheme}://{parsed_url.netloc}{page_href}" if page_href.startswith('/') else f"{page_url.rstrip('/')}/{page_href}"
                        else:
                            full_page_url = page_href
                        
                        if full_page_url not in processed_pages and full_page_url not in next_pages:
                            next_pages.append(full_page_url)
                            print(f"    + Tìm thấy trang phân trang: {full_page_url}")
            
            return local_product_urls, next_pages
        except Exception as e:
            print(f"Lỗi khi xử lý trang {page_url}: {str(e)}")
            import traceback
            traceback.print_exc()
            return [], []
    
    # Số luồng tối đa
    max_workers = 10
    
    while pages_to_process:
        current_batch = pages_to_process[:max_workers]
        pages_to_process = pages_to_process[max_workers:]
        
        # Xử lý theo batch để tránh quá tải
        with ThreadPoolExecutor(max_workers=min(max_workers, len(current_batch))) as executor:
            futures = [executor.submit(process_page, page_url) for page_url in current_batch]
            
            for future in as_completed(futures):
                try:
                    page_products, page_pagination = future.result()
                    
                    # Thêm sản phẩm vào danh sách
                    for prod_url in page_products:
                        if prod_url not in product_urls:
                            product_urls.append(prod_url)
                    
                    # Thêm trang phân trang vào danh sách cần xử lý
                    for page_url in page_pagination:
                        if page_url not in processed_pages and page_url not in pages_to_process:
                            pages_to_process.append(page_url)
                            
                except Exception as e:
                    print(f"Lỗi khi xử lý future: {str(e)}")
        
        # Đánh dấu các trang đã xử lý
        for page_url in current_batch:
            processed_pages.add(page_url)
    
    print(f"Đã xử lý {len(processed_pages)} trang, tìm thấy {len(product_urls)} URL sản phẩm")
    return product_urls

def get_product_info(url, required_fields=None):
    """
    Trích xuất thông tin sản phẩm dựa trên mẫu URL
    """
    product_info_list = []
    
    # Kiểm tra nếu URL là trang danh mục
    if is_category_url(url):
        # Trích xuất tất cả URL sản phẩm hợp lệ từ trang danh mục
        product_urls = extract_product_urls(url)
        
        # Xử lý từng URL sản phẩm hợp lệ
        for i, product_url in enumerate(product_urls):
            product_info = extract_product_info(product_url, required_fields, i+1)
            if product_info:
                # Đảm bảo trường URL được loại bỏ nếu không có trong required_fields
                if required_fields and 'URL' not in required_fields and 'URL' in product_info:
                    product_info.pop('URL')
                product_info_list.append(product_info)
    
    # Kiểm tra nếu URL là trang sản phẩm hợp lệ
    elif is_product_url(url):
        product_info = extract_product_info(url, required_fields, 1)
        if product_info:
            # Đảm bảo trường URL được loại bỏ nếu không có trong required_fields
            if required_fields and 'URL' not in required_fields and 'URL' in product_info:
                product_info.pop('URL')
            product_info_list.append(product_info)
    
    return product_info_list

def scrape_product_info(product_urls, excel_template_path):
    """Thu thập thông tin từ danh sách URL sản phẩm và xuất ra file Excel"""
    # Lọc các URL là URL sản phẩm hợp lệ
    valid_product_urls = [url for url in product_urls if is_product_url(url)]
    print(f"Tìm thấy {len(valid_product_urls)} URL sản phẩm hợp lệ")
    
    # Gửi thông báo bắt đầu
    socketio.emit('progress_update', {'percent': 0, 'message': f'Bắt đầu thu thập thông tin từ {len(valid_product_urls)} sản phẩm'})
    
    # Các trường cần thu thập
    required_fields = ['STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Giá', 'Tổng quan']
    
    # Sử dụng ThreadPoolExecutor để xử lý đa luồng
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    all_products_info = []
    total_products = len(valid_product_urls)
    processed_count = 0
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Tạo các task xử lý sản phẩm
        future_to_url = {
            executor.submit(extract_product_info, url, required_fields, i+1): url 
            for i, url in enumerate(valid_product_urls)
        }
        
        # Xử lý kết quả khi hoàn thành
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                product_info = future.result()
                if product_info:
                    all_products_info.append(product_info)
                
                processed_count += 1
                progress = int((processed_count / total_products) * 100)
                socketio.emit('progress_update', {
                    'percent': progress,
                    'message': f'Đã xử lý {processed_count}/{total_products} sản phẩm'
                })
                
            except Exception as e:
                print(f"Lỗi khi xử lý {url}: {str(e)}")
    
    # Gửi thông báo hoàn thành
    socketio.emit('progress_update', {
        'percent': 100,
        'message': f'Đã hoàn thành việc thu thập thông tin từ {len(all_products_info)} sản phẩm'
    })
    
    # Tạo DataFrame và xuất Excel
    results_df = pd.DataFrame(all_products_info)
    results_df = results_df[required_fields]
    
    # Tạo file Excel tạm
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Ghi DataFrame vào Excel
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False, sheet_name='Sản phẩm')
        
        # Định dạng các cột
        worksheet = writer.sheets['Sản phẩm']
        worksheet.column_dimensions['A'].width = 5  # STT
        worksheet.column_dimensions['B'].width = 20  # Mã sản phẩm
        worksheet.column_dimensions['C'].width = 40  # Tên sản phẩm
        worksheet.column_dimensions['D'].width = 15  # Giá
        worksheet.column_dimensions['E'].width = 80  # Tổng quan
    
    return temp_file.name

def search_autonics_product(product_code):
    """
    Tìm kiếm sản phẩm trên website Autonics theo mã sản phẩm
    
    Args:
        product_code (str): Mã sản phẩm cần tìm
        
    Returns:
        str: URL của sản phẩm nếu tìm thấy, None nếu không tìm thấy
    """
    try:
        # Xây dựng URL tìm kiếm
        search_url = f"https://www.autonics.com/vn/search/total?keyword={product_code}"
        print(f"Tìm kiếm sản phẩm {product_code} trên Autonics.com: {search_url}")
        
        # Gửi request đến trang tìm kiếm
        response = requests.get(search_url, headers=HEADERS, timeout=10)
        response.raise_for_status()
        
        # Parse HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Tìm các sản phẩm trong kết quả tìm kiếm
        product_items = soup.select('.product-item')
        
        if not product_items:
            print(f"Không tìm thấy sản phẩm {product_code} trên Autonics.com")
            return None
        
        # Lấy liên kết của sản phẩm đầu tiên
        first_product = product_items[0]
        product_link = first_product.select_one('a')
        
        if not product_link:
            print(f"Không tìm thấy liên kết cho sản phẩm {product_code}")
            return None
        
        # Xây dựng URL đầy đủ
        product_url = product_link.get('href')
        if not product_url.startswith('http'):
            product_url = f"https://www.autonics.com{product_url}"
        
        print(f"Đã tìm thấy sản phẩm {product_code}: {product_url}")
        return product_url
    
    except Exception as e:
        print(f"Lỗi khi tìm kiếm sản phẩm {product_code}: {str(e)}")
        return None

def get_product_url(product_code):
    """
    Tạo URL sản phẩm từ mã sản phẩm
    
    Args:
        product_code (str): Mã sản phẩm
        
    Returns:
        str: URL của sản phẩm nếu tìm thấy, None nếu không tìm thấy
    """
    try:
        # Làm sạch mã sản phẩm
        clean_code = product_code.strip()
        
        # Cách 1: Tạo URL trực tiếp Autonics
        direct_url = f"https://www.autonics.com/vn/model/{clean_code}"
        print(f"Thử URL Autonics trực tiếp: {direct_url}")
        
        # Kiểm tra URL trực tiếp
        try:
            response = requests.head(direct_url, headers=HEADERS, timeout=5)
            if response.status_code == 200:
                print(f"URL Autonics trực tiếp hợp lệ: {direct_url}")
                return direct_url
        except:
            pass
        
        # Cách 2: Tìm kiếm trên Autonics
        autonics_url = search_autonics_product(clean_code)
        if autonics_url:
            return autonics_url
        
        # Cách 3: Thử các mẫu URL BAA.vn
        baa_patterns = [
            f"https://baa.vn/vn/san-pham/bo-chuyen-doi-nhiet-do-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/cam-bien-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/cam-bien-tinh-tien-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/cam-bien-quang-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/bo-dieu-khien-nhiet-do-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/dong-ho-hien-thi-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/bo-dem-autonics-{clean_code.lower()}",
            f"https://baa.vn/vn/san-pham/man-hinh-logic-autonics-{clean_code.lower()}"
        ]
        
        for url in baa_patterns:
            try:
                print(f"Thử URL BAA.vn: {url}")
                response = requests.head(url, timeout=3)
                if response.status_code == 200 and 'baa.vn' in response.url and 'tim-kiem' not in response.url:
                    print(f"URL BAA.vn hợp lệ: {response.url}")
                    return response.url
            except:
                continue
        
        # Cách 4: Tìm kiếm trên BAA.vn
        try:
            search_url = f"https://baa.vn/tim-kiem?q={clean_code}"
            print(f"Tìm kiếm trên BAA.vn: {search_url}")
            
            response = requests.get(search_url, timeout=5)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                product_links = soup.select('.product-name a, .product-item a.name, .product-list a')
                
                for link in product_links:
                    href = link.get('href')
                    link_text = link.text.strip()
                    
                    if href and clean_code.lower() in link_text.lower():
                        if not href.startswith('http'):
                            href = urljoin('https://baa.vn/', href)
                        print(f"Tìm thấy sản phẩm trên BAA.vn: {href}")
                        return href
        except:
            pass
        
        print(f"Không tìm thấy URL cho sản phẩm {clean_code}")
        return None
        
    except Exception as e:
        print(f"Lỗi khi tạo URL sản phẩm cho {product_code}: {str(e)}")
        return None

def extract_product_image(product_url):
    """Trích xuất hình ảnh sản phẩm từ trang chi tiết sản phẩm"""
    try:
        # Làm sạch URL trước khi truy cập
        clean_url = product_url.strip()
        
        # Xử lý URL có chứa kí tự đặc biệt như dấu ngoặc
        if '(' in clean_url or ')' in clean_url:
            # Trích xuất mã sản phẩm từ URL (phần trước dấu ngoặc)
            product_code = clean_url.split('/')[-1].split('(')[0].strip()
            # Tạo URL mới không chứa dấu ngoặc
            clean_url = f"https://www.autonics.com/vn/model/{product_code}"
            print(f"URL có chứa kí tự đặc biệt, đã chuyển sang: {clean_url}")
        else:
            # Lấy mã sản phẩm từ URL
            product_code = clean_url.split('/')[-1]
        
        # Tải trang chi tiết sản phẩm không giới hạn timeout
        try:
            response = requests.get(clean_url, headers=HEADERS)
            response.raise_for_status()
        except requests.RequestException as e:
            print(f"Lỗi khi tải trang sản phẩm {clean_url}: {str(e)}")
            
            # Giới hạn chỉ tìm kiếm thêm nếu không thể mở URL trực tiếp
            alternative_url = search_autonics_product(product_code)
            if alternative_url:
                print(f"Tìm thấy URL thay thế: {alternative_url}")
                response = requests.get(alternative_url, headers=HEADERS)
                response.raise_for_status()
            else:
                print(f"Không tìm thấy URL thay thế cho {product_code}")
                return None
        
        # Parse HTML - Sử dụng html.parser nhanh hơn lxml
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Tìm hình ảnh sản phẩm chính với bộ selector tối ưu
        # Ưu tiên các selector phổ biến nhất
        img_element = None
        
        # Selector chính - thử trước tiên và nhanh nhất
        primary_selectors = [
            '.img img#img-chg',
            '.product-detail-image img',
            '.product-image img'
        ]
        
        # Thử các selector chính trước
        for selector in primary_selectors:
            img_element = soup.select_one(selector)
            if img_element and img_element.get('src'):
                break
        
        # Nếu không tìm thấy, thử tìm theo mã sản phẩm trong thuộc tính src/alt
        if not img_element:
            # Tìm tất cả các ảnh
            all_images = soup.select('img')
            
            # Ưu tiên tìm ảnh theo mã sản phẩm
            for img in all_images:
                src = img.get('src', '')
                alt = img.get('alt', '')
                
                # Kiểm tra nếu mã sản phẩm có trong src hoặc alt
                if product_code.lower() in src.lower() or product_code.lower() in alt.lower():
                    img_element = img
                    print(f"Tìm thấy ảnh theo mã sản phẩm: {product_code}")
                    break
            
            # Nếu vẫn không tìm thấy, tìm theo các từ khóa phổ biến
            if not img_element:
                keywords = ['product', 'item', 'main', 'gallery', 'detail']
                for img in all_images:
                    src = img.get('src', '')
                    for keyword in keywords:
                        if keyword in src.lower():
                            img_element = img
                            print(f"Tìm thấy ảnh theo từ khóa: {keyword}")
                            break
                    if img_element:
                        break
            
            # Nếu vẫn không tìm thấy, lấy ảnh đầu tiên có kích thước hợp lý
            if not img_element:
                for img in all_images:
                    src = img.get('src', '')
                    if src and not ('icon' in src.lower() or 'logo' in src.lower()):
                        width = img.get('width', '0')
                        height = img.get('height', '0')
                        
                        # Nếu có thuộc tính width, height > 100px thì có thể là ảnh sản phẩm
                        try:
                            if int(width) > 100 and int(height) > 100:
                                img_element = img
                                print(f"Tìm thấy ảnh có kích thước lớn: {width}x{height}")
                                break
                        except ValueError:
                            # Nếu không chuyển được sang số, bỏ qua
                            pass
                
                # Nếu vẫn không tìm được, lấy ảnh đầu tiên có src hợp lệ
                if not img_element:
                    for img in all_images:
                        src = img.get('src', '')
                        if src and src.endswith(('.jpg', '.jpeg', '.png', '.webp')):
                            img_element = img
                            print(f"Tìm thấy ảnh đầu tiên có định dạng hợp lệ")
                            break
        
        # Nếu không tìm thấy ảnh
        if not img_element or not img_element.get('src'):
            print(f"Không tìm thấy hình ảnh trong trang {clean_url}")
            return None
        
        # Lấy URL hình ảnh và xử lý để chắc chắn là URL đầy đủ
        img_url = img_element.get('src')
        if not img_url.startswith('http'):
            # Xử lý URL tương đối
            base_url = "https://www.autonics.com"
            if img_url.startswith('//'):
                img_url = 'https:' + img_url
            elif img_url.startswith('/'):
                img_url = base_url + img_url
            else:
                img_url = urljoin(base_url, img_url)
            
        print(f"Đã tìm thấy ảnh sản phẩm {product_code}: {img_url}")
        return {
            'url': img_url,
            'code': product_code
        }
    
    except Exception as e:
        print(f"Lỗi khi trích xuất hình ảnh từ {product_url}: {str(e)}")
        return None

def download_product_image(img_info, output_folder):
    """Tải về hình ảnh sản phẩm"""
    img_path = None
    try:
        if not img_info or not img_info.get('url'):
            print("Không có thông tin ảnh để tải")
            return None
        
        # Tạo tên file ảnh theo định dạng mới
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        
        # Tạo đường dẫn mới theo định dạng yêu cầu
        new_path = f"/wp-content/uploads/{year}/{month:02d}/{img_info['code']}.webp"
        
        # Tạo tên file an toàn - loại bỏ tất cả các kí tự không an toàn cho tên file
        safe_name = re.sub(r'[^\w\-_]', '_', img_info['code'])
        img_filename = f"{safe_name}.webp"
        
        # Tạo cấu trúc thư mục đầy đủ trong output_folder
        year_month_folder = os.path.join(output_folder, str(year), f"{month:02d}")
        os.makedirs(year_month_folder, exist_ok=True)
        
        # Đường dẫn đầy đủ để lưu file
        img_path = os.path.join(year_month_folder, img_filename)
        
        print(f"Đang tải ảnh từ: {img_info['url']}")
        print(f"Lưu vào: {img_path}")
        
        # Tải ảnh với timeout cao hơn
        max_retries = 2
        retry_count = 0
        last_error = None
        
        while retry_count < max_retries:
            try:
                # Tạo session mới để thêm các header cần thiết
                session = requests.Session()
                session.headers.update(HEADERS)
                
                # Thêm referer vào header để tránh bị chặn
                session.headers.update({
                    'Referer': 'https://www.autonics.com/',
                    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8'
                })
                
                # Tải ảnh với stream=True và timeout 10s
                response = session.get(img_info['url'], stream=True, timeout=10)
                response.raise_for_status()
                
                # Kiểm tra Content-Type
                content_type = response.headers.get('Content-Type', '')
                if not content_type.startswith('image/'):
                    print(f"Lỗi: URL không trả về hình ảnh (Content-Type: {content_type})")
                    # Thử lại nếu chưa đạt số lần thử tối đa
                    if retry_count == max_retries - 1:
                        return None
                    retry_count += 1
                    time.sleep(0.5)
                    continue
                    
                # Kiểm tra kích thước file
                content_length = int(response.headers.get('Content-Length', 0))
                if content_length < 100:  # Ảnh quá nhỏ có thể là lỗi
                    print(f"Cảnh báo: Kích thước ảnh quá nhỏ ({content_length} bytes)")
                    if content_length == 0:
                        print("Lỗi: File ảnh rỗng")
                        # Thử lại nếu chưa đạt số lần thử tối đa
                        if retry_count == max_retries - 1:
                            return None
                        retry_count += 1
                        time.sleep(0.5)
                        continue
                
                # Lưu ảnh với xử lý lỗi
                try:
                    with open(img_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    
                    # Kiểm tra kích thước file đã tải
                    file_size = os.path.getsize(img_path)
                    if file_size < 100:  # File quá nhỏ, có thể là lỗi
                        print(f"Cảnh báo: File đã tải có kích thước nhỏ: {file_size} bytes")
                        if file_size == 0:
                            print("Lỗi: File ảnh đã tải bị rỗng, sẽ thử lại")
                            os.remove(img_path)
                            # Thử lại nếu chưa đạt số lần thử tối đa
                            if retry_count == max_retries - 1:
                                return None
                            retry_count += 1
                            time.sleep(0.5)
                            continue
                    
                    # Xác nhận thành công
                    print(f"Đã tải thành công: {img_filename} ({file_size} bytes)")
                    return {
                        'path': img_path,
                        'url': f"https://haiphongtech.vn{new_path}"
                    }
                except Exception as e:
                    print(f"Lỗi khi lưu file: {str(e)}")
                    # Xóa file lỗi nếu có
                    if os.path.exists(img_path):
                        os.remove(img_path)
                    # Thử lại nếu chưa đạt số lần thử tối đa
                    if retry_count == max_retries - 1:
                        raise e
                    retry_count += 1
                    last_error = e
                    time.sleep(0.5)
                    continue
            
            except Exception as e:
                print(f"Lỗi khi tải ảnh (lần {retry_count+1}): {str(e)}")
                # Thử lại nếu chưa đạt số lần thử tối đa
                if retry_count == max_retries - 1:
                    raise e
                retry_count += 1
                last_error = e
                time.sleep(0.5)
                continue
        
        # Nếu đã thử tối đa nhưng vẫn thất bại
        if last_error:
            print(f"Đã thử {max_retries} lần nhưng không thành công: {str(last_error)}")
        return None
    
    except Exception as e:
        print(f"Lỗi khi tải hình ảnh: {str(e)}")
        # Xóa file lỗi nếu có
        if img_path and os.path.exists(img_path):
            try:
                os.remove(img_path)
                print(f"Đã xóa file lỗi: {img_path}")
            except:
                print(f"Không thể xóa file lỗi: {img_path}")
        return None

def download_jpg_product_image(img_info, output_folder):
    """Tải về hình ảnh sản phẩm chất lượng cao dưới định dạng JPG"""
    img_path = None
    try:
        if not img_info or not img_info.get('url'):
            print("Không có thông tin ảnh để tải")
            return None
        
        # Tạo tên file ảnh theo định dạng mới
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        
        # Tạo đường dẫn mới theo định dạng yêu cầu
        new_path = f"/wp-content/uploads/{year}/{month:02d}/{img_info['code']}.jpg"
        
        # Tạo tên file an toàn - loại bỏ tất cả các kí tự không an toàn cho tên file
        safe_name = re.sub(r'[^\w\-_]', '_', img_info['code'])
        img_filename = f"{safe_name}.jpg"
        
        # Tạo cấu trúc thư mục đầy đủ trong output_folder
        year_month_folder = os.path.join(output_folder, str(year), f"{month:02d}")
        os.makedirs(year_month_folder, exist_ok=True)
        
        # Đường dẫn đầy đủ để lưu file
        img_path = os.path.join(year_month_folder, img_filename)
        
        print(f"Đang tải ảnh JPG chất lượng cao từ: {img_info['url']}")
        print(f"Lưu vào: {img_path}")
        
        # Tải ảnh với timeout cao hơn
        max_retries = 2
        retry_count = 0
        last_error = None
        
        while retry_count < max_retries:
            try:
                # Tạo session mới để thêm các header cần thiết
                session = requests.Session()
                session.headers.update(HEADERS)
                
                # Thêm referer vào header để tránh bị chặn
                session.headers.update({
                    'Referer': 'https://www.autonics.com/',
                    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8'
                })
                
                # Tải ảnh với stream=True và timeout 10s
                response = session.get(img_info['url'], stream=True, timeout=10)
                response.raise_for_status()
                
                # Kiểm tra Content-Type
                content_type = response.headers.get('Content-Type', '')
                if not content_type.startswith('image/'):
                    print(f"Lỗi: URL không trả về hình ảnh (Content-Type: {content_type})")
                    # Thử lại nếu chưa đạt số lần thử tối đa
                    if retry_count == max_retries - 1:
                        return None
                    retry_count += 1
                    time.sleep(0.5)
                    continue
                    
                # Kiểm tra kích thước file
                content_length = int(response.headers.get('Content-Length', 0))
                if content_length < 100:  # Ảnh quá nhỏ có thể là lỗi
                    print(f"Cảnh báo: Kích thước ảnh quá nhỏ ({content_length} bytes)")
                    if content_length == 0:
                        print("Lỗi: File ảnh rỗng")
                        # Thử lại nếu chưa đạt số lần thử tối đa
                        if retry_count == max_retries - 1:
                            return None
                        retry_count += 1
                        time.sleep(0.5)
                        continue
                
                # Đọc dữ liệu ảnh vào bộ nhớ
                image_data = io.BytesIO(response.content)
                
                try:
                    # Xử lý ảnh bằng Pillow
                    with Image.open(image_data) as img:
                        # Chuyển đổi sang RGB nếu cần
                        if img.mode in ('RGBA', 'P'):
                            img = img.convert('RGB')
                        
                        # Đảm bảo thư mục tồn tại
                        os.makedirs(os.path.dirname(img_path), exist_ok=True)
                        
                        # Lưu ảnh dưới định dạng JPEG với chất lượng cao nhất
                        img.save(img_path, 'JPEG', quality=100, optimize=True, subsampling=0)
                        
                    # Kiểm tra kích thước file đã tải
                    file_size = os.path.getsize(img_path)
                    if file_size < 100:  # File quá nhỏ, có thể là lỗi
                        print(f"Cảnh báo: File đã tải có kích thước nhỏ: {file_size} bytes")
                        if file_size == 0:
                            print("Lỗi: File ảnh đã tải bị rỗng, sẽ thử lại")
                            os.remove(img_path)
                            # Thử lại nếu chưa đạt số lần thử tối đa
                            if retry_count == max_retries - 1:
                                return None
                            retry_count += 1
                            time.sleep(0.5)
                            continue
                    
                    # Xác nhận thành công
                    print(f"Đã tải và lưu thành công ảnh JPG chất lượng cao: {img_filename} ({file_size} bytes)")
                    return {
                        'path': img_path,
                        'url': f"https://haiphongtech.vn{new_path}"
                    }
                except Exception as e:
                    print(f"Lỗi khi xử lý ảnh: {str(e)}")
                    # Xóa file lỗi nếu có
                    if os.path.exists(img_path):
                        os.remove(img_path)
                    # Thử lại nếu chưa đạt số lần thử tối đa
                    if retry_count == max_retries - 1:
                        raise e
                    retry_count += 1
                    last_error = e
                    time.sleep(0.5)
                    continue
            
            except Exception as e:
                print(f"Lỗi khi tải ảnh (lần {retry_count+1}): {str(e)}")
                # Thử lại nếu chưa đạt số lần thử tối đa
                if retry_count == max_retries - 1:
                    raise e
                retry_count += 1
                last_error = e
                time.sleep(0.5)
                continue
        
        # Nếu đã thử tối đa nhưng vẫn thất bại
        if last_error:
            print(f"Đã thử {max_retries} lần nhưng không thành công: {str(last_error)}")
        return None
    
    except Exception as e:
        print(f"Lỗi khi tải hình ảnh: {str(e)}")
        # Xóa file lỗi nếu có
        if img_path and os.path.exists(img_path):
            try:
                os.remove(img_path)
                print(f"Đã xóa file lỗi: {img_path}")
            except:
                print(f"Không thể xóa file lỗi: {img_path}")
        return None

def download_autonics_images(product_codes, output_folder):
    """Tải nhiều hình ảnh sản phẩm từ danh sách mã sản phẩm sử dụng đa luồng"""
    # Đảm bảo thư mục đầu ra tồn tại
    os.makedirs(output_folder, exist_ok=True)
    
    total_products = len(product_codes)
    successful_downloads = 0
    failed_downloads = 0
    skipped_codes = 0
    download_results = []
    
    # Gửi thông báo bắt đầu
    socketio.emit('progress_update', {
        'percent': 0, 
        'message': f'Bắt đầu tải {total_products} hình ảnh sản phẩm...'
    })
    
    # Lập danh sách các mã sản phẩm hợp lệ và đã làm sạch
    clean_product_codes = []
    for product_code in product_codes:
        code = product_code.strip()
        if not code:  # Bỏ qua các dòng trống
            skipped_codes += 1
            continue
        
        # Làm sạch mã sản phẩm
        clean_code = re.sub(r'[\(\)\s]+', '', code)
        if clean_code != code:
            print(f"Đã làm sạch mã sản phẩm: {code} -> {clean_code}")
        
        clean_product_codes.append(clean_code)
    
    # Số lượng sản phẩm thực tế
    actual_total = len(clean_product_codes)
    
    # Khóa để đồng bộ hóa cập nhật kết quả
    result_lock = threading.Lock()
    progress_lock = threading.Lock()
    current_processed = 0
    
    # Hàm xử lý tải ảnh một sản phẩm
    def process_product(product_code, index):
        nonlocal current_processed, successful_downloads, failed_downloads
        
        # Kết quả ban đầu
        result = {
            'product_code': product_code,
            'status': 'Thất bại',
            'image_path': None,
            'image_url': None,
            'error': None
        }
        
        try:
            # Tạo URL sản phẩm
            product_url = f"https://www.autonics.com/vn/model/{product_code}"
            print(f"Đang truy cập URL: {product_url}")
            
            # Trích xuất thông tin hình ảnh
            img_info = extract_product_image(product_url)
            
            if not img_info:
                with result_lock:
                    result['error'] = 'Không tìm thấy hình ảnh'
                    failed_downloads += 1
                print(f"Không tìm thấy hình ảnh cho sản phẩm {product_code}")
            else:
                # Tải hình ảnh
                img_result = download_product_image(img_info, output_folder)
                
                if img_result:
                    with result_lock:
                        result['status'] = 'Thành công'
                        result['image_path'] = img_result['path']
                        result['image_url'] = img_result['url']
                        successful_downloads += 1
                    print(f"Tải thành công ảnh cho sản phẩm {product_code}")
                else:
                    with result_lock:
                        result['error'] = 'Không thể tải hình ảnh'
                        failed_downloads += 1
                    print(f"Tải ảnh thất bại cho sản phẩm {product_code}")
        
        except Exception as e:
            error_msg = str(e)
            with result_lock:
                result['error'] = error_msg
                failed_downloads += 1
            print(f"Lỗi khi xử lý sản phẩm {product_code}: {error_msg}")
        
        # Cập nhật tiến trình
        with progress_lock:
            current_processed += 1
            progress = int((current_processed / actual_total) * 100)
            socketio.emit('progress_update', {
                'percent': progress,
                'message': f'Đang xử lý sản phẩm {product_code} ({current_processed}/{actual_total})'
            })
        
        return result
    
    # Số lượng luồng tối đa
    max_workers = min(10, actual_total)  # Tối đa 10 luồng hoặc bằng số lượng sản phẩm nếu ít hơn
    
    # Sử dụng ThreadPoolExecutor để xử lý đa luồng
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Tạo các task để xử lý mã sản phẩm
        futures = {executor.submit(process_product, code, i+1): code for i, code in enumerate(clean_product_codes)}
        
        # Xử lý kết quả khi hoàn thành
        for future in as_completed(futures):
            product_code = futures[future]
            try:
                result = future.result()
                download_results.append(result)
            except Exception as e:
                print(f"Lỗi không xác định khi xử lý {product_code}: {str(e)}")
                # Tạo kết quả lỗi nếu có ngoại lệ không xử lý được
                download_results.append({
                    'product_code': product_code,
                    'status': 'Thất bại',
                    'image_path': None,
                    'image_url': None,
                    'error': f'Lỗi không xử lý được: {str(e)}'
                })
    
    # Gửi thông báo hoàn thành
    completion_message = f'Hoàn thành! Đã tải {successful_downloads}/{actual_total} hình ảnh'
    if failed_downloads > 0:
        completion_message += f', {failed_downloads} thất bại'
    if skipped_codes > 0:
        completion_message += f', {skipped_codes} bỏ qua'
        
    socketio.emit('progress_update', {
        'percent': 100,
        'message': completion_message
    })
    
    # Tạo báo cáo kết quả
    summary = {
        'total': actual_total,
        'successful': successful_downloads,
        'failed': failed_downloads,
        'skipped': skipped_codes,
        'output_folder': output_folder,
        'results': download_results
    }
    
    # Tạo file Excel báo cáo kết quả
    try:
        report_path = os.path.join(output_folder, 'download_report.xlsx')
        report_df = pd.DataFrame(download_results)
        report_df.to_excel(report_path, index=False)
        print(f"Đã tạo báo cáo tại: {report_path}")
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo: {str(e)}")
    
    return summary

def download_autonics_jpg_images(product_codes, output_folder):
    """Tải nhiều hình ảnh sản phẩm chất lượng cao dưới định dạng JPG từ danh sách mã sản phẩm với đa luồng"""
    # Đảm bảo thư mục đầu ra tồn tại
    os.makedirs(output_folder, exist_ok=True)
    
    total_products = len(product_codes)
    successful_downloads = 0
    failed_downloads = 0
    skipped_codes = 0
    download_results = []
    
    # Gửi thông báo bắt đầu
    socketio.emit('progress_update', {
        'percent': 0, 
        'message': f'Bắt đầu tải {total_products} hình ảnh JPG chất lượng cao...'
    })
    
    # Lập danh sách các mã sản phẩm hợp lệ và đã làm sạch
    clean_product_codes = []
    for product_code in product_codes:
        code = product_code.strip()
        if not code:  # Bỏ qua các dòng trống
            skipped_codes += 1
            continue
        
        # Làm sạch mã sản phẩm
        clean_code = re.sub(r'[\(\)\s]+', '', code)
        if clean_code != code:
            print(f"Đã làm sạch mã sản phẩm: {code} -> {clean_code}")
        
        clean_product_codes.append(clean_code)
    
    # Số lượng sản phẩm thực tế
    actual_total = len(clean_product_codes)
    
    # Khóa để đồng bộ hóa cập nhật kết quả
    result_lock = threading.Lock()
    progress_lock = threading.Lock()
    current_processed = 0
    
    # Hàm xử lý tải ảnh một sản phẩm
    def process_product(product_code, index):
        nonlocal current_processed, successful_downloads, failed_downloads
        
        # Đặt giới hạn thời gian xử lý tối đa cho mỗi sản phẩm
        start_time = time.time()
        
        # Kết quả ban đầu
        result = {
            'product_code': product_code,
            'status': 'Thất bại',
            'image_path': None,
            'image_url': None,
            'error': None
        }
        
        try:
            # Tạo URL sản phẩm
            product_url = f"https://www.autonics.com/vn/model/{product_code}"
            print(f"Đang truy cập URL: {product_url}")
            
            # Trích xuất thông tin hình ảnh
            img_info = extract_product_image(product_url)
            
            if not img_info:
                with result_lock:
                    result['error'] = 'Không tìm thấy hình ảnh'
                    failed_downloads += 1
                print(f"Không tìm thấy hình ảnh cho sản phẩm {product_code}")
            else:
                # Tải hình ảnh JPG chất lượng cao
                img_result = download_jpg_product_image(img_info, output_folder)
                
                if img_result:
                    with result_lock:
                        result['status'] = 'Thành công'
                        result['image_path'] = img_result['path']
                        result['image_url'] = img_result['url']
                        successful_downloads += 1
                    print(f"Tải thành công ảnh JPG chất lượng cao cho sản phẩm {product_code}")
                else:
                    with result_lock:
                        result['error'] = 'Không thể tải hình ảnh'
                        failed_downloads += 1
                    print(f"Tải ảnh JPG thất bại cho sản phẩm {product_code}")
            
            # Kiểm tra thời gian xử lý
            elapsed_time = time.time() - start_time
            print(f"Đã xử lý sản phẩm {product_code} trong {elapsed_time:.2f} giây")
            
            # Tạm dừng nếu còn thời gian trong giới hạn 10 giây
            remaining_time = 10 - elapsed_time
            if remaining_time > 0 and remaining_time < 1:
                time.sleep(remaining_time)  # Chỉ chờ nếu còn ít hơn 1 giây
            
        except Exception as e:
            error_msg = str(e)
            with result_lock:
                result['error'] = error_msg
                failed_downloads += 1
            print(f"Lỗi khi xử lý sản phẩm {product_code}: {error_msg}")
            
            # Kiểm tra thời gian đã trôi qua
            elapsed_time = time.time() - start_time
            print(f"Xử lý thất bại sau {elapsed_time:.2f} giây")
        
        # Cập nhật tiến trình
        with progress_lock:
            current_processed += 1
            progress = int((current_processed / actual_total) * 100)
            socketio.emit('progress_update', {
                'percent': progress,
                'message': f'Đang xử lý sản phẩm {product_code} ({current_processed}/{actual_total})'
            })
        
        return result
    
    # Số lượng luồng tối đa
    max_workers = min(10, actual_total)  # Tối đa 10 luồng hoặc bằng số lượng sản phẩm nếu ít hơn
    
    # Sử dụng ThreadPoolExecutor để xử lý đa luồng
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Tạo các task để xử lý mã sản phẩm
        futures = {executor.submit(process_product, code, i+1): code for i, code in enumerate(clean_product_codes)}
        
        # Xử lý kết quả khi hoàn thành
        for future in as_completed(futures):
            product_code = futures[future]
            try:
                result = future.result()
                download_results.append(result)
            except Exception as e:
                print(f"Lỗi không xác định khi xử lý {product_code}: {str(e)}")
                # Tạo kết quả lỗi nếu có ngoại lệ không xử lý được
                download_results.append({
                    'product_code': product_code,
                    'status': 'Thất bại',
                    'image_path': None,
                    'image_url': None,
                    'error': f'Lỗi không xử lý được: {str(e)}'
                })
    
    # Gửi thông báo hoàn thành
    completion_message = f'Hoàn thành! Đã tải {successful_downloads}/{actual_total} hình ảnh JPG chất lượng cao'
    if failed_downloads > 0:
        completion_message += f', {failed_downloads} thất bại'
    if skipped_codes > 0:
        completion_message += f', {skipped_codes} bỏ qua'
        
    socketio.emit('progress_update', {
        'percent': 100,
        'message': completion_message
    })
    
    # Tạo báo cáo kết quả
    summary = {
        'total': actual_total,
        'successful': successful_downloads,
        'failed': failed_downloads,
        'skipped': skipped_codes,
        'output_folder': output_folder,
        'results': download_results
    }
    
    # Tạo file Excel báo cáo kết quả
    try:
        report_path = os.path.join(output_folder, 'download_report.xlsx')
        report_df = pd.DataFrame(download_results)
        report_df.to_excel(report_path, index=False)
        print(f"Đã tạo báo cáo tại: {report_path}")
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo: {str(e)}")
    
    return summary

def download_product_document(doc_info, output_folder):
    """
    Tải tài liệu PDF từ URL và lưu vào thư mục đầu ra
    
    Args:
        doc_info (dict): Thông tin tài liệu cần tải
            - url: URL của tài liệu
            - name: Tên tài liệu (nếu không có sẽ lấy từ URL)
        output_folder (str): Đường dẫn thư mục đầu ra
        
    Returns:
        dict: Thông tin kết quả tải tài liệu
            - success: True nếu tải thành công, False nếu thất bại
            - path: Đường dẫn đến file đã tải (nếu thành công)
            - filename: Tên file đã tải (nếu thành công)
            - error: Thông báo lỗi (nếu thất bại)
            - url: URL của tài liệu
            - name: Tên gốc của tài liệu
    """
    try:
        # Tạo thư mục đầu ra nếu chưa tồn tại
        os.makedirs(output_folder, exist_ok=True)
        
        # Lấy URL tài liệu
        doc_url = doc_info['url']
        
        # Tạo tên file an toàn từ tên tài liệu
        doc_name = doc_info.get('name', '')
        if not doc_name:
            doc_name = doc_url.split('/')[-1]
        
        # Loại bỏ các ký tự không hợp lệ trong tên file
        safe_filename = re.sub(r'[\\/*?:"<>|]', '', doc_name)
        # Thêm .pdf nếu chưa có
        if not safe_filename.lower().endswith('.pdf'):
            safe_filename += '.pdf'
        
        # Tạo đường dẫn đầy đủ đến file đích
        file_path = os.path.join(output_folder, safe_filename)
        
        print(f"Đang tải tài liệu từ: {doc_url}")
        print(f"Lưu vào: {file_path}")
        
        # Tải tài liệu với số lần thử lại
        max_retries = 5
        current_retry = 0
        success = False
        error_msg = ""
        
        while current_retry < max_retries and not success:
            try:
                # Tạo session với User-Agent của trình duyệt
                session = requests.Session()
                session.headers.update({
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
                    'Accept': 'application/pdf,application/x-pdf,application/octet-stream,*/*',
                    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
                    'Connection': 'keep-alive',
                    'Referer': 'https://baa.vn/'
                })
                
                # Tạo một stream request để tải file mà không cần đọc toàn bộ vào bộ nhớ
                response = session.get(doc_url, stream=True, timeout=60)  # Tăng timeout lên 60 giây
                
                # Kiểm tra content-type
                content_type = response.headers.get('Content-Type', '').lower()
                if 'application/pdf' not in content_type and not (content_type.startswith('application/') or 'octet-stream' in content_type):
                    print(f"CẢNH BÁO: Nội dung không phải PDF (Content-Type: {content_type})")
                    
                    # Kiểm tra nếu header đầu tiên của file là %PDF-
                    first_chunk = next(response.iter_content(chunk_size=10), None)
                    if first_chunk and not first_chunk.startswith(b'%PDF-'):
                        raise Exception(f"Nội dung tải về không phải là file PDF hợp lệ (Content-Type: {content_type})")
                
                # Kiểm tra content-length
                content_length = response.headers.get('Content-Length')
                if content_length:
                    content_length = int(content_length)
                    if content_length < 1000:
                        raise Exception(f"File quá nhỏ ({content_length} bytes), có thể không phải file PDF hợp lệ")
                    
                # Lưu file
                with open(file_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                # Kiểm tra kích thước file
                file_size = os.path.getsize(file_path)
                if file_size < 1000:
                    raise Exception(f"File quá nhỏ ({file_size} bytes), có thể không phải file PDF hợp lệ")
                
                # Kiểm tra nội dung file
                with open(file_path, 'rb') as f:
                    header = f.read(5)
                    # Kiểm tra header PDF
                    if not header.startswith(b'%PDF-'):
                        raise Exception("File không có signature PDF hợp lệ (%PDF-)")
                
                print(f"Tải tài liệu thành công: {file_path} ({file_size} bytes)")
                success = True
                
                # Ngừng vòng lặp nếu thành công
                break
                
            except Exception as e:
                current_retry += 1
                error_msg = str(e)
                
                # Tăng thời gian chờ mỗi lần thử lại
                wait_time = current_retry * 2
                print(f"Lỗi khi tải tài liệu (lần thử {current_retry}/{max_retries}): {error_msg}")
                print(f"Thử lại sau {wait_time} giây...")
                
                if os.path.exists(file_path):
                    os.remove(file_path)
                
                time.sleep(wait_time)
        
        if success:
            return {
                'success': True,
                'path': file_path,  # Đổi từ filepath thành path
                'filename': safe_filename,
                'url': doc_url,
                'name': doc_name
            }
        else:
            return {
                'success': False,
                'error': error_msg,
                'url': doc_url,
                'name': doc_name
            }
    
    except Exception as e:
        error_msg = str(e)
        print(f"Lỗi khi tải tài liệu: {error_msg}")
        traceback.print_exc()
        
        return {
            'success': False,
            'error': error_msg,
            'url': doc_info.get('url', ''),
            'name': doc_info.get('name', '')
        }

def download_product_documents(product_codes_or_urls, output_folder='output_documents'):
    """
    Tải tất cả tài liệu cho danh sách mã sản phẩm hoặc URL sử dụng đa luồng
    
    Args:
        product_codes_or_urls (list): Danh sách mã sản phẩm hoặc URL
        output_folder (str): Thư mục đầu ra
        
    Returns:
        dict: Báo cáo về quá trình tải với các thông tin:
            - total_products: Tổng số sản phẩm
            - successful_products: Số sản phẩm tải thành công
            - failed_products: Số sản phẩm tải thất bại
            - product_results: Kết quả chi tiết cho mỗi sản phẩm
            - skipped_codes: Danh sách mã đã bị bỏ qua
            - output_folder: Thư mục đầu ra
    """
    # Kiểm tra input
    if not product_codes_or_urls:
        print("Không có mã sản phẩm hoặc URL nào được cung cấp")
        return None
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    os.makedirs(output_folder, exist_ok=True)
    
    # Khởi tạo biến theo dõi
    total_products = len(product_codes_or_urls)
    successful_products = 0
    failed_products = 0
    skipped_codes = []
    product_results = {}
    
    # Khóa đồng bộ hóa cho các biến toàn cục
    result_lock = threading.Lock()
    
    # Định nghĩa hàm xử lý cho mỗi sản phẩm
    def process_product(item, index):
        try:
            print(f"\n[{index}/{total_products}] Đang xử lý: {item}")
            
            # Xác định xem đầu vào là URL hay mã sản phẩm
            if item.startswith('http'):
                product_url = item
                # Trích xuất mã sản phẩm từ URL nếu có thể
                try:
                    parts = product_url.split('/')
                    product_code = parts[-1] if parts[-1] else parts[-2]
                    
                    # Trích xuất tên sản phẩm từ URL (với URL từ baa.vn)
                    product_name = ""
                    if 'baa.vn' in product_url:
                        # Tìm tên sản phẩm từ phần cuối URL - thường có dạng ten-san-pham_xyz
                        product_code_parts = product_code.split('_')
                        if len(product_code_parts) > 0:
                            product_slug = product_code_parts[0]
                            # Chuyển slug thành tên đẹp hơn
                            product_name = ' '.join([part.capitalize() for part in product_slug.split('-')])
                    
                    # Nếu không thể trích xuất tên từ URL, thử tải trang sản phẩm
                    if not product_name or product_name == product_code:
                        try:
                            # Tạo session với User-Agent giống trình duyệt
                            session = requests.Session()
                            session.headers.update({
                                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
                            })
                            response = session.get(product_url, timeout=30)
                            soup = BeautifulSoup(response.text, 'html.parser')
                            
                            # Thử tìm tên sản phẩm từ các selector thường gặp
                            name_selectors = [
                                'h1.product-title', '.product-name h1', '.page-title h1', 
                                '.product__name', '.product-detail h1', '.product-title'
                            ]
                            
                            for selector in name_selectors:
                                name_elem = soup.select_one(selector)
                                if name_elem:
                                    product_name = name_elem.text.strip()
                                    print(f"  Tìm thấy tên sản phẩm: {product_name}")
                                    break
                        except Exception as e:
                            print(f"  Không thể trích xuất tên sản phẩm từ trang: {str(e)}")
                except:
                    product_code = f"product_{index}"
                    product_name = f"Product {index}"
            else:
                product_code = item
                product_url = get_product_url(product_code)
                product_name = f"Product {product_code}"  # Default name
                
                # Thử lấy tên sản phẩm từ URL
                if product_url:
                    try:
                        # Tạo session với User-Agent giống trình duyệt
                        session = requests.Session()
                        session.headers.update({
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
                        })
                        response = session.get(product_url, timeout=30)
                        soup = BeautifulSoup(response.text, 'html.parser')
                        
                        # Thử tìm tên sản phẩm từ các selector thường gặp
                        name_selectors = [
                            'h1.product-title', '.product-name h1', '.page-title h1', 
                            '.product__name', '.product-detail h1', '.product-title'
                        ]
                        
                        for selector in name_selectors:
                            name_elem = soup.select_one(selector)
                            if name_elem:
                                product_name = name_elem.text.strip()
                                print(f"  Tìm thấy tên sản phẩm: {product_name}")
                                break
                    except Exception as e:
                        print(f"  Không thể trích xuất tên sản phẩm từ trang: {str(e)}")
            
            # Bỏ qua nếu không tìm thấy URL sản phẩm
            if not product_url:
                print(f"  Không thể tạo URL cho mã sản phẩm: {product_code}")
                with result_lock:
                    skipped_codes.append(product_code)
                    product_results[product_code] = {
                        'product_name': product_name,
                        'success': False,
                        'message': 'Không thể tạo URL sản phẩm',
                        'documents': [],
                        'failed_documents': [],
                        'successful_documents': 0,
                        'total_documents': 0,
                        'failed_documents_count': 0
                    }
                return None
            
            # Tạo thư mục cho sản phẩm - tất cả tài liệu của cùng một sản phẩm sẽ được lưu trong cùng một thư mục
            product_folder = os.path.join(output_folder, product_code)
            os.makedirs(product_folder, exist_ok=True)
            
            # Trích xuất liên kết tài liệu
            document_links = extract_product_documents(product_url)
            
            if not document_links:
                print(f"  Không tìm thấy tài liệu nào cho {product_code}")
                with result_lock:
                    failed_products += 1
                    product_results[product_code] = {
                        'product_name': product_name,
                        'success': False,
                        'message': 'Không tìm thấy tài liệu',
                        'documents': [],
                        'failed_documents': [],
                        'successful_documents': 0,
                        'total_documents': 0,
                        'failed_documents_count': 0
                    }
                return None
            
            print(f"  Tìm thấy {len(document_links)} tài liệu")
            
            # Tải tài liệu sử dụng đa luồng trong mỗi sản phẩm
            documents = []
            failed_documents = []
            
            # Sử dụng ThreadPoolExecutor để tải các tài liệu của sản phẩm
            with ThreadPoolExecutor(max_workers=5) as doc_executor:
                doc_futures = {doc_executor.submit(download_product_document, doc_link, product_folder): doc_link for doc_link in document_links}
                
                for future in as_completed(doc_futures):
                    doc_link = doc_futures[future]
                    try:
                        result = future.result()
                        if result['success']:
                            documents.append(result)
                            print(f"    ✓ Đã tải: {result['path']}")
                        else:
                            failed_documents.append(result)
                            print(f"    ✗ Lỗi: {result['error']}")
                    except Exception as e:
                        print(f"    ✗ Lỗi khi tải tài liệu {doc_link['url']}: {str(e)}")
                        failed_documents.append({
                            'success': False,
                            'error': str(e),
                            'url': doc_link.get('url', ''),
                            'name': doc_link.get('name', '')
                        })
            
            # Cập nhật số liệu thống kê
            successful_documents = len(documents)
            failed_documents_count = len(failed_documents)
            
            product_result = {
                'product_name': product_name,
                'success': successful_documents > 0,
                'message': f"Tải thành công {successful_documents}/{len(document_links)} tài liệu" if successful_documents > 0 
                          else "Không tài liệu nào được tải thành công",
                'documents': documents,
                'failed_documents': failed_documents,
                'successful_documents': successful_documents,
                'total_documents': len(document_links),
                'failed_documents_count': failed_documents_count
            }
            
            # Cập nhật kết quả tổng hợp
            with result_lock:
                if successful_documents > 0:
                    successful_products += 1
                else:
                    failed_products += 1
                product_results[product_code] = product_result
            
            # In kết quả tạm thời
            print(f"  Kết quả: {successful_documents} thành công, {failed_documents_count} thất bại")
            return product_result
            
        except Exception as e:
            print(f"  Lỗi khi tải tài liệu cho {item}: {str(e)}")
            traceback.print_exc()
            
            with result_lock:
                failed_products += 1
                # Đảm bảo chúng ta có product_code ngay cả khi xảy ra lỗi sớm
                if 'product_code' not in locals():
                    product_code = f"unknown_{index}"
                if 'product_name' not in locals():
                    product_name = f"Unknown Product {index}"
                
                product_results[product_code] = {
                    'product_name': product_name,
                    'success': False,
                    'message': f"Lỗi: {str(e)}",
                    'documents': [],
                    'failed_documents': [],
                    'successful_documents': 0,
                    'total_documents': 0,
                    'failed_documents_count': 0
                }
            return None
    
    # Sử dụng ThreadPoolExecutor để xử lý nhiều sản phẩm cùng lúc
    max_workers = min(10, total_products)  # Tối đa 10 luồng
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for i, item in enumerate(product_codes_or_urls, 1):
            futures.append(executor.submit(process_product, item, i))
        
        # Đợi tất cả hoàn thành
        for future in futures:
            try:
                future.result()
            except Exception as e:
                print(f"Lỗi không mong muốn trong quá trình xử lý: {str(e)}")
    
    # Tạo báo cáo tổng hợp
    print("\n--- Kết quả tải tài liệu ---")
    print(f"Tổng số sản phẩm: {total_products}")
    print(f"Thành công: {successful_products}")
    print(f"Thất bại: {failed_products}")
    print(f"Bỏ qua: {len(skipped_codes)}")
    
    # Tạo cấu trúc báo cáo
    report = {
        'total_products': total_products,
        'successful_products': successful_products,
        'failed_products': failed_products,
        'product_results': product_results,
        'skipped_codes': skipped_codes,
        'output_folder': output_folder
    }
    
    # Tạo báo cáo Excel
    try:
        create_documents_report(report, output_folder)
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo Excel: {str(e)}")
    
    return report

def create_documents_report(report_data, output_folder):
    """
    Tạo báo cáo Excel cho quá trình tải tài liệu
    
    Args:
        report_data (dict): Dữ liệu báo cáo
        output_folder (str): Thư mục đầu ra
    """
    report_path = os.path.join(output_folder, 'documents_report.xlsx')
    
    try:
        # Tạo DataFrame đơn giản hợp nhất tất cả thông tin
        all_data = []
        
        # Thêm hàng tiêu đề với thông tin tổng quan
        all_data.append({
            'Loại': 'THÔNG TIN TỔNG QUAN',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': ''
        })
        
        # Thêm thông tin tổng quan
        all_data.append({
            'Loại': 'Tổng số sản phẩm',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': str(report_data['total_products'])
        })
        
        all_data.append({
            'Loại': 'Sản phẩm thành công',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': str(report_data['successful_products'])
        })
        
        all_data.append({
            'Loại': 'Sản phẩm thất bại',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': str(report_data['failed_products'])
        })
        
        all_data.append({
            'Loại': 'Mã bị bỏ qua',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': str(len(report_data['skipped_codes']))
        })
        
        all_data.append({
            'Loại': 'Thư mục đầu ra',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': report_data['output_folder']
        })
        
        # Thêm một hàng trống để phân tách
        all_data.append({
            'Loại': '',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': ''
        })
        
        # Thêm tiêu đề cho phần chi tiết sản phẩm và tài liệu
        all_data.append({
            'Loại': 'CHI TIẾT SẢN PHẨM VÀ TÀI LIỆU',
            'Tên sản phẩm': '',
            'Mã sản phẩm': '',
            'Tài liệu đã tải': '',
            'Trạng thái': ''
        })
        
        # Thêm dữ liệu từng sản phẩm và tài liệu
        for product_code, result in report_data['product_results'].items():
            product_name = result.get('product_name', '')
            success_docs = result.get('documents', [])
            
            if not success_docs:
                # Nếu không có tài liệu thành công, thêm hàng với thông tin sản phẩm nhưng không có tài liệu
                all_data.append({
                    'Loại': 'Sản phẩm',
                    'Tên sản phẩm': product_name,
                    'Mã sản phẩm': product_code,
                    'Tài liệu đã tải': 'Không có tài liệu',
                    'Trạng thái': result.get('message', 'Không có tài liệu')
                })
            else:
                # Nếu có tài liệu, kết hợp tất cả tài liệu của sản phẩm vào một hàng
                document_names = []
                document_paths = []
                
                for doc in success_docs:
                    document_names.append(doc.get('name', ''))
                    document_paths.append(doc.get('path', ''))
                
                # Nối tất cả tài liệu thành một chuỗi, ngăn cách bởi dấu ";"
                all_docs = "; ".join([f"{name} ({os.path.basename(path)})" for name, path in zip(document_names, document_paths)])
                
                all_data.append({
                    'Loại': 'Sản phẩm',
                    'Tên sản phẩm': product_name,
                    'Mã sản phẩm': product_code,
                    'Tài liệu đã tải': all_docs,
                    'Trạng thái': f"Thành công ({len(success_docs)} tài liệu)"
                })
        
        # Tạo DataFrame từ tất cả dữ liệu
        all_df = pd.DataFrame(all_data)
        
        # Lưu vào file Excel với một sheet duy nhất
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            all_df.to_excel(writer, sheet_name='Báo cáo tài liệu', index=False)
            
            # Tùy chỉnh độ rộng cột
            worksheet = writer.sheets['Báo cáo tài liệu']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 100)  # Giới hạn độ rộng tối đa là 100
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Thêm định dạng cho các tiêu đề
            bold_font = openpyxl.styles.Font(bold=True)
            title_fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Áp dụng định dạng cho tiêu đề cột
            for cell in worksheet[1]:
                cell.font = bold_font
                cell.fill = title_fill
            
            # Áp dụng định dạng cho tiêu đề các phần
            for row_idx, row in enumerate(worksheet.iter_rows(), 1):
                if row[0].value == 'THÔNG TIN TỔNG QUAN' or row[0].value == 'CHI TIẾT SẢN PHẨM VÀ TÀI LIỆU':
                    for cell in row:
                        cell.font = bold_font
                        cell.fill = title_fill
            
            # Thiết lập wrap text cho cột tài liệu
            for row in worksheet.iter_rows(min_row=2):
                doc_cell = row[3]  # Cột 'Tài liệu đã tải' là cột thứ 4 (index 3)
                doc_cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')
        
        print(f"Đã tạo báo cáo Excel tại: {report_path}")
        return True
    
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo Excel: {str(e)}")
        traceback.print_exc()
        return False

def extract_product_code_from_url(url):
    """
    Trích xuất mã sản phẩm từ URL sản phẩm BAA.vn
    
    Args:
        url (str): URL sản phẩm
        
    Returns:
        str: Mã sản phẩm hoặc None nếu không tìm thấy
    """
    try:
        # Mẫu 1: URL có dạng example.com/path/product-name_12345
        pattern1 = r'[^/]+_(\d+)/?$'
        match1 = re.search(pattern1, url)
        if match1:
            return match1.group(1)
            
        # Mẫu 2: URL có dạng example.com/path/MODEL-NUMBER
        parts = url.strip('/').split('/')
        if parts:
            last_part = parts[-1]
            
            # Tìm phần cuối cùng sau dấu gạch ngang
            code_match = re.search(r'.*[-_]([a-zA-Z0-9-]+)$', last_part)
            if code_match:
                return code_match.group(1)
                
            # Nếu không có dấu gạch ngang, lấy phần cuối
            return last_part
            
        return None
    except Exception as e:
        print(f"Lỗi khi trích xuất mã sản phẩm: {str(e)}")
        return None

def standardize_product_code(code):
    """
    Chuyển mã sản phẩm thành chữ hoa.
    
    Args:
        code (str): Mã sản phẩm cần chuyển đổi
        
    Returns:
        str: Mã sản phẩm đã được chuyển thành chữ hoa
    """
    if not code or pd.isna(code):
        return ''
        
    # Chuyển thành chuỗi và viết hoa
    return str(code).upper()

def get_random_headers():
    """Tạo các header ngẫu nhiên để tránh bị chặn"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Connection': 'keep-alive',
        'Referer': 'https://google.com'
    }
    return headers

def download_baa_product_images(url_file, output_folder):
    """Tải ảnh sản phẩm từ BAA.vn dựa trên file URL sử dụng đa luồng"""
    results = []
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # Đọc danh sách URL từ file
    with open(url_file, 'r', encoding='utf-8') as f:
        urls = [line.strip() for line in f.readlines() if line.strip()]
    
    total_urls = len(urls)
    print(f"Tổng số URL cần xử lý: {total_urls}")
    
    # Đếm số URL đã xử lý và số lượng thành công
    processed_count = 0
    success_count = 0
    
    # Khóa đồng bộ hóa
    result_lock = threading.Lock()
    counter_lock = threading.Lock()
    
    # Hàm xử lý riêng cho mỗi URL
    def process_url(url, index):
        nonlocal processed_count, success_count
        
        print(f"[{index}/{total_urls}] Đang xử lý: {url}")
        try:
            # Tải nội dung trang
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            # Thử tải trang nhiều lần nếu có lỗi
            for attempt in range(3):
                try:
                    response = requests.get(url, headers=headers, timeout=30)
                    response.raise_for_status()
                    break
                except (requests.RequestException, Exception) as e:
                    print(f"  > Lỗi khi tải trang ({attempt+1}/3): {str(e)}")
                    if attempt == 2:  # Lần thử cuối cùng
                        raise
                    time.sleep(2)
            
            # Tạo soup
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Trích xuất mã sản phẩm từ URL
            product_code = ""
            try:
                # Thử trích xuất mã sản phẩm từ trang web
                product_code_element = soup.select_one('span.product__symbol__value')
                if product_code_element:
                    product_code = product_code_element.get_text(strip=True)
                    print(f"  > Trích xuất mã sản phẩm từ span.product__symbol__value: {product_code}")
                else:
                    # Hoặc trích xuất từ URL nếu không tìm thấy trong trang
                    if "_" in url:
                        product_code = url.split('_')[-1].split('/')[-1]
                    else:
                        product_code = url.split('/')[-1]
                    print(f"  > Trích xuất mã sản phẩm từ URL: {product_code}")
            except Exception as e:
                print(f"  > Lỗi khi trích xuất mã sản phẩm: {str(e)}")
                # Sử dụng URL làm mã sản phẩm trong trường hợp lỗi
                product_code = url.split('/')[-1]
            
            print(f"  > Mã sản phẩm: {product_code}")
            
            # Tìm ảnh sản phẩm
            image_url = None
            
            # Thử tìm ảnh từ div.modal-body-image với các class cụ thể
            print(f"  > Đang tìm div.modal-body-image với các class")
            modal_div = soup.select_one('div.modal-body-image.position-absolute.w-100.h-100.d-none.active')
            if modal_div:
                print(f"  > Tìm thấy div.modal-body-image.position-absolute.w-100.h-100.d-none.active")
                # Trích xuất URL ảnh từ thuộc tính style
                style = modal_div.get('style', '')
                if 'background-image' in style:
                    image_url_match = re.search(r'url\([\'"]?(.*?)[\'"]?\)', style)
                    if image_url_match:
                        image_url = image_url_match.group(1)
                        print(f"  > Trích xuất được URL ảnh từ style: {image_url}")
            
            # Nếu không tìm được từ selector cụ thể, thử tìm bất kỳ div.modal-body-image nào
            if not image_url:
                print(f"  > Không tìm thấy div với class đầy đủ, thử tìm bất kỳ div.modal-body-image nào")
                modal_divs = soup.select('div.modal-body-image')
                for div in modal_divs:
                    style = div.get('style', '')
                    if 'background-image' in style:
                        image_url_match = re.search(r'url\([\'"]?(.*?)[\'"]?\)', style)
                        if image_url_match:
                            image_url = image_url_match.group(1)
                            print(f"  > Trích xuất được URL ảnh từ div.modal-body-image: {image_url}")
                            break
            
            # Nếu không tìm thấy từ div.modal-body-image, thử tìm từ img.btn-image-view-360
            if not image_url:
                print(f"  > Không tìm thấy ảnh từ div.modal-body-image, thử tìm từ img.btn-image-view-360")
                img_element = soup.select_one('img.btn-image-view-360')
                if img_element:
                    image_url = img_element.get('src')
                    # Thay đổi kích thước ảnh (từ nhỏ sang lớn)
                    image_url = image_url.replace('/s/', '/l/')
                    print(f"  > Tìm thấy ảnh từ img.btn-image-view-360: {image_url}")
            
            # Thử tìm từ các img khác
            if not image_url:
                print(f"  > Không tìm thấy ảnh từ selector cụ thể, tìm ảnh đầu tiên phù hợp")
                img_elements = soup.select('img.img-fluid')
                for img in img_elements:
                    src = img.get('src')
                    if src and ('product' in src.lower() or 'prod_img' in src.lower()):
                        image_url = src
                        # Thay đổi kích thước ảnh (từ nhỏ sang lớn)
                        image_url = image_url.replace('/s/', '/l/')
                        print(f"  > Tìm thấy ảnh từ img.img-fluid: {image_url}")
                        break
            
            # Nếu vẫn không tìm thấy, thử tìm bất kỳ ảnh nào
            if not image_url:
                print(f"  > Không tìm thấy ảnh từ các selector cụ thể, tìm bất kỳ ảnh nào")
                img_elements = soup.select('img')
                for img in img_elements:
                    src = img.get('src')
                    if src and ('product' in src.lower() or 'prod_img' in src.lower() or '/img/' in src.lower()):
                        image_url = src
                        # Thay đổi kích thước ảnh (từ nhỏ sang lớn)
                        image_url = image_url.replace('/s/', '/l/')
                        print(f"  > Tìm thấy ảnh: {image_url}")
                        break
            
            result = {
                'URL': url,
                'Mã sản phẩm': product_code,
                'Trạng thái': '',
                'Ảnh sản phẩm': '',
                'Lỗi': ''
            }
            
            if image_url:
                # Đảm bảo URL ảnh là đầy đủ
                if not image_url.startswith('http'):
                    # Xử lý trường hợp URL bắt đầu bằng //
                    if image_url.startswith('//'):
                        image_url = 'https:' + image_url
                    else:
                        # Xử lý URL tương đối
                        parsed_url = urlparse(url)
                        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
                        image_url = urljoin(base_url, image_url)
                
                try:
                    # Tạo tên file từ mã sản phẩm
                    image_filename = f"{product_code}.webp"
                    image_path = os.path.join(output_folder, image_filename)
                    
                    # Tải ảnh
                    print(f"  > Đang tải ảnh từ: {image_url}")
                    for attempt in range(3):
                        try:
                            img_response = requests.get(image_url, headers=headers, timeout=30)
                            img_response.raise_for_status()
                            
                            # Chuyển đổi sang WebP và lưu
                            img = Image.open(BytesIO(img_response.content))
                            img.save(image_path, 'WEBP', quality=90)
                            
                            print(f"  > Đã lưu ảnh: {image_path}")
                            result['Trạng thái'] = 'Thành công'
                            result['Ảnh sản phẩm'] = image_filename
                            
                            with counter_lock:
                                success_count += 1
                            break
                        except (requests.RequestException, Exception) as e:
                            print(f"  > Lỗi khi tải ảnh ({attempt+1}/3): {str(e)}")
                            if attempt == 2:  # Lần thử cuối cùng
                                result['Trạng thái'] = 'Lỗi'
                                result['Lỗi'] = f"Không thể tải ảnh: {str(e)}"
                            time.sleep(2)
                except Exception as e:
                    print(f"  > Lỗi khi xử lý ảnh: {str(e)}")
                    result['Trạng thái'] = 'Lỗi'
                    result['Lỗi'] = f"Lỗi xử lý ảnh: {str(e)}"
            else:
                print(f"  > Không tìm thấy ảnh sản phẩm")
                result['Trạng thái'] = 'Lỗi'
                result['Lỗi'] = 'Không tìm thấy ảnh sản phẩm'
            
            with result_lock:
                results.append(result)
            
            with counter_lock:
                processed_count += 1
            
            return result
            
        except Exception as e:
            print(f"  > Lỗi khi xử lý URL {url}: {str(e)}")
            error_result = {
                'URL': url,
                'Mã sản phẩm': 'Không xác định',
                'Trạng thái': 'Lỗi',
                'Ảnh sản phẩm': '',
                'Lỗi': str(e)
            }
            
            with result_lock:
                results.append(error_result)
            
            with counter_lock:
                processed_count += 1
                
            return error_result
    
    # Sử dụng ThreadPoolExecutor để xử lý đa luồng
    max_workers = min(10, total_urls)  # Tối đa 10 luồng
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_url = {executor.submit(process_url, url, i+1): url for i, url in enumerate(urls)}
        
        # Xử lý kết quả khi hoàn thành
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            try:
                future.result()
            except Exception as e:
                print(f"Lỗi không mong muốn khi xử lý {url}: {str(e)}")
    
    # Tạo báo cáo Excel
    report_file = os.path.join(output_folder, 'baa_images_report.xlsx')
    create_image_report(results, report_file)
    print(f"Đã tạo báo cáo: {report_file}")
    
    # Tạo file ZIP
    zip_file = f"{output_folder}.zip"
    create_zip_from_folder(output_folder, zip_file)
    print(f"Đã tạo file ZIP: {zip_file}")
    
    # Trả về đường dẫn đến file ZIP
    return zip_file

def resize_image_to_square(image, size=800):
    """
    Thay đổi kích thước ảnh thành hình vuông có kích thước được chỉ định
    
    Args:
        image (PIL.Image): Đối tượng ảnh Pillow
        size (int): Kích thước mong muốn cho ảnh vuông
        
    Returns:
        PIL.Image: Ảnh đã được thay đổi kích thước thành hình vuông
    """
    # Lấy kích thước ảnh gốc
    width, height = image.size
    
    # Tạo ảnh vuông mới với nền trắng
    square_img = Image.new('RGBA' if image.mode == 'RGBA' else 'RGB', (size, size), (255, 255, 255))
    
    # Tính toán tỷ lệ để giữ nguyên tỷ lệ khung hình gốc
    if width > height:
        new_width = size
        new_height = int(height * size / width)
        resized_img = image.resize((new_width, new_height), Image.LANCZOS)
        # Căn giữa theo chiều dọc
        paste_y = (size - new_height) // 2
        square_img.paste(resized_img, (0, paste_y))
    else:
        new_height = size
        new_width = int(width * size / height)
        resized_img = image.resize((new_width, new_height), Image.LANCZOS)
        # Căn giữa theo chiều ngang
        paste_x = (size - new_width) // 2
        square_img.paste(resized_img, (paste_x, 0))
    
    return square_img

def download_baa_product_images_fixed(product_urls, output_folder=None, create_report=True):
    try:
        # Chuyển đổi input thành list nếu nhận được string
        if isinstance(product_urls, str):
            product_urls = [product_urls]
        
        # Tạo thư mục output nếu không được cung cấp
        if not output_folder:
            # Tạo output folder trong thư mục hiện tại
            output_folder = os.path.join(os.getcwd(), "output_images")
        
        # Đảm bảo thư mục output tồn tại
        os.makedirs(output_folder, exist_ok=True)
        
        print(f"Thư mục lưu ảnh: {output_folder}")
        
        # Khởi tạo kết quả trả về
        results = {
            'total': len(product_urls),
            'success': 0,
            'failed': 0,
            'image_paths': [],
            'report_data': [],
            'report_file': None
        }
        
        # Khóa đồng bộ hóa kết quả
        result_lock = threading.Lock()
        
        # Hàm xử lý cho mỗi URL sản phẩm
        def process_product_url(url, i):
            if len(product_urls) > 1:  # Chỉ in tiến độ khi có nhiều hơn 1 URL
                print(f"\n[{i}/{len(product_urls)}] Đang xử lý URL: {url}")
            
            # Khởi tạo dữ liệu báo cáo cho URL này
            report_item = {
                'STT': i,
                'URL': url,
                'Mã sản phẩm': '',
                'Đường dẫn ảnh': '',
                'Kích thước ảnh': '',
                'Trạng thái': 'Thất bại',
                'Lý do lỗi': '',
                'Thời gian xử lý': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            images_found = False
            
            try:
                # Lấy nội dung trang sản phẩm
                headers = get_random_headers()
                response = requests.get(url, headers=headers, timeout=30)
                response.raise_for_status()
                
                # Sử dụng hàm trích xuất URL ảnh và mã sản phẩm từ HTML
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Trích xuất mã sản phẩm
                product_code = None
                product_symbol = soup.select_one('span.product__symbol__value')
                if product_symbol and product_symbol.text.strip():
                    product_code = product_symbol.text.strip()
                    if len(product_urls) == 1:  # Chỉ in chi tiết khi xử lý 1 URL
                        print(f"  ✓ Tìm thấy mã sản phẩm từ span.product__symbol__value: {product_code}")
                
                # Ưu tiên sử dụng mã sản phẩm từ HTML
                if product_code:
                    original_product_code = product_code
                    if len(product_urls) == 1:
                        print(f"  ✓ Sử dụng mã sản phẩm từ HTML: {product_code}")
                else:
                    # Nếu không tìm được mã từ HTML, thử trích xuất từ URL
                    product_code = extract_product_code_from_url(url)
                    original_product_code = product_code
                    if len(product_urls) == 1:
                        print(f"  ✓ Sử dụng mã sản phẩm từ URL: {product_code}")
                
                # Chuẩn hóa mã sản phẩm cho tên file - thay thế các ký tự đặc biệt
                if product_code:
                    # Lưu lại mã sản phẩm gốc cho báo cáo
                    report_product_code = product_code
                    # Chuẩn hóa mã sản phẩm để sử dụng làm tên file
                    from app.utils import standardize_filename
                    product_code = standardize_filename(product_code)
                else:
                    report_product_code = f"unknown_product_{i}"
                    product_code = f"unknown_product_{i}"
                
                # Ghi nhận mã sản phẩm vào báo cáo
                report_item['Mã sản phẩm'] = report_product_code
                
                # PHƯƠNG PHÁP 1: Ưu tiên tìm ảnh 800px từ div.modal-body-image.active
                img_url = None
                
                # Tìm tất cả image tags trong modal-body-image.active với class w-100 h-100
                modal_active_img = soup.select('div.modal-body-image.active img.w-100.h-100')
                if modal_active_img:
                    for img in modal_active_img:
                        if img.get('src'):
                            src = img.get('src')
                            if src and 'LOGO-BAA' not in src:
                                # Đảm bảo là ảnh kích thước 800
                                if '800' in src:
                                    img_url = src
                                    if len(product_urls) == 1:
                                        print(f"  ✓ Tìm thấy ảnh 800px từ modal-body-image.active: {img_url}")
                                    break
                                else:
                                    # Nếu không có kích thước 800, thử chuyển đổi
                                    pattern = r'/(series|model)/(\d+)/(\d+)/'
                                    match = re.search(pattern, src)
                                    if match:
                                        type_folder = match.group(1)  # series hoặc model
                                        folder_id = match.group(2)    # ID thư mục (ví dụ: 274)
                                        img_url = re.sub(pattern, f'/{type_folder}/{folder_id}/800/', src)
                                        if len(product_urls) == 1:
                                            print(f"  ✓ Chuyển đổi ảnh từ modal-body-image.active sang kích thước 800px: {img_url}")
                                        break
                
                # PHƯƠNG PHÁP 2: Tìm trong div.modal-body__view-image
                if not img_url:
                    modal_view_img = soup.select('div.modal-body__view-image img.w-100.h-100')
                    for img in modal_view_img:
                        if img.get('src'):
                            src = img.get('src')
                            if src and 'LOGO-BAA' not in src:
                                # Ưu tiên ảnh 800px
                                if '800' in src:
                                    img_url = src
                                    if len(product_urls) == 1:
                                        print(f"  ✓ Tìm thấy ảnh 800px từ div.modal-body__view-image: {img_url}")
                                    break
                                else:
                                    # Nếu không có kích thước 800, thử chuyển đổi
                                    pattern = r'/(series|model)/(\d+)/(\d+)/'
                                    match = re.search(pattern, src)
                                    if match:
                                        type_folder = match.group(1)  # series hoặc model
                                        folder_id = match.group(2)    # ID thư mục (ví dụ: 274)
                                        img_url = re.sub(pattern, f'/{type_folder}/{folder_id}/800/', src)
                                        if len(product_urls) == 1:
                                            print(f"  ✓ Chuyển đổi ảnh từ div.modal-body__view-image sang kích thước 800px: {img_url}")
                                        break
                
                # PHƯƠNG PHÁP 3: Nếu không tìm thấy, lấy từ img.btn-image-view-360 và chuyển sang 800
                if not img_url:
                    main_image = soup.select_one('img.btn-image-view-360')
                    if main_image and main_image.get('src'):
                        src = main_image.get('src')
                        if src and 'LOGO-BAA' not in src:
                            # Chuyển từ kích thước nhỏ (thường là 300) sang 800
                            pattern = r'/(series|model)/(\d+)/(\d+)/'
                            match = re.search(pattern, src)
                            if match:
                                type_folder = match.group(1)  # series hoặc model
                                folder_id = match.group(2)    # ID thư mục (ví dụ: 274)
                                img_url = re.sub(pattern, f'/{type_folder}/{folder_id}/800/', src)
                                if len(product_urls) == 1:
                                    print(f"  ✓ Chuyển đổi ảnh từ img.btn-image-view-360 sang kích thước 800px: {img_url}")
                            else:
                                img_url = src
                                if len(product_urls) == 1:
                                    print(f"  ✓ Tìm thấy ảnh từ img.btn-image-view-360: {img_url}")
                
                # PHƯƠNG PHÁP 4: Nếu vẫn không tìm thấy, thử og:image
                if not img_url:
                    og_image = soup.find('meta', property='og:image')
                    if og_image:
                        img_url = og_image.get('content', '')
                        if img_url and 'LOGO-BAA' not in img_url:
                            # Chuyển từ kích thước nhỏ (thường là 300) sang 800
                            pattern = r'/(series|model)/(\d+)/(\d+)/'
                            match = re.search(pattern, img_url)
                            if match:
                                type_folder = match.group(1)  # series hoặc model
                                folder_id = match.group(2)    # ID thư mục (ví dụ: 274)
                                img_url = re.sub(pattern, f'/{type_folder}/{folder_id}/800/', img_url)
                                if len(product_urls) == 1:
                                    print(f"  ✓ Chuyển đổi ảnh từ og:image sang kích thước 800px: {img_url}")
                        else:
                            if len(product_urls) == 1:
                                print(f"  ✓ Tìm thấy ảnh từ og:image: {img_url}")
                
                # Nếu không tìm thấy ảnh nào, thông báo thất bại
                if not img_url:
                    with result_lock:
                        results['failed'] += 1
                    report_item['Lý do lỗi'] = 'Không tìm thấy ảnh sản phẩm'
                    with result_lock:
                        results['report_data'].append(report_item)
                    if len(product_urls) == 1:
                        print(f"  ✗ Không tìm thấy ảnh cho URL: {url}")
                    return report_item
                
                # Chuẩn hóa URL ảnh
                base_url = '{uri.scheme}://{uri.netloc}'.format(uri=urlparse(url))
                
                # Chuyển URL tương đối thành tuyệt đối
                if img_url.startswith('//'):
                    img_url = 'https:' + img_url
                elif not img_url.startswith(('http://', 'https://')):
                    img_url = urljoin(base_url, img_url)
                
                if len(product_urls) == 1:
                    print(f"  → Lưu ảnh vào thư mục: {output_folder}")
                
                # Thử tải ảnh - giữ nguyên định dạng ảnh
                try:
                    # Tạo tên file ảnh dựa trên mã sản phẩm 
                    img_filename = f"{product_code}.webp"
                    img_path = os.path.join(output_folder, img_filename)
                    
                    if len(product_urls) == 1:
                        print(f"  → Đang tải ảnh: {img_url}")
                    
                    # Tải ảnh
                    img_response = requests.get(img_url, headers=headers, timeout=15)
                    img_response.raise_for_status()
                    
                    # Kiểm tra MIME type
                    content_type = img_response.headers.get('Content-Type', '')
                    if not content_type.startswith('image/'):
                        raise ValueError(f"Không phải file ảnh: {content_type}")
                    
                    # Xử lý ảnh
                    img = Image.open(BytesIO(img_response.content))
                    img = img.convert("RGB")
                    
                    # Resize ảnh về kích thước vuông 800x800
                    img = resize_image_to_square(img, 800)
                    
                    # Lưu thông tin kích thước ảnh
                    img_size = f"{img.width}x{img.height}"
                    
                    # Lưu ảnh dưới dạng WebP với chất lượng cao
                    img.save(img_path, 'WEBP', quality=95)
                    
                    # Cập nhật kết quả
                    with result_lock:
                        results['success'] += 1
                        results['image_paths'].append(img_path)
                    report_item['Trạng thái'] = 'Thành công'
                    report_item['Đường dẫn ảnh'] = img_path
                    report_item['Kích thước ảnh'] = img_size
                    
                    if len(product_urls) == 1:
                        print(f"  ✓ Đã lưu: {img_filename} ({img_size})")
                    
                except requests.exceptions.HTTPError as e:
                    # Nếu lỗi 404 với ảnh 800px, thử với 300px
                    if '404' in str(e) and '800' in img_url:
                        try:
                            # Thử lại với kích thước 300px
                            img_url_300 = re.sub(r'/800/', '/300/', img_url)
                            if len(product_urls) == 1:
                                print(f"  → Thử lại với ảnh kích thước 300px: {img_url_300}")
                            
                            # Tải ảnh kích thước 300px
                            img_response = requests.get(img_url_300, headers=headers, timeout=15)
                            img_response.raise_for_status()
                            
                            # Kiểm tra MIME type
                            content_type = img_response.headers.get('Content-Type', '')
                            if not content_type.startswith('image/'):
                                raise ValueError(f"Không phải file ảnh: {content_type}")
                            
                            # Xử lý ảnh
                            img = Image.open(BytesIO(img_response.content))
                            img = img.convert("RGB")
                            
                            # Lưu thông tin kích thước ảnh
                            img_size = f"{img.width}x{img.height}"
                            
                            # Lưu ảnh dưới dạng WebP với chất lượng cao
                            img.save(img_path, 'WEBP', quality=95)
                            
                            # Cập nhật kết quả
                            with result_lock:
                                results['success'] += 1
                                results['image_paths'].append(img_path)
                            report_item['Trạng thái'] = 'Thành công'
                            report_item['Đường dẫn ảnh'] = img_path
                            report_item['Kích thước ảnh'] = img_size
                            
                            if len(product_urls) == 1:
                                print(f"  ✓ Đã lưu ảnh kích thước 300px: {img_filename} ({img_size})")
                        except Exception as inner_e:
                            with result_lock:
                                results['failed'] += 1
                            report_item['Lý do lỗi'] = f"Lỗi khi tải ảnh: {str(e)} và {str(inner_e)}"
                    else:
                        with result_lock:
                            results['failed'] += 1
                        report_item['Lý do lỗi'] = f"Lỗi HTTP: {str(e)}"
                
                except Exception as e:
                    if len(product_urls) == 1:
                        print(f"  ✗ Lỗi khi tải ảnh: {str(e)}")
                    with result_lock:
                        results['failed'] += 1
                    report_item['Lý do lỗi'] = f"Lỗi khi tải ảnh: {str(e)}"
                
                # Thêm vào dữ liệu báo cáo
                with result_lock:
                    results['report_data'].append(report_item)
                
                # Nghỉ giữa các yêu cầu để tránh bị chặn
                time.sleep(0.5)
                
                return report_item
                
            except Exception as e:
                with result_lock:
                    results['failed'] += 1
                report_item['Lý do lỗi'] = f"Lỗi: {str(e)}"
                with result_lock:
                    results['report_data'].append(report_item)
                if len(product_urls) == 1:
                    print(f"Lỗi khi xử lý URL {url}: {str(e)}")
                traceback.print_exc()
                return report_item
        
        # Xử lý đa luồng
        max_workers = min(10, len(product_urls))  # Tối đa 10 luồng
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(process_product_url, url, i+1): url for i, url in enumerate(product_urls)}
            
            for future in as_completed(futures):
                url = futures[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"Lỗi không xử lý được: {str(e)}")
        
        # Tạo báo cáo Excel chỉ khi được yêu cầu
        if create_report and len(product_urls) > 1:  # Chỉ tạo báo cáo khi xử lý nhiều URL
            try:
                print("\nĐang tạo báo cáo Excel...")
                
                # Tạo DataFrame từ dữ liệu báo cáo
                df = pd.DataFrame(results['report_data'])
                
                # Tạo đường dẫn file báo cáo
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                report_filename = f"baa_image_report_{timestamp}.xlsx"
                report_path = os.path.join(output_folder, report_filename)
                
                # Lưu vào file Excel một cách an toàn
                with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Báo cáo tải ảnh')
                    
                    # Lấy workbook và worksheet
                    workbook = writer.book
                    sheet = writer.sheets['Báo cáo tải ảnh']
                
                # Định dạng các cột
                for idx, col in enumerate(df.columns, 1):
                    # Đặt độ rộng cột
                    column_letter = openpyxl.utils.get_column_letter(idx)
                    if col == 'URL' or col == 'Đường dẫn ảnh':
                        sheet.column_dimensions[column_letter].width = 50
                    elif col == 'Lý do lỗi':
                        sheet.column_dimensions[column_letter].width = 30
                    else:
                        sheet.column_dimensions[column_letter].width = 20
                        
                    # Định dạng tiêu đề
                    header_cell = sheet.cell(row=1, column=idx)
                    header_cell.font = openpyxl.styles.Font(bold=True)
                    header_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Tạo bảng tóm tắt
                summary_row = sheet.max_row + 3
                sheet.cell(row=summary_row, column=1).value = "Tổng số URL:"
                sheet.cell(row=summary_row, column=2).value = results['total']
                
                sheet.cell(row=summary_row+1, column=1).value = "Tải thành công:"
                sheet.cell(row=summary_row+1, column=2).value = results['success']
                
                sheet.cell(row=summary_row+2, column=1).value = "Tải thất bại:"
                sheet.cell(row=summary_row+2, column=2).value = results['failed']
                
                sheet.cell(row=summary_row+3, column=1).value = "Tổng số ảnh đã tải:"
                sheet.cell(row=summary_row+3, column=2).value = len(results['image_paths'])
                
                # Đặt độ rộng cho cột
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 10
                
                # In đậm tiêu đề tóm tắt
                for i in range(4):
                    cell = sheet.cell(row=summary_row+i, column=1)
                    cell.font = openpyxl.styles.Font(bold=True)
                
                # Thêm đường dẫn file báo cáo vào kết quả
                results['report_file'] = report_path
                print(f"Đã tạo báo cáo Excel: {report_path}")
                
            except Exception as e:
                print(f"Lỗi khi tạo báo cáo Excel: {str(e)}")
                traceback.print_exc()
        
        # Trả về kết quả
        return results
        
    except Exception as e:
        print(f"Lỗi trong quá trình tải ảnh: {str(e)}")
        traceback.print_exc()
        return {
            'total': len(product_urls) if isinstance(product_urls, list) else 1, 
            'success': 0,
            'failed': len(product_urls) if isinstance(product_urls, list) else 1,
            'image_paths': [],
            'report_data': [],
            'report_file': None
        }

def extract_baa_image_url_from_html(html_content):
    """
    Trích xuất URL ảnh sản phẩm từ HTML
    """
    try:
        # Tạo đối tượng BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Trích xuất URL ảnh và mã sản phẩm
        image_urls = []
        product_code = None
        
        # Kiểm tra có phần tử image-view-360 không
        img_element = soup.select_one('img.btn-image-view-360')
        if img_element and 'src' in img_element.attrs:
            img_url = img_element['src']
            if img_url:
                # Chuyển đổi URL sang kích thước lớn
                full_url = convert_to_large_image_url(img_url)
                image_urls.append(full_url)
                print(f"Đã tìm thấy ảnh chính: {full_url}")
        
        # Tìm mã sản phẩm
        product_code_element = soup.select_one('.product__symbol__value')
        if product_code_element:
            product_code = product_code_element.text.strip()
            print(f"Mã sản phẩm: {product_code}")
        
        return {
            'image_urls': image_urls,
            'product_code': product_code
        }
        
    except Exception as e:
        print(f"Lỗi khi trích xuất URL ảnh: {str(e)}")
        return {'image_urls': [], 'product_code': None}

def clean_price(price_str):
    """
    Làm sạch giá sản phẩm: 
    - Giữ lại số 
    - Giữ lại đơn vị tiền tệ (đ, ₫, VND)
    - Loại bỏ các ký tự không cần thiết khác
    
    Args:
        price_str (str): Chuỗi giá ban đầu
    
    Returns:
        str: Chuỗi giá đã làm sạch
    """
    if not price_str:
        return ""
    
    # Giữ nguyên chuỗi nếu có đơn vị tiền tệ
    if "₫" in price_str or "đ" in price_str or "VND" in price_str:
        return price_str.strip()
    
    # Thử chuyển về số
    try:
        # Loại bỏ tất cả ký tự không phải số
        price_digits = re.sub(r'[^\d]', '', price_str)
        if price_digits:
            # Định dạng lại số với dấu phân cách hàng nghìn
            price_val = int(price_digits)
            return f"{price_val:,}₫".replace(",", ".")
        return price_str.strip()
    except:
        return price_str.strip()

def extract_product_price(url, index=1):
    """
    Chỉ trích xuất mã sản phẩm và giá từ URL
    
    Args:
        url (str): URL của trang sản phẩm
        index (int): STT của sản phẩm trong danh sách
        
    Returns:
        dict: Thông tin mã và giá sản phẩm
    """
    print(f"Đang trích xuất giá từ {url}")
    
    # Số lần thử tối đa
    max_retries = 3
    current_retry = 0
    
    while current_retry < max_retries:
        try:
            # Tải nội dung trang không giới hạn timeout
            response = requests.get(url, headers=HEADERS)
            response.raise_for_status()
            html_content = response.text
            
            # Parse HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Khởi tạo kết quả
            product_info = {
                'STT': index,
                'URL': url,
                'Mã sản phẩm': "",
                'Giá': ""
            }
            
            # Trích xuất mã sản phẩm
            product_code = ""
            code_element = soup.select_one('.product__symbol__value')
            if code_element:
                product_code = code_element.text.strip()
            
            # Nếu không tìm thấy bằng phương pháp thông thường, thử các phương pháp khác
            if not product_code:
                # Thử các CSS selector khác
                code_selectors = [
                    '.product-sku',
                    '.sku',
                    '[itemprop="sku"]',
                    '.product-id'
                ]
                
                for selector in code_selectors:
                    code_element = soup.select_one(selector)
                    if code_element:
                        product_code = code_element.text.strip()
                        break
                
                # Nếu vẫn không tìm thấy, trích xuất từ URL hoặc tên sản phẩm
                if not product_code:
                    # Trích xuất từ URL
                    url_path = urlparse(url).path
                    path_parts = url_path.split("/")
                    if path_parts:
                        last_part = path_parts[-1]
                        # Loại bỏ các số ID, giữ lại phần mã sản phẩm
                        product_code = re.sub(r'_\d+$', '', last_part)
                    
                    # Trích xuất từ tên sản phẩm
                    if not product_code:
                        product_name = ""
                        name_element = soup.select_one('.product__name, h1.product-title')
                        if name_element:
                            product_name = name_element.text.strip()
                            code_match = re.search(r'([A-Z]{2,3}-[A-Z0-9]{3,5}(?:-[A-Z0-9]{4,5})+)', product_name)
                            if code_match:
                                product_code = code_match.group(1)
            
            if product_code:
                print(f"Mã sản phẩm: {product_code}")
                product_info['Mã sản phẩm'] = product_code
            
            # ------ PHẦN XỬ LÝ GIÁ SẢN PHẨM ------
            # 1. ƯU TIÊN SỐ 1: Trích xuất từ phần tử span.product__price-print có thuộc tính data-root
            price_element_with_data = soup.select_one('span.product__price-print[data-root]')
            if price_element_with_data and 'data-root' in price_element_with_data.attrs:
                data_root = price_element_with_data.get('data-root')
                if data_root:
                    try:
                        price_value = int(data_root)
                        formatted_price = f"{price_value:,}".replace(",", ".")
                        
                        # Tìm đơn vị tiền tệ
                        price_unit = ""
                        
                        # Tìm trong cùng cell hoặc gần kề
                        parent_td = price_element_with_data.find_parent('td')
                        if parent_td:
                            unit_element = parent_td.select_one('span.product__price-unit')
                            if unit_element:
                                price_unit = unit_element.text.strip()
                        
                        # Nếu không tìm thấy, tìm element kế tiếp
                        if not price_unit:
                            next_element = price_element_with_data.find_next_sibling()
                            if next_element and 'product__price-unit' in next_element.get('class', []):
                                price_unit = next_element.text.strip()
                        
                        # Nếu vẫn không tìm thấy
                        if not price_unit:
                            unit_element = soup.select_one('span.product__price-unit')
                            if unit_element:
                                price_unit = unit_element.text.strip()
                            else:
                                price_unit = "₫"  # Mặc định
                        
                        # Định dạng giá cuối cùng
                        product_info['Giá'] = formatted_price + price_unit
                        print(f"Giá từ data-root: {product_info['Giá']}")
                    except ValueError as e:
                        print(f"Lỗi khi xử lý giá từ data-root: {str(e)}")
            
            # 2. NẾU KHÔNG TÌM THẤY: Tìm giá từ các vị trí thông thường
            if not product_info['Giá']:
                price_selectors = [
                    'div.product__card--price span.fw-bold.text-danger.text-start',
                    '.product__card--price span',
                    '.product__card--price .fw-bold',
                    '.product-price',
                    '.product__price-print',
                    '.price-box .price',
                    '.special-price .price',
                    '[data-price-type="finalPrice"] .price'
                ]
                
                for selector in price_selectors:
                    price_element = soup.select_one(selector)
                    if price_element:
                        product_price = price_element.text.strip()
                        
                        # Tìm đơn vị tiền tệ
                        price_unit = ""
                        unit_element = soup.select_one('span.product__price-unit')
                        if unit_element:
                            price_unit = unit_element.text.strip()
                        
                        # Định dạng giá
                        if price_unit and price_unit not in product_price:
                            product_info['Giá'] = product_price + price_unit
                        else:
                            product_info['Giá'] = product_price
                        
                        print(f"Giá sản phẩm: {product_info['Giá']}")
                        break
            
            # 3. KIỂM TRA LẠI: Nếu vẫn không tìm thấy, tìm bất kỳ phần tử nào có thuộc tính data-root
            if not product_info['Giá']:
                elements_with_data_root = soup.select('[data-root]')
                if elements_with_data_root:
                    for element in elements_with_data_root:
                        data_root = element.get('data-root')
                        if data_root:
                            try:
                                price_value = int(data_root)
                                product_info['Giá'] = f"{price_value:,}₫".replace(",", ".")
                                print(f"Giá từ data-root (phương pháp 3): {product_info['Giá']}")
                                break
                            except ValueError:
                                continue
            
            # Làm sạch giá trước khi trả về
            if product_info['Giá']:
                product_info['Giá'] = clean_price(product_info['Giá'])
            
            return product_info
            
        except requests.exceptions.RequestException as e:
            current_retry += 1
            print(f"Lỗi tải trang (lần thử {current_retry}): {str(e)}")
            if current_retry < max_retries:
                print(f"Thử lại trong 3 giây...")
                time.sleep(3)
            else:
                print(f"Đã thử {max_retries} lần, bỏ qua URL {url}")
    
    # Trả về dữ liệu tối thiểu nếu không thể tải trang
    return {
        'STT': index,
        'URL': url,
        'Mã sản phẩm': "",
        'Giá': ""
    }

def extract_product_documents(product_url_or_code):
    """
    Trích xuất danh sách các tài liệu PDF từ trang BAA.vn - phiên bản tối ưu
    
    Hàm này tìm kiếm tài liệu PDF từ website BAA.vn bằng cách:
    1. Truy cập trang sản phẩm
    2. Phân tích DOM để tìm các link PDF trong các phần tử như Link_download và feature__metadata__link--download
    
    Cải tiến:
    - Tăng timeout lên 60 giây
    - Tập trung vào các selector cụ thể trong HTML
    - Xử lý lỗi chi tiết
    """
    # Cache đơn giản để lưu trữ kết quả tìm kiếm
    cache_key = str(product_url_or_code).strip().lower()
    static_cache = getattr(extract_product_documents, 'cache', {})
    if cache_key in static_cache:
        print(f"Lấy kết quả từ cache cho: {product_url_or_code}")
        return static_cache[cache_key]
    
    try:
        # 1. Xác định URL sản phẩm
        product_url = None
        
        # Nếu là URL, sử dụng trực tiếp
        if isinstance(product_url_or_code, str) and product_url_or_code.startswith('http'):
            product_url = product_url_or_code
            print(f"Sử dụng URL trực tiếp: {product_url}")
        else:
            # Nếu là mã sản phẩm, tìm URL tương ứng
            product_url = get_product_url(product_url_or_code)
            if not product_url:
                print(f"Không tìm thấy URL sản phẩm cho mã: {product_url_or_code}")
                static_cache[cache_key] = []
                return []
        
        # 2. Tải trang sản phẩm với timeout tăng lên 60 giây
        print(f"Đang tải trang sản phẩm: {product_url}")
        
        # Tạo session với User-Agent giống trình duyệt
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
            'Referer': 'https://baa.vn/'
        })
        
        # Tải trang với timeout 60 giây
        product_response = session.get(product_url, timeout=60)
        product_html = product_response.text
        
        # Parse HTML
        product_soup = BeautifulSoup(product_html, 'html.parser')
        
        # 3. Tìm kiếm các link PDF
        document_links = []
        
        # Tìm kiếm các liên kết tải tài liệu theo mẫu trong ví dụ
        # a. Tìm Link_download
        download_links = product_soup.select('a#Link_download[href$=".pdf"]')
        for link in download_links:
            href = link.get('href')
            if href:
                # Tìm tên file từ các phần tử liên quan
                parent_tr = link.find_parent('tr')
                if parent_tr:
                    filelink_elem = parent_tr.select_one('.filelink')
                    if filelink_elem:
                        name = filelink_elem.text.strip()
                    else:
                        name = href.split('/')[-1]
                else:
                    name = href.split('/')[-1]
                
                document_links.append({
                    'url': href,
                    'name': name
                })
                print(f"Đã tìm thấy link tải tài liệu (#Link_download): {name} - {href}")
        
        # b. Tìm .feature__metadata__link--download
        feature_links = product_soup.select('.feature__metadata__link--download[href$=".pdf"], span.feature__metadata__link--download')
        for link in feature_links:
            href = link.get('href')
            
            # Trường hợp span chứa href
            if not href and isinstance(link, Tag):
                href = link.get('href')
            
            # Một số trường hợp span không có href trực tiếp
            if not href:
                parent = link.parent
                if parent and parent.name == 'a':
                    href = parent.get('href')
                elif parent:
                    a_tag = parent.find('a')
                    if a_tag:
                        href = a_tag.get('href')
            
            if href and '.pdf' in href.lower():
                # Tìm tên file
                filelink_elem = link.select_one('.filelink')
                if filelink_elem:
                    name = filelink_elem.text.strip()
                else:
                    name = href.split('/')[-1]
                
                # Đảm bảo URL đầy đủ
                if not href.startswith('http'):
                    href = urljoin(product_url, href)
                
                document_links.append({
                    'url': href,
                    'name': name
                })
                print(f"Đã tìm thấy link tải tài liệu (.feature__metadata__link--download): {name} - {href}")
        
        # c. Tìm kiếm trong các bảng có chứa liên kết tài liệu
        catalog_tables = product_soup.select('.product-tab-content table, .tab-content table')
        for table in catalog_tables:
            for row in table.find_all('tr'):
                # Tìm các ô có chứa icon file và link PDF
                file_cells = row.select('td i.bi-file-earmark-text, td i.bi-file-pdf, td .filelink')
                if file_cells:
                    # Tìm link download trong hàng này
                    download_link = row.select_one('a[href$=".pdf"]')
                    if download_link:
                        href = download_link.get('href')
                        
                        # Tìm tên file
                        filelink_elem = row.select_one('.filelink')
                        if filelink_elem:
                            name = filelink_elem.text.strip()
                        else:
                            name = href.split('/')[-1]
                        
                        # Đảm bảo URL đầy đủ
                        if not href.startswith('http'):
                            href = urljoin(product_url, href)
                        
                        document_links.append({
                            'url': href,
                            'name': name
                        })
                        print(f"Đã tìm thấy link tải tài liệu (tìm trong bảng): {name} - {href}")
        
        # d. Kiểm tra tất cả các link có chứa PDF
        all_pdf_links = product_soup.select('a[href$=".pdf"]')
        for link in all_pdf_links:
            href = link.get('href')
            if href and 'document' in href.lower():
                name = link.text.strip()
                if not name:
                    name = href.split('/')[-1]
                
                # Đảm bảo URL đầy đủ
                if not href.startswith('http'):
                    href = urljoin(product_url, href)
                
                document_links.append({
                    'url': href,
                    'name': name
                })
                print(f"Đã tìm thấy link PDF khác: {name} - {href}")
        
        # 4. Loại bỏ các link trùng lặp
        unique_links = []
        added_urls = set()
        
        for link in document_links:
            if link['url'] not in added_urls:
                added_urls.add(link['url'])
                unique_links.append(link)
        
        print(f"Tìm thấy {len(unique_links)} tài liệu PDF không trùng lặp")
        
        # Lưu vào cache
        static_cache[cache_key] = unique_links
        
        return unique_links
    
    except Exception as e:
        print(f"Lỗi khi trích xuất tài liệu từ {product_url_or_code}: {str(e)}")
        traceback.print_exc()
        # Khởi tạo cache nếu chưa có
        if not hasattr(extract_product_documents, 'cache'):
            extract_product_documents.cache = {}
        static_cache[cache_key] = []
        return []

# Khởi tạo cache cho hàm
extract_product_documents.cache = {}

def debug_extract_products_from_url(url):
    """
    Hàm debug để kiểm tra cấu trúc HTML và trích xuất sản phẩm
    """
    print(f"\n=== DEBUG: Phân tích URL {url} ===")
    
    # Kiểm tra loại URL
    if is_category_url(url):
        print("✓ Được nhận diện là URL danh mục")
    elif is_product_url(url):
        print("✓ Được nhận diện là URL sản phẩm")
    else:
        print("✗ Không được nhận diện là URL hợp lệ")
        return []
    
    # Tải HTML
    html = get_html_content(url)
    if not html:
        print("✗ Không thể tải HTML")
        return []
    
    print(f"✓ Đã tải HTML, kích thước: {len(html)} ký tự")
    
    # Parse HTML
    soup = BeautifulSoup(html, 'html.parser')
    
    # Debug: Hiển thị title của trang
    title = soup.select_one('title')
    if title:
        print(f"✓ Title trang: {title.get_text(strip=True)}")
    
    # Debug: Kiểm tra các class và ID chính
    print("\n--- PHÂN TÍCH CẤU TRÚC HTML ---")
    
    # Tìm các div chính có thể chứa sản phẩm
    main_containers = soup.select('div[class*="product"], div[class*="item"], div[class*="card"], .row, .container')
    print(f"Tìm thấy {len(main_containers)} container có thể chứa sản phẩm")
    
    # Debug: Hiển thị một số class phổ biến
    common_classes = set()
    for element in soup.select('[class]')[:100]:  # Chỉ lấy 100 element đầu
        for cls in element.get('class', []):
            if any(keyword in cls.lower() for keyword in ['product', 'item', 'card', 'list', 'grid']):
                common_classes.add(cls)
    
    if common_classes:
        print(f"Các class phổ biến liên quan: {', '.join(sorted(common_classes)[:10])}")
    
    # Thử các selector khác nhau
    selectors_to_try = [
        ('a[href*="/san-pham/"]', 'Link có /san-pham/'),
        ('a[href*="/product/"]', 'Link có /product/'),
        ('a[href*="/vn/"]', 'Link có /vn/ (BAA pattern)'),
        ('.product-item a', 'Link trong .product-item'),
        ('.product-card a', 'Link trong .product-card'),
        ('.product-list a', 'Link trong .product-list'),
        ('a.product-link', 'Link với class .product-link'),
        ('a.product-item-link', 'Link với class .product-item-link'),
        ('.col-product a', 'Link trong .col-product'),
        ('.card.product__card', 'Card sản phẩm BAA'),
        ('a.card', 'Link card chung'),
        ('.row a', 'Link trong row'),
        ('a[href*="_"]', 'Link có chứa dấu gạch dưới'),
        ('.col-md-3 a', 'Link trong col-md-3'),
        ('.col-lg-3 a', 'Link trong col-lg-3'),
        ('.col-sm-6 a', 'Link trong col-sm-6'),
        ('div[class*="product"] a', 'Link trong div có class product'),
        ('div[class*="item"] a', 'Link trong div có class item'),
        ('.d-block a', 'Link trong .d-block'),
        ('a[title]', 'Link có attribute title'),
        ('img[alt]', 'Ảnh có alt (kiểm tra parent link)'),
        ('.grid-item a', 'Link trong .grid-item'),
        ('.list-item a', 'Link trong .list-item')
    ]
    
    all_found_links = []
    product_links = []
    
    for selector, description in selectors_to_try:
        elements = soup.select(selector)
        print(f"\n--- {description} ({selector}) ---")
        print(f"Tìm thấy {len(elements)} phần tử")
        
        # Xử lý đặc biệt cho img[alt]
        if selector == 'img[alt]':
            for img in elements:
                parent_a = img.find_parent('a')
                if parent_a and parent_a.get('href'):
                    href = parent_a.get('href')
                    alt_text = img.get('alt', '')[:50]
                    
                    if href:
                        # Tạo URL đầy đủ
                        if not href.startswith('http'):
                            parsed_url = urlparse(url)
                            full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}" if href.startswith('/') else f"{url.rstrip('/')}/{href}"
                        else:
                            full_url = href
                        
                        all_found_links.append(full_url)
                        is_product = is_product_url(full_url)
                        if is_product:
                            product_links.append(full_url)
                        
                        print(f"  [IMG] {full_url[:80]}{'...' if len(full_url) > 80 else ''}")
                        print(f"        Alt: {alt_text}")
                        print(f"        Is Product: {'✓' if is_product else '✗'}")
        else:
            for i, element in enumerate(elements[:5]):  # Chỉ hiển thị 5 phần tử đầu
                href = element.get('href', '')
                text = element.get_text(strip=True)[:50]  # Giới hạn text
                
                if href:
                    # Tạo URL đầy đủ
                    if not href.startswith('http'):
                        parsed_url = urlparse(url)
                        full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}" if href.startswith('/') else f"{url.rstrip('/')}/{href}"
                    else:
                        full_url = href
                    
                    all_found_links.append(full_url)
                    is_product = is_product_url(full_url)
                    if is_product:
                        product_links.append(full_url)
                    
                    print(f"  [{i+1}] {full_url[:80]}{'...' if len(full_url) > 80 else ''}")
                    print(f"      Text: {text}")
                    print(f"      Is Product: {'✓' if is_product else '✗'}")
    
    # Debug: Hiển thị một phần HTML để phân tích
    print(f"\n--- MẪU HTML (1000 ký tự đầu) ---")
    print(html[:1000])
    print("...")
    
    # Debug: Tìm tất cả các link và phân tích pattern
    print(f"\n--- PHÂN TÍCH TẤT CẢ LINK ---")
    all_links = soup.select('a[href]')
    print(f"Tổng số link trên trang: {len(all_links)}")
    
    # Phân loại link theo pattern
    patterns = {
        'san-pham': [],
        'product': [],
        'vn_with_id': [],  # pattern /vn/name_id
        'category': [],
        'other': []
    }
    
    for link in all_links[:20]:  # Chỉ phân tích 20 link đầu
        href = link.get('href', '')
        if '/san-pham/' in href:
            patterns['san-pham'].append(href)
        elif '/product/' in href:
            patterns['product'].append(href)
        elif re.search(r'/vn/[^/]+_\d+', href):
            patterns['vn_with_id'].append(href)
        elif 'category' in href.lower():
            patterns['category'].append(href)
        else:
            patterns['other'].append(href)
    
    for pattern_name, links in patterns.items():
        if links:
            print(f"Pattern '{pattern_name}': {len(links)} link")
            for link in links[:3]:
                print(f"  - {link}")
    
    # Loại bỏ trùng lặp
    unique_links = list(dict.fromkeys(all_found_links))
    unique_product_links = list(dict.fromkeys(product_links))
    
    print(f"\n=== KẾT QUẢ DEBUG ===")
    print(f"Tổng số link tìm thấy: {len(unique_links)}")
    print(f"Số link sản phẩm hợp lệ: {len(unique_product_links)}")
    
    if unique_product_links:
        print("\nCác URL sản phẩm hợp lệ:")
        for i, link in enumerate(unique_product_links[:10]):  # Hiển thị tối đa 10
            print(f"  {i+1}. {link}")
    else:
        print("\n⚠️  KHÔNG TÌM THẤY URL SẢN PHẨM NÀO!")
        print("Điều này có thể do:")
        print("  1. Cấu trúc HTML của trang đã thay đổi")
        print("  2. Selector không phù hợp với trang này")
        print("  3. Trang không chứa sản phẩm")
        print("  4. Cần cập nhật hàm is_product_url()")
    
    return unique_product_links

def test_baa_url_direct(url="https://baa.vn/vn/Category/den-thap-led-qlight_1479_1/"):
    """
    Hàm test trực tiếp với URL BAA.vn để debug
    """
    print(f"\n=== TEST TRỰC TIẾP URL BAA.VN ===")
    print(f"URL: {url}")
    
    try:
        # Test tải HTML
        html = get_html_content(url)
        if not html:
            print("❌ Không thể tải HTML")
            return
        
        print(f"✅ Đã tải HTML: {len(html)} ký tự")
        
        # Parse HTML
        soup = BeautifulSoup(html, 'html.parser')
        
        # Tìm title
        title = soup.select_one('title')
        if title:
            print(f"📄 Title: {title.get_text(strip=True)}")
        
        # Tìm tất cả link trên trang
        all_links = soup.select('a[href]')
        print(f"🔗 Tổng số link trên trang: {len(all_links)}")
        
        # Tìm các pattern URL phổ biến
        print(f"\n--- PHÂN TÍCH PATTERN URL ---")
        patterns_found = {
            'vn_with_number': [],
            'san_pham': [],
            'category': [],
            'other': []
        }
        
        for link in all_links:
            href = link.get('href', '')
            if not href:
                continue
                
            # Tạo full URL
            if not href.startswith('http'):
                parsed_url = urlparse(url)
                full_url = f"{parsed_url.scheme}://{parsed_url.netloc}{href}" if href.startswith('/') else f"{url.rstrip('/')}/{href}"
            else:
                full_url = href
            
            # Phân loại
            if re.search(r'/vn/[^/]+_\d+', full_url):
                patterns_found['vn_with_number'].append(full_url)
            elif '/san-pham/' in full_url:
                patterns_found['san_pham'].append(full_url)
            elif 'category' in full_url.lower():
                patterns_found['category'].append(full_url)
            else:
                patterns_found['other'].append(full_url)
        
        # Hiển thị kết quả
        for pattern_name, urls in patterns_found.items():
            if urls:
                print(f"\n{pattern_name.upper()}: {len(urls)} URL")
                for i, link_url in enumerate(urls[:5]):  # Chỉ hiển thị 5 URL đầu
                    is_product = is_product_url(link_url)
                    is_category = is_category_url(link_url)
                    print(f"  {i+1}. {link_url}")
                    print(f"     Product: {'✅' if is_product else '❌'} | Category: {'✅' if is_category else '❌'}")
        
        # Test extract_product_urls
        print(f"\n--- TEST EXTRACT_PRODUCT_URLS ---")
        product_urls = extract_product_urls(url)
        print(f"🎯 Kết quả extract_product_urls: {len(product_urls)} URL")
        
        if product_urls:
            print("Các URL sản phẩm tìm được:")
            for i, prod_url in enumerate(product_urls[:5]):
                print(f"  {i+1}. {prod_url}")
        else:
            print("❌ Không tìm được URL sản phẩm nào!")
        
        return product_urls
        
    except Exception as e:
        print(f"❌ Lỗi khi test: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def _is_baa_product_url(url):
    """Kiểm tra URL sản phẩm đặc biệt cho BAA.vn"""
    if 'baa.vn' not in url.lower():
        return False
    
    # Pattern 1: /vn/san-pham/...
    if '/san-pham/' in url:
        return True
    
    # Pattern 2: /vn/name_number (nhưng không phải Category)
    if re.search(r'/vn/[^/]+_\d+/?$', url) and '/Category/' not in url:
        # Loại bỏ các URL category hoặc các URL không phải sản phẩm
        excluded_keywords = ['category', 'list', 'search', 'page', 'menu', 'navigation']
        if not any(keyword in url.lower() for keyword in excluded_keywords):
            return True
    
    return False

def extract_baa_product_price(soup, product_name=""):
    """
    Trích xuất giá sản phẩm từ BAA.vn với logic đặc biệt:
    - Chỉ lấy giá từ span.product__price-print có data-root
    - Trả về chuỗi rỗng nếu không có cấu trúc giá hợp lệ
    - Giảm 5% cho sản phẩm có giá
    
    Args:
        soup: BeautifulSoup object của trang sản phẩm
        product_name: Tên sản phẩm để kiểm tra thêm
        
    Returns:
        str: Giá đã xử lý hoặc "" nếu không có giá
    """
    try:
        # 1. Chỉ tìm giá từ span.product__price-print có data-root
        price_element = soup.select_one('span.product__price-print[data-root]')
        
        if not price_element or not price_element.get('data-root'):
            print(f"  → Không có cấu trúc giá hợp lệ - để trống giá")
            return ""
        
        # 2. Lấy giá từ data-root
        try:
            price_value = float(price_element.get('data-root'))
            print(f"  → Tìm thấy giá từ product__price-print data-root: {price_value:,.0f}")
        except (ValueError, TypeError):
            print(f"  → data-root không hợp lệ: {price_element.get('data-root')} - để trống giá")
            return ""
        
        # 3. Kiểm tra giá có hợp lệ không
        if price_value <= 0:
            print(f"  → Giá không hợp lệ: {price_value} - để trống giá")
            return ""
        
        # 4. Áp dụng giảm giá 5%
        discounted_price = price_value * 0.95  # Giảm 5%
        
        # 5. Tìm đơn vị tiền tệ từ span.product__price-unit
        unit_element = soup.select_one('span.product__price-unit')
        currency = "₫"  # Mặc định
        if unit_element:
            currency = unit_element.get_text(strip=True).replace('&nbsp;', ' ').strip()
        
        # 6. Định dạng giá cuối cùng
        formatted_price = f"{discounted_price:,.0f}{currency}".replace(",", ".")
        
        print(f"  → Giá gốc: {price_value:,.0f}{currency} → Giá sau giảm 5%: {formatted_price}")
        
        return formatted_price
        
    except Exception as e:
        print(f"  → Lỗi khi trích xuất giá: {str(e)}")
        return ""
