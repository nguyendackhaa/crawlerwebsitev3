import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, send_from_directory, jsonify, session
from werkzeug.utils import secure_filename
import tempfile
import traceback
from app.crawler import extract_category_links, scrape_product_info, is_product_url, get_product_info, download_autonics_images, download_autonics_jpg_images, download_product_documents, extract_product_urls, is_category_url, download_baa_product_images, download_baa_product_images_fixed, extract_product_price, debug_extract_products_from_url
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime
from flask import current_app
from app import utils, socketio
from app.category_crawler import CategoryCrawler
import time
import openpyxl
import re
import zipfile
import shutil
from app.progress_bar import (
    TerminalProgressBar, 
    progress_tracker, 
    simple_progress, 
    batch_progress, 
    create_child_progress
)
from urllib.parse import urlparse
import concurrent.futures
from app.baa_crawler import BaaProductCrawler
from app.product_categorizer import ProductCategorizer
from app.resize import ImageResizer
from app.webp_converter import WebPConverter

main_bp = Blueprint('main', __name__)

ALLOWED_EXTENSIONS_TXT = {'txt'}
ALLOWED_EXTENSIONS_EXCEL = {'xlsx', 'xls'}

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@main_bp.route('/')
def index():
    return render_template('index.html')

@main_bp.route('/extract-links', methods=['POST'])
@progress_tracker(name="Trích xuất liên kết sản phẩm", total_steps=100, verbose=True)
def extract_links(progress: TerminalProgressBar):
    """
    Trích xuất liên kết sản phẩm từ file txt chứa danh sách URL danh mục
    """
    try:
        progress.update(5, "Kiểm tra file đầu vào")
        
        if 'link_file' not in request.files:
            progress.error("Không tìm thấy file!")
            flash('Không tìm thấy file!', 'error')
            return redirect(url_for('main.index'))
            
        file = request.files['link_file']
        
        if file.filename == '':
            progress.error("Không có file nào được chọn!")
            flash('Không có file nào được chọn!', 'error')
            return redirect(url_for('main.index'))
            
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            progress.error("Chỉ cho phép file .txt!")
            flash('Chỉ cho phép file .txt!', 'error')
            return redirect(url_for('main.index'))
        
        # Đọc nội dung file
        progress.update(10, "Đang đọc nội dung file", f"File: {file.filename}")
        content = file.read().decode('utf-8')
        
        # Tách thành danh sách URL, bỏ qua dòng trống
        raw_urls = content.strip().split('\n')
        urls = [url.strip() for url in raw_urls if url.strip()]
        
        progress.update(20, "Đang kiểm tra tính hợp lệ của URLs", f"Tìm thấy {len(urls)} URLs")
        
        # Lọc các URL hợp lệ
        valid_urls = []
        invalid_urls = []
        
        # Gửi thông báo bắt đầu (kết hợp với socketio)
        socketio.emit('progress_update', {'percent': 20, 'message': 'Đang kiểm tra URL...'})
        
        # Kiểm tra các URL với progress bar con
        url_check_progress = create_child_progress(progress, "Kiểm tra URLs", len(urls))
        
        for i, url in enumerate(urls):
            url_check_progress.update(i+1, f"Đang kiểm tra URL {i+1}/{len(urls)}")
            
            # Kiểm tra URL đặc biệt của đèn tháp LED
            is_led_url = 'den-thap-led-sang-tinh-chop-nhay-d45mm-qlight-st45l-and-st45ml-series_4779' in url
            
            if utils.is_valid_url(url) and (is_led_url or is_category_url(url)):
                valid_urls.append(url)
            else:
                invalid_urls.append(url)
        
        url_check_progress.complete("success", f"Đã kiểm tra {len(urls)} URLs", {
            "URLs hợp lệ": len(valid_urls),
            "URLs không hợp lệ": len(invalid_urls)
        })
        
        if not valid_urls:
            progress.error('Không có URL danh mục hợp lệ trong file!')
            flash('Không có URL danh mục hợp lệ trong file!', 'error')
            return redirect(url_for('main.index'))
            
        progress.update(40, f"Tìm thấy {len(valid_urls)} URL danh mục hợp lệ", f"Bỏ qua {len(invalid_urls)} URL không hợp lệ")
            
        # Gửi thông báo cập nhật
        socketio.emit('progress_update', {'percent': 40, 'message': f'Đã tìm thấy {len(valid_urls)} URL danh mục hợp lệ'})
        
        # Trích xuất liên kết sản phẩm từ các URL danh mục
        progress.update(50, "Đang trích xuất liên kết sản phẩm từ các danh mục")
        extract_progress = create_child_progress(progress, "Trích xuất sản phẩm", len(valid_urls))
        
        product_links = extract_category_links(valid_urls)
        
        extract_progress.complete("success", f"Trích xuất hoàn tất", {
            "Sản phẩm tìm thấy": len(product_links)
        })
        
        if not product_links:
            progress.error('Không tìm thấy liên kết sản phẩm nào!')
            flash('Không tìm thấy liên kết sản phẩm nào!', 'error')
            return redirect(url_for('main.index'))
            
        progress.update(80, f"Đã trích xuất {len(product_links)} liên kết sản phẩm", "Đang tạo file kết quả")
            
        # Gửi thông báo hoàn thành trích xuất liên kết
        socketio.emit('progress_update', {'percent': 80, 'message': f'Đã trích xuất xong {len(product_links)} liên kết sản phẩm'})
        
        # Tạo file kết quả
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        result_filename = f"product_links_{timestamp}.txt"
        result_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], result_filename)
        
        # Lưu danh sách liên kết vào file
        with open(result_file_path, 'w', encoding='utf-8') as f:
            for link in product_links:
                f.write(link + '\n')
        
        progress.update(95, "Đang tạo đường dẫn tải xuống")
        
        # Tạo đường dẫn download
        download_url = url_for('main.download_file', filename=result_filename)
        
        # Thông báo thành công
        success_message = f'Đã tìm thấy {len(product_links)} liên kết sản phẩm. '
        if invalid_urls:
            success_message += f'Có {len(invalid_urls)} URL không hợp lệ bị bỏ qua.'
            
        flash(success_message, 'success')
        
        # Gửi thông báo hoàn thành
        socketio.emit('progress_update', {'percent': 100, 'message': 'Hoàn thành!'})
        
        # Hoàn thành progress với thống kê chi tiết
        progress.complete("success", "Trích xuất liên kết hoàn tất", {
            "File đầu vào": file.filename,
            "URLs đầu vào": len(urls),
            "URLs hợp lệ": len(valid_urls),
            "URLs không hợp lệ": len(invalid_urls),
            "Sản phẩm tìm thấy": len(product_links),
            "File kết quả": result_filename
        })
        
        # Hiển thị trang kết quả
        return render_template('index.html', download_url=download_url)
        
    except Exception as e:
        error_message = str(e)
        # In chi tiết lỗi
        traceback.print_exc()
        progress.error(f'Lỗi: {error_message}', traceback.format_exc())
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/scrape-products', methods=['POST'])
def scrape_products():
    try:
        if 'product_link_file' not in request.files or 'excel_template' not in request.files:
            return render_template('index.html', error="Thiếu file")
        
        link_file = request.files['product_link_file']
        excel_file = request.files['excel_template']
        
        if link_file.filename == '' or excel_file.filename == '':
            return render_template('index.html', error="Không có file nào được chọn")
        
        if not allowed_file(link_file.filename, ALLOWED_EXTENSIONS_TXT):
            return render_template('index.html', error="File liên kết phải là file .txt")
        
        if not allowed_file(excel_file.filename, ALLOWED_EXTENSIONS_EXCEL):
            return render_template('index.html', error="File mẫu phải là file .xlsx hoặc .xls")
        
        # Đọc file txt chứa danh sách URL sản phẩm
        product_urls = []
        invalid_urls = []
        content = link_file.read().decode('utf-8')
        
        for line in content.splitlines():
            line = line.strip()
            if line.startswith('http'):
                if is_product_url(line):
                    product_urls.append(line)
                else:
                    invalid_urls.append(line)
        
        if invalid_urls:
            warning_msg = f"Có {len(invalid_urls)} URL không phải là trang sản phẩm và sẽ bị bỏ qua."
            print(warning_msg)
            print(f"Các URL không hợp lệ: {invalid_urls}")
            
        if not product_urls:
            return render_template('index.html', error="Không tìm thấy URL sản phẩm hợp lệ trong file")
        
        # Lưu file Excel mẫu
        excel_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
        excel_file.save(excel_path)
        
        # Thu thập thông tin sản phẩm
        result_file = scrape_product_info(product_urls, excel_path)
        
        # Xóa file Excel mẫu sau khi sử dụng
        os.unlink(excel_path)
        
        success_message = f"Đã thu thập thông tin từ {len(product_urls)} sản phẩm."
        if invalid_urls:
            success_message += f" Đã bỏ qua {len(invalid_urls)} URL không hợp lệ."
            
        print(success_message)
        
        return send_file(result_file, as_attachment=True, download_name="product_info.xlsx")
    except Exception as e:
        error_msg = f"Lỗi khi xử lý: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        return render_template('index.html', error=error_msg)

@main_bp.route('/process', methods=['POST'])
def process_url():
    """
    Process URL from form submission and scrape product information
    """
    try:
        url = request.form.get('url')
        required_fields = request.form.getlist('required_fields')
        
        if not url:
            flash('URL không được để trống!', 'error')
            return redirect(url_for('index'))
        
        if not utils.is_valid_url(url):
            flash('URL không hợp lệ!', 'error')
            return redirect(url_for('index'))

        if not required_fields:
            flash('Hãy chọn ít nhất một trường dữ liệu!', 'error')
            return redirect(url_for('index'))

        # Convert checkbox values to field names
        field_mapping = {
            'field_id': 'STT',
            'field_code': 'Mã sản phẩm',
            'field_name': 'Tên sản phẩm',
            'field_overview': 'Tổng quan',
            'field_url': 'URL'
        }
        
        selected_fields = [field_mapping[field] for field in required_fields if field in field_mapping]
        
        # Scrape product information
        product_info_list = get_product_info(url, selected_fields)
        
        if not product_info_list:
            flash('Không tìm thấy thông tin sản phẩm!', 'error')
            return redirect(url_for('index'))
            
        # Generate a temporary file name
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        temp_file = f"product_info_{timestamp}.xlsx"
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], temp_file)
        
        # Save to Excel file
        utils.save_to_excel(product_info_list, file_path)
        
        # Generate download URL
        download_url = url_for('main.download_file', filename=temp_file)
        
        return render_template('index.html', download_url=download_url)
    
    except Exception as e:
        current_app.logger.error(f"Error processing URL: {str(e)}")
        flash(f'Đã xảy ra lỗi: {str(e)}', 'error')
        return redirect(url_for('index'))

@main_bp.route('/download/<path:filename>')
def download_file(filename):
    download_dir = os.path.join(current_app.root_path, 'downloads')
    file_path = os.path.join(download_dir, filename)
    
    # Kiểm tra xem file có tồn tại không
    if not os.path.exists(file_path):
        # Trường hợp file ZIP không tồn tại nhưng thư mục tồn tại (chưa được nén)
        if filename.endswith('.zip'):
            folder_name = filename[:-4]  # Bỏ phần .zip để lấy tên thư mục
            folder_path = os.path.join(download_dir, folder_name)
            
            # Kiểm tra xem thư mục có tồn tại không
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                try:
                    # Tạo file ZIP mới
                    utils.create_zip_from_folder(folder_path, file_path)
                    print(f"Đã tạo file ZIP mới: {file_path}")
                except Exception as e:
                    print(f"Lỗi khi tạo file ZIP: {str(e)}")
                    flash(f'Lỗi khi tạo file ZIP: {str(e)}', 'error')
                    return redirect(url_for('main.index'))
    
    # Tách filename thành thư mục và tên file nếu có '/'
    if '/' in filename:
        # Ví dụ: reports/product_comparison_report_20250429_170113.xlsx
        folder, file = filename.split('/', 1)
        download_path = os.path.join(download_dir, folder)
        return send_from_directory(directory=download_path, path=file, as_attachment=True)
    else:
        # Trường hợp file nằm trực tiếp trong thư mục downloads
        return send_from_directory(directory=download_dir, path=filename, as_attachment=True)

@main_bp.route('/view-images')
def view_images():
    """Hiển thị danh sách các ảnh đã tải xuống"""
    try:
        # Thư mục chứa tất cả các thư mục ảnh
        download_dir = os.path.join(current_app.root_path, 'downloads')
        
        # Tìm tất cả các thư mục images_*
        image_folders = [d for d in os.listdir(download_dir) if os.path.isdir(os.path.join(download_dir, d)) and d.startswith('images_')]
        
        if not image_folders:
            return render_template('view_images.html', error="Không tìm thấy thư mục ảnh nào")
        
        # Sắp xếp theo thời gian tạo (mới nhất trước)
        image_folders.sort(reverse=True)
        
        # Lấy thư mục mới nhất
        latest_folder = image_folders[0]
        latest_folder_path = os.path.join(download_dir, latest_folder)
        
        # Tìm tất cả các file ảnh trong thư mục
        all_images = []
        for root, dirs, files in os.walk(latest_folder_path):
            for file in files:
                if file.endswith('.webp'):
                    # Tạo đường dẫn tương đối từ thư mục tải xuống
                    rel_path = os.path.relpath(os.path.join(root, file), download_dir)
                    # Chuyển dấu gạch ngược thành gạch chéo cho URL
                    rel_path = rel_path.replace('\\', '/')
                    
                    # Lấy mã sản phẩm từ tên file
                    product_code = os.path.splitext(file)[0]
                    
                    all_images.append({
                        'path': rel_path,
                        'code': product_code,
                        'url': url_for('main.view_image', image_path=rel_path)
                    })
        
        return render_template('view_images.html', images=all_images, folder=latest_folder)
        
    except Exception as e:
        error_msg = f"Lỗi khi hiển thị ảnh: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        return render_template('view_images.html', error=error_msg)

@main_bp.route('/view-image/<path:image_path>')
def view_image(image_path):
    """Hiển thị một ảnh cụ thể"""
    download_dir = os.path.join(current_app.root_path, 'downloads')
    return send_from_directory(directory=download_dir, path=image_path)

@main_bp.route('/convert-to-webp', methods=['POST'])
def convert_to_webp():
    """Chuyển đổi ảnh từ JPG/PNG sang WebP THỰC SỰ"""
    try:
        if 'image_files' not in request.files:
            return render_template('index.html', error="Không tìm thấy file")
        
        files = request.files.getlist('image_files')
        
        if not files or files[0].filename == '':
            return render_template('index.html', error="Không có file nào được chọn")
        
        # Import WebPConverter
        from app.webp_converter import WebPConverter
        
        # Kiểm tra các file có đúng định dạng không
        allowed_formats = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.tif']
        for file in files:
            filename = file.filename.lower()
            if not any(filename.endswith(fmt) for fmt in allowed_formats):
                return render_template('index.html', error=f"Chỉ chấp nhận file: {', '.join(allowed_formats)}")
        
        # Tạo thư mục đầu ra cho ảnh đã chuyển đổi
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        webp_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'webp_converted_{timestamp}')
        temp_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'temp_{timestamp}')
        os.makedirs(webp_folder, exist_ok=True)
        os.makedirs(temp_folder, exist_ok=True)
        
        print(f"Đã tạo thư mục lưu ảnh webp: {webp_folder}")
        
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 5, 
            'message': f'Bắt đầu chuyển đổi {len(files)} ảnh sang định dạng WebP THỰC SỰ',
            'detail': 'Đang kiểm tra và xử lý từng file...'
        })
        
        # Biến lưu trữ kết quả chuyển đổi
        converted_files = []
        conversion_results = []
        total_input_size = 0
        total_output_size = 0
        
        # Lưu các file tạm thời trước
        temp_files = []
        for i, file in enumerate(files):
            original_filename = secure_filename(file.filename)
            temp_path = os.path.join(temp_folder, original_filename)
            file.save(temp_path)
            temp_files.append(temp_path)
        
        # Xử lý từng file
        for i, temp_path in enumerate(temp_files):
            # Tính phần trăm tiến trình
            progress = int(5 + ((i / len(temp_files)) * 85))
            
            try:
                # Lấy tên file gốc
                original_filename = os.path.basename(temp_path)
                base_name = os.path.splitext(original_filename)[0]
                webp_filename = f"{base_name}.webp"
                
                # Đường dẫn đầy đủ cho file đầu ra
                webp_path = os.path.join(webp_folder, webp_filename)
                
                # Cập nhật tiến trình
                socketio.emit('progress_update', {
                    'percent': progress,
                    'message': f'Đang xử lý ảnh {i+1}/{len(temp_files)}: {original_filename}',
                    'detail': 'Chuyển đổi sang WebP với kiểm tra nghiêm ngặt...'
                })
                
                # Sử dụng WebPConverter để chuyển đổi
                result = WebPConverter.convert_to_webp(
                    input_path=temp_path,
                    output_path=webp_path,
                    quality=90,  # Chất lượng cao
                    lossless=False,  # Nén có mất dữ liệu cho kích thước nhỏ hơn
                    method=6  # Phương pháp nén tốt nhất (chậm hơn nhưng chất lượng cao)
                )
                
                if result['success']:
                    # Kiểm tra lại MIME type
                    mime_type = WebPConverter.get_mime_type(webp_path)
                    if mime_type != 'image/webp':
                        raise Exception(f"MIME type không đúng: {mime_type}, cần là image/webp")
                    
                    converted_files.append(webp_path)
                    conversion_results.append({
                        'original': original_filename,
                        'converted': webp_filename,
                        'path': webp_path,
                        'status': 'success',
                        'input_size': result['input_size'],
                        'output_size': result['output_size'],
                        'compression_ratio': result['compression_ratio'],
                        'webp_verified': result['webp_verified'],
                        'webp_format': result['webp_info']['format'] if result['webp_info'] else 'N/A',
                        'mime_type': mime_type
                    })
                    
                    total_input_size += result['input_size']
                    total_output_size += result['output_size']
                    
                    print(f"✓ Chuyển đổi thành công: {original_filename} -> {webp_filename}")
                    print(f"  - WebP verified: {result['webp_verified']}")
                    print(f"  - Format: {result['webp_info']['format'] if result['webp_info'] else 'N/A'}")
                    print(f"  - MIME type: {mime_type}")
                    print(f"  - Giảm {result['compression_ratio']}%")
                else:
                    raise Exception(result['error'] or 'Lỗi không xác định')
                
            except Exception as e:
                error_msg = f"Lỗi khi xử lý {original_filename}: {str(e)}"
                print(f"✗ {error_msg}")
                conversion_results.append({
                    'original': original_filename,
                    'status': 'error',
                    'error': str(e)
                })
        
        # Xóa thư mục tạm
        try:
            import shutil
            shutil.rmtree(temp_folder)
        except:
            pass
        
        # Gửi thông báo hoàn thành
        socketio.emit('progress_update', {
            'percent': 90, 
            'message': f'Đã chuyển đổi thành công {len(converted_files)}/{len(files)} ảnh. Đang tạo file ZIP...',
            'detail': f'Tổng dung lượng giảm: {round((1 - total_output_size / total_input_size) * 100, 2)}%' if total_input_size > 0 else ''
        })
        
        # Tạo báo cáo Excel về kết quả chuyển đổi
        report_path = os.path.join(webp_folder, 'conversion_report.xlsx')
        try:
            report_df = pd.DataFrame(conversion_results)
            
            # Sắp xếp các cột
            if 'status' in report_df.columns and 'success' in report_df['status'].values:
                success_df = report_df[report_df['status'] == 'success']
                if not success_df.empty:
                    report_df = pd.DataFrame(conversion_results)
                    
            report_df.to_excel(report_path, index=False)
            print(f"Đã tạo báo cáo: {report_path}")
        except Exception as e:
            print(f"Lỗi khi tạo báo cáo: {str(e)}")
        
        # Tạo file ZIP để tải xuống
        zip_filename = f'webp_images_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        
        # Tạo file zip từ thư mục ảnh
        if utils.create_zip_from_folder(webp_folder, zip_path):
            print(f"Đã tạo file ZIP thành công: {zip_path}")
            
            # Cập nhật hoàn thành
            socketio.emit('progress_update', {
                'percent': 100, 
                'message': f'Đã chuyển đổi và nén {len(converted_files)} ảnh WebP THỰC SỰ thành công!',
                'detail': 'Tất cả file đã được kiểm tra và xác nhận là WebP chuẩn.'
            })
            
            # Thông báo thành công chi tiết
            success_message = f"""
            Đã chuyển đổi thành công {len(converted_files)}/{len(files)} ảnh sang định dạng WebP THỰC SỰ.
            <br><br>
            <strong>Thông tin chi tiết:</strong><br>
            • Tổng dung lượng gốc: {round(total_input_size / 1024 / 1024, 2)} MB<br>
            • Tổng dung lượng WebP: {round(total_output_size / 1024 / 1024, 2)} MB<br>
            • Tiết kiệm: {round((1 - total_output_size / total_input_size) * 100, 2)}%<br>
            • Tất cả file đã được xác minh là WebP chuẩn (RIFF/WEBP signature)
            """
            
            # Tạo URL để tải xuống file zip
            download_url = url_for('main.download_file', filename=zip_filename)
            
            return render_template('index.html', 
                                 download_url=download_url, 
                                 success_message=success_message,
                                 active_tab='view-images-tab')
        else:
            return render_template('index.html', error="Không thể tạo file ZIP từ ảnh đã chuyển đổi")
    
    except Exception as e:
        error_msg = f"Lỗi khi xử lý: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=error_msg)

@main_bp.route('/download-images', methods=['POST'])
@progress_tracker(name="Tải ảnh sản phẩm", total_steps=100, verbose=True)
def download_images(progress: TerminalProgressBar):
    try:
        # Đặt thời gian bắt đầu
        start_time = time.time()
        
        progress.update(5, "Kiểm tra file đầu vào")
        
        if 'product_code_file' not in request.files:
            progress.error("Không tìm thấy file")
            return render_template('index.html', error="Không tìm thấy file")
        
        file = request.files['product_code_file']
        
        if file.filename == '':
            progress.error("Không có file nào được chọn")
            return render_template('index.html', error="Không có file nào được chọn")
        
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            progress.error("Chỉ chấp nhận file .txt")
            return render_template('index.html', error="Chỉ chấp nhận file .txt")
        
        progress.update(10, "Đang đọc danh sách mã sản phẩm", f"File: {file.filename}")
        
        # Đọc file txt chứa danh sách mã sản phẩm
        product_codes = []
        content = file.read().decode('utf-8')
        for line in content.splitlines():
            line = line.strip()
            if line:  # Bỏ qua dòng trống
                product_codes.append(line)
        
        if not product_codes:
            progress.error("Không tìm thấy mã sản phẩm hợp lệ trong file")
            return render_template('index.html', error="Không tìm thấy mã sản phẩm hợp lệ trong file")
        
        progress.update(15, f"Đã đọc {len(product_codes)} mã sản phẩm", "Đang tạo thư mục lưu ảnh")
        
        # Tạo thư mục đầu ra cho hình ảnh
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        images_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'images_{timestamp}')
        os.makedirs(images_folder, exist_ok=True)
        print(f"Đã tạo thư mục lưu ảnh: {images_folder}")
        
        # Gửi thông báo bắt đầu (kết hợp với socketio)
        socketio.emit('progress_update', {
            'percent': 20, 
            'message': f'Chuẩn bị tải {len(product_codes)} hình ảnh sản phẩm Autonics'
        })
        
        progress.update(20, "Đang bắt đầu tải ảnh sản phẩm", f"Đích: {images_folder}")
        
        # Tạo progress bar con cho việc tải ảnh
        download_progress = create_child_progress(progress, "Download Autonics Images", 60)
        
        # Tải ảnh sản phẩm với giới hạn thời gian
        summary = download_autonics_images(product_codes, images_folder)
        
        download_progress.complete("success", "Tải ảnh hoàn tất", {
            "Thành công": summary['successful'],
            "Thất bại": summary['total'] - summary['successful'],
            "Tổng cộng": summary['total']
        })
        
        if summary['successful'] == 0:
            elapsed_time = time.time() - start_time
            print(f"Quá trình xử lý thất bại sau {elapsed_time:.2f} giây")
            progress.error("Không tải được ảnh nào")
            return render_template('index.html', error="Không tải được ảnh nào")
        
        progress.update(85, f"Đã tải {summary['successful']}/{summary['total']} ảnh", "Đang tạo file ZIP")
        
        # Tạo đường dẫn cho file zip
        zip_filename = f'product_images_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        print(f"Bắt đầu tạo file ZIP: {zip_path}")
        
        # Tạo progress bar con cho việc tạo ZIP
        zip_progress = create_child_progress(progress, "Tạo file ZIP", 10)
        
        # Tạo file zip từ thư mục ảnh
        if utils.create_zip_from_folder(images_folder, zip_path):
            zip_progress.complete("success", "Đã tạo file ZIP thành công")
            
            print(f"Đã tạo file ZIP thành công: {zip_path}")
            success_message = f"Đã tải thành công {summary['successful']}/{summary['total']} ảnh sản phẩm."
            
            # Tạo URL để tải xuống file zip
            download_url = url_for('main.download_file', filename=zip_filename)
            
            # Tính toán thời gian xử lý tổng cộng
            elapsed_time = time.time() - start_time
            print(f"Hoàn thành sau {elapsed_time:.2f} giây")
            avg_time = elapsed_time / len(product_codes) if product_codes else 0
            print(f"Thời gian xử lý trung bình: {avg_time:.2f} giây/sản phẩm")
            
            # Bổ sung thông tin hiệu năng vào thông báo thành công
            if avg_time <= 10:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (tốt)."
            else:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (cần cải thiện)."
            
            # Hoàn thành progress với thống kê chi tiết
            progress.complete("success", "Tải ảnh sản phẩm hoàn tất", {
                "File đầu vào": file.filename,
                "Mã sản phẩm": len(product_codes),
                "Ảnh thành công": summary['successful'],
                "Ảnh thất bại": summary['total'] - summary['successful'],
                "Thời gian xử lý": f"{elapsed_time:.2f}s",
                "Tốc độ TB": f"{avg_time:.2f}s/sản phẩm",
                "File ZIP": zip_filename
            })
            
            return render_template('index.html', download_url=download_url, success_message=success_message)
        else:
            zip_progress.error("Không thể tạo file ZIP")
            progress.error("Không thể tạo file ZIP từ ảnh đã tải")
            return render_template('index.html', error="Không thể tạo file ZIP từ ảnh đã tải")
    
    except Exception as e:
        error_msg = f"Lỗi khi xử lý: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        progress.error(f'Lỗi: {error_msg}', traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=error_msg)

@main_bp.route('/download-jpg-images', methods=['POST'])
def download_jpg_images():
    try:
        # Đặt thời gian bắt đầu
        start_time = time.time()
        
        if 'product_code_file' not in request.files:
            return render_template('index.html', error="Không tìm thấy file")
        
        file = request.files['product_code_file']
        
        if file.filename == '':
            return render_template('index.html', error="Không có file nào được chọn")
        
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            return render_template('index.html', error="Chỉ chấp nhận file .txt")
        
        # Đọc file txt chứa danh sách mã sản phẩm
        product_codes = []
        content = file.read().decode('utf-8')
        for line in content.splitlines():
            line = line.strip()
            if line:  # Bỏ qua dòng trống
                product_codes.append(line)
        
        if not product_codes:
            return render_template('index.html', error="Không tìm thấy mã sản phẩm hợp lệ trong file")
        
        # Tạo thư mục đầu ra cho hình ảnh
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        images_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'jpg_images_{timestamp}')
        os.makedirs(images_folder, exist_ok=True)
        print(f"Đã tạo thư mục lưu ảnh JPG: {images_folder}")
        
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 5, 
            'message': f'Chuẩn bị tải {len(product_codes)} hình ảnh JPG chất lượng cao'
        })
        
        # Tải ảnh sản phẩm với chất lượng cao
        summary = download_autonics_jpg_images(product_codes, images_folder)
        
        if summary['successful'] == 0:
            elapsed_time = time.time() - start_time
            print(f"Quá trình xử lý thất bại sau {elapsed_time:.2f} giây")
            return render_template('index.html', error="Không tải được ảnh nào")
        
        # Tạo đường dẫn cho file zip
        zip_filename = f'jpg_product_images_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        print(f"Bắt đầu tạo file ZIP: {zip_path}")
        
        # Tạo file zip từ thư mục ảnh
        if utils.create_zip_from_folder(images_folder, zip_path):
            print(f"Đã tạo file ZIP thành công: {zip_path}")
            success_message = f"Đã tải thành công {summary['successful']}/{summary['total']} ảnh JPG chất lượng cao."
            
            # Tạo URL để tải xuống file zip
            download_url = url_for('main.download_file', filename=zip_filename)
            
            # Tính toán thời gian xử lý tổng cộng
            elapsed_time = time.time() - start_time
            print(f"Hoàn thành sau {elapsed_time:.2f} giây")
            avg_time = elapsed_time / len(product_codes) if product_codes else 0
            print(f"Thời gian xử lý trung bình: {avg_time:.2f} giây/sản phẩm")
            
            # Bổ sung thông tin hiệu năng vào thông báo thành công
            if avg_time <= 10:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (tốt)."
            else:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (cần cải thiện)."
            
            return render_template('index.html', download_url=download_url, success_message=success_message)
        else:
            return render_template('index.html', error="Không thể tạo file ZIP từ ảnh đã tải")
    
    except Exception as e:
        error_msg = f"Lỗi khi xử lý: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=error_msg)

@main_bp.route('/download-documents', methods=['POST'])
def download_documents():
    try:
        # Đặt thời gian bắt đầu
        start_time = time.time()
        
        if 'product_code_file' not in request.files:
            return render_template('index.html', error="Không tìm thấy file")
        
        file = request.files['product_code_file']
        
        if file.filename == '':
            return render_template('index.html', error="Không có file nào được chọn")
        
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            return render_template('index.html', error="Chỉ chấp nhận file .txt")
        
        # Đọc file txt chứa danh sách URL sản phẩm thay vì mã sản phẩm
        product_urls = []
        content = file.read().decode('utf-8')
        for line in content.splitlines():
            line = line.strip()
            if line and line.startswith('http'):  # Chỉ nhận các dòng có chứa URL
                product_urls.append(line)
        
        if not product_urls:
            return render_template('index.html', error="Không tìm thấy URL sản phẩm hợp lệ trong file")
        
        # Tạo thư mục đầu ra cho tài liệu PDF
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        documents_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'pdf_documents_{timestamp}')
        os.makedirs(documents_folder, exist_ok=True)
        print(f"Đã tạo thư mục lưu tài liệu PDF: {documents_folder}")
        
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 5, 
            'message': f'Chuẩn bị tải tài liệu PDF cho {len(product_urls)} sản phẩm'
        })
        
        # Tải tài liệu PDF cho các sản phẩm
        summary = download_product_documents(product_urls, documents_folder)
        
        # Tính tổng số tài liệu đã tải thành công
        total_documents = 0
        for product_code, result in summary['product_results'].items():
            total_documents += result.get('successful_documents', 0)
        
        if total_documents == 0:
            elapsed_time = time.time() - start_time
            print(f"Quá trình xử lý thất bại sau {elapsed_time:.2f} giây")
            return render_template('index.html', error="Không tải được tài liệu PDF nào")
        
        # Tạo đường dẫn cho file zip
        zip_filename = f'product_documents_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        print(f"Bắt đầu tạo file ZIP: {zip_path}")
        
        # Tạo file zip từ thư mục tài liệu
        if utils.create_zip_from_folder(documents_folder, zip_path):
            print(f"Đã tạo file ZIP thành công: {zip_path}")
            success_message = f"Đã tải thành công {total_documents} tài liệu PDF từ {summary['successful_products']}/{summary['total_products']} sản phẩm."
            
            # Tạo URL để tải xuống file zip
            download_url = url_for('main.download_file', filename=zip_filename)
            
            # Tính toán thời gian xử lý tổng cộng
            elapsed_time = time.time() - start_time
            print(f"Hoàn thành sau {elapsed_time:.2f} giây")
            avg_time = elapsed_time / len(product_urls) if product_urls else 0
            print(f"Thời gian xử lý trung bình: {avg_time:.2f} giây/sản phẩm")
            
            # Bổ sung thông tin hiệu năng vào thông báo thành công
            if avg_time <= 15:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (tốt)."
            else:
                success_message += f" Tốc độ xử lý: {avg_time:.2f} giây/sản phẩm (cần cải thiện)."
            
            return render_template('index.html', download_url=download_url, success_message=success_message)
        else:
            return render_template('index.html', error="Không thể tạo file ZIP từ tài liệu đã tải")
    
    except Exception as e:
        error_msg = f"Lỗi khi xử lý: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=error_msg)

@main_bp.route('/compare-categories', methods=['POST'])
def compare_categories():
    try:
        # Kiểm tra các file đã tải lên
        if 'urls_file_1' not in request.files or 'urls_file_2' not in request.files:
            return render_template('index.html', error="Vui lòng tải lên cả hai file TXT chứa danh sách URL")
        
        urls_file_1 = request.files['urls_file_1']
        urls_file_2 = request.files['urls_file_2']
        
        if urls_file_1.filename == '' or urls_file_2.filename == '':
            return render_template('index.html', error="Không có file nào được chọn")
        
        if not allowed_file(urls_file_1.filename, ALLOWED_EXTENSIONS_TXT) or not allowed_file(urls_file_2.filename, ALLOWED_EXTENSIONS_TXT):
            return render_template('index.html', error="Chỉ chấp nhận file .txt cho danh sách URL")
        
        # Thông báo tiến trình
        socketio.emit('progress_update', {
            'percent': 5, 
            'message': 'Đang đọc file URLs...'
        })
        
        # Đọc danh sách URL từ file
        urls_1 = []  # URLs từ BAA.vn
        content = urls_file_1.read().decode('utf-8')
        for line in content.splitlines():
            line = line.strip()
            if line and line.startswith('http'):
                urls_1.append(line)
        
        urls_2 = []  # URLs từ HaiphongTech
        content = urls_file_2.read().decode('utf-8')
        for line in content.splitlines():
            line = line.strip()
            if line and line.startswith('http'):
                urls_2.append(line)
        
        if not urls_1:
            return render_template('index.html', error="Không tìm thấy URL hợp lệ trong file thứ nhất")
        
        if not urls_2:
            return render_template('index.html', error="Không tìm thấy URL hợp lệ trong file thứ hai")
        
        # Thông báo trước khi bắt đầu thu thập
        socketio.emit('progress_update', {
            'percent': 10, 
            'message': f'Đã tìm thấy {len(urls_1)} URL từ file BAA.vn. Đang xử lý...'
        })
        
        # Thu thập sản phẩm từ BAA.vn
        from app.crawler import is_product_url, extract_product_urls
        
        # Phân loại URL thành URL danh mục và URL sản phẩm
        product_urls_1 = []
        category_urls_1 = []
        
        for url in urls_1:
            if is_product_url(url):
                product_urls_1.append(url)
            else:
                category_urls_1.append(url)
        
        # Thu thập thêm từ các URL danh mục
        for i, category_url in enumerate(category_urls_1):
            progress = 10 + int((i / len(category_urls_1)) * 30)
            socketio.emit('progress_update', {
                'percent': progress, 
                'message': f'Đang xử lý URL danh mục BAA.vn ({i+1}/{len(category_urls_1)}): {category_url}'
            })
            
            # Lấy sản phẩm từ danh mục
            category_products = extract_product_urls(category_url)
            for product_url in category_products:
                if product_url not in product_urls_1:
                    product_urls_1.append(product_url)
        
        if not product_urls_1:
            return render_template('index.html', error="Không tìm thấy sản phẩm nào từ URLs BAA.vn")
        
        socketio.emit('progress_update', {
            'percent': 40, 
            'message': f'Đã tìm thấy {len(product_urls_1)} sản phẩm từ BAA.vn. Đang xử lý HaiphongTech...'
        })
        
        # Thu thập sản phẩm từ HaiphongTech
        product_urls_2 = []
        category_urls_2 = []
        
        for url in urls_2:
            if is_product_url(url):
                product_urls_2.append(url)
            else:
                category_urls_2.append(url)
        
        # Thu thập thêm từ các URL danh mục
        for i, category_url in enumerate(category_urls_2):
            progress = 40 + int((i / len(category_urls_2)) * 30)
            socketio.emit('progress_update', {
                'percent': progress, 
                'message': f'Đang xử lý URL danh mục HaiphongTech ({i+1}/{len(category_urls_2)}): {category_url}'
            })
            
            # Lấy sản phẩm từ danh mục
            category_products = extract_product_urls(category_url)
            for product_url in category_products:
                if product_url not in product_urls_2:
                    product_urls_2.append(product_url)
        
        if not product_urls_2:
            return render_template('index.html', error="Không tìm thấy sản phẩm nào từ URLs HaiphongTech")
        
        socketio.emit('progress_update', {
            'percent': 70, 
            'message': f'Đã tìm thấy {len(product_urls_2)} sản phẩm từ HaiphongTech. Đang trích xuất và chuẩn hóa mã sản phẩm...'
        })
        
        # Trích xuất và chuẩn hóa mã sản phẩm từ các URL của BAA.vn
        product_codes_1 = []  # Chuẩn hóa mã sản phẩm
        product_urls_by_code_1 = {}  # Map mã sản phẩm -> URL
        original_codes_by_std_1 = {}  # Map mã chuẩn hóa -> mã gốc
        
        for i, url in enumerate(product_urls_1):
            if i % 10 == 0:
                progress = 70 + int((i / len(product_urls_1)) * 5)
                socketio.emit('progress_update', {
                    'percent': progress, 
                    'message': f'Đang trích xuất mã sản phẩm từ BAA.vn ({i}/{len(product_urls_1)})...'
                })
            
            # Trích xuất mã sản phẩm từ URL hoặc nội dung trang
            original_code = extract_product_code_from_url(url)
            if original_code:
                # Function extract_product_code_from_url đã gọi standardize_product_code
                product_codes_1.append(original_code)
                product_urls_by_code_1[original_code] = url
                original_codes_by_std_1[original_code] = original_code
                print(f"Sản phẩm BAA.vn: {original_code} - {url}")
        
        # Trích xuất và chuẩn hóa mã sản phẩm từ các URL của HaiphongTech
        product_codes_2 = []  # Chuẩn hóa mã sản phẩm
        product_urls_by_code_2 = {}  # Map mã sản phẩm -> URL
        original_codes_by_std_2 = {}  # Map mã chuẩn hóa -> mã gốc
        
        for i, url in enumerate(product_urls_2):
            if i % 10 == 0:
                progress = 75 + int((i / len(product_urls_2)) * 5)
                socketio.emit('progress_update', {
                    'percent': progress, 
                    'message': f'Đang trích xuất mã sản phẩm từ HaiphongTech ({i}/{len(product_urls_2)})...'
                })
            
            # Trích xuất mã sản phẩm từ URL hoặc nội dung trang
            original_code = extract_product_code_from_url(url)
            if original_code:
                # Function extract_product_code_from_url đã gọi standardize_product_code
                product_codes_2.append(original_code)
                product_urls_by_code_2[original_code] = url
                original_codes_by_std_2[original_code] = original_code
                print(f"Sản phẩm HaiphongTech: {original_code} - {url}")
        
        socketio.emit('progress_update', {
            'percent': 80, 
            'message': f'Đã tìm thấy {len(product_codes_1)} mã sản phẩm từ BAA.vn và {len(product_codes_2)} mã sản phẩm từ HaiphongTech. Đang so sánh...'
        })
        
        # So sánh các mã sản phẩm (đã được chuẩn hóa)
        socketio.emit('progress_update', {
            'percent': 80, 
            'message': f'Đã tìm thấy {len(product_codes_1)} mã sản phẩm từ BAA.vn và {len(product_codes_2)} mã sản phẩm từ HaiphongTech. Đang so sánh...'
        })
        
        # Tạo tập hợp (set) của tất cả mã sản phẩm
        all_product_codes = set(product_codes_1).union(set(product_codes_2))
        all_product_codes = sorted(list(all_product_codes))
        
        # Tạo các tập hợp để phân loại
        # 1. Các mã sản phẩm có trên cả hai website
        common_codes = list(set(product_codes_1).intersection(set(product_codes_2)))
        common_codes.sort()
        
        # 2. Các mã sản phẩm chỉ có trên BAA.vn
        unique_codes_baa = list(set(product_codes_1) - set(product_codes_2))
        unique_codes_baa.sort()
        
        # 3. Các mã sản phẩm chỉ có trên HaiphongTech
        unique_codes_hpt = list(set(product_codes_2) - set(product_codes_1))
        unique_codes_hpt.sort()
        
        # Tạo danh sách URLs tương ứng
        unique_urls_baa = [product_urls_by_code_1.get(code, '') for code in unique_codes_baa]
        unique_urls_hpt = [product_urls_by_code_2.get(code, '') for code in unique_codes_hpt]
        
        socketio.emit('progress_update', {
            'percent': 85, 
            'message': f'Phân tích hoàn tất. Tìm thấy {len(all_product_codes)} mã sản phẩm khác nhau. Đang tạo báo cáo...'
        })
        
        # Tạo báo cáo Excel
        # 1. Tạo một DataFrame chứa tất cả mã sản phẩm và URLs tương ứng từ hai website
        print(f"Tổng số mã sản phẩm: {len(all_product_codes)}")
        print(f"Số mã sản phẩm chung: {len(common_codes)}")
        print(f"Số mã sản phẩm chỉ có trên BAA.vn: {len(unique_codes_baa)}")
        print(f"Số mã sản phẩm chỉ có trên HaiphongTech: {len(unique_codes_hpt)}")
        
        comparison_data = []
        for code in all_product_codes:
            # URL từ BAA.vn
            url_baa = product_urls_by_code_1.get(code, '')
            # URL từ HaiphongTech
            url_hpt = product_urls_by_code_2.get(code, '')
            # Trạng thái: Chung, Chỉ BAA.vn, Chỉ HaiphongTech
            if code in common_codes:
                status = "Chung"
            elif code in unique_codes_baa:
                status = "Chỉ BAA.vn"
            else:
                status = "Chỉ HaiphongTech"
            
            comparison_data.append({
                'Mã sản phẩm': code,
                'URL BAA.vn': url_baa,
                'URL HaiphongTech': url_hpt,
                'Trạng thái': status
            })
        
        # Tạo output directory
        output_dir = os.path.join(current_app.static_folder, 'output_compare')
        os.makedirs(output_dir, exist_ok=True)
        
        # Lưu báo cáo Excel
        excel_path = os.path.join(output_dir, 'product_comparison.xlsx')
        df = pd.DataFrame(comparison_data)
        df.to_excel(excel_path, index=False)
        
        # Lưu danh sách URLs sản phẩm độc nhất vào file TXT
        # 1. URLs từ BAA.vn không có trên HaiphongTech
        unique_baa_txt_path = os.path.join(output_dir, 'unique_baa_urls.txt')
        with open(unique_baa_txt_path, 'w', encoding='utf-8') as f:
            for i, url in enumerate(unique_urls_baa):
                if url:
                    f.write(f"{unique_codes_baa[i]}\t{url}\n")
        
        # 2. URLs từ HaiphongTech không có trên BAA.vn
        unique_hpt_txt_path = os.path.join(output_dir, 'unique_hpt_urls.txt')
        with open(unique_hpt_txt_path, 'w', encoding='utf-8') as f:
            for i, url in enumerate(unique_urls_hpt):
                if url:
                    f.write(f"{unique_codes_hpt[i]}\t{url}\n")
        
        # Nén các file kết quả lại thành ZIP
        zip_path = os.path.join(current_app.static_folder, 'downloads', 'product_comparison.zip')
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            # Thêm file Excel
            zipf.write(excel_path, os.path.basename(excel_path))
            # Thêm các file TXT
            zipf.write(unique_baa_txt_path, os.path.basename(unique_baa_txt_path))
            zipf.write(unique_hpt_txt_path, os.path.basename(unique_hpt_txt_path))
            
            socketio.emit('progress_update', {
                'percent': 100, 
            'message': 'So sánh danh mục sản phẩm hoàn tất!'
        })
        
        # Trả về URL để tải xuống file ZIP
        download_url = url_for('main.download_file', filename='product_comparison.zip')
        return render_template('index.html', download_url=download_url, success_message="So sánh danh mục sản phẩm thành công!")
            
    except Exception as e:
        error_msg = f"Lỗi: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=error_msg)

@main_bp.route('/compare-product-codes', methods=['POST'])
def compare_product_codes():
    try:
        # Kiểm tra file đầu tiên (HPT)
        if 'excel_file_1' not in request.files:
            return render_template('index.html', error="Vui lòng tải lên file Excel HPT")
        
        excel_file_1 = request.files['excel_file_1']
        
        if excel_file_1.filename == '':
            return render_template('index.html', error="Không có file HPT nào được chọn")
        
        # Kiểm tra các file thứ hai
        excel_files_2 = request.files.getlist('excel_file_2')
        
        if not excel_files_2 or all(f.filename == '' for f in excel_files_2):
            return render_template('index.html', error="Vui lòng tải lên ít nhất một file Excel thứ hai")
        
        # Loại bỏ các file trống
        excel_files_2 = [f for f in excel_files_2 if f.filename != '']
        
        if not excel_files_2:
            return render_template('index.html', error="Không có file Excel thứ hai nào được chọn")
        
        allowed_extensions = set(['xlsx', 'xls', 'csv'])
        
        # Kiểm tra file HPT
        if not excel_file_1.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
            return render_template('index.html', error="File HPT phải có định dạng .xlsx, .xls hoặc .csv")
        
        # Kiểm tra các file thứ hai
        for excel_file in excel_files_2:
            if not excel_file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                return render_template('index.html', error=f"File {excel_file.filename} phải có định dạng .xlsx, .xls hoặc .csv")
        
        # Kiểm tra tùy chọn phân loại danh mục
        categorize_results = 'categorize_results' in request.form
        
        # Thông báo tiến trình
        socketio.emit('progress_update', {
            'percent': 10, 
            'message': f'Đang đọc dữ liệu từ {1 + len(excel_files_2)} file Excel...'
        })
        
        # Lưu các file tạm thời
        temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        file1_path = os.path.join(temp_dir, secure_filename(excel_file_1.filename))
        excel_file_1.save(file1_path)
        
        file2_paths = []
        for i, excel_file in enumerate(excel_files_2):
            file2_path = os.path.join(temp_dir, secure_filename(excel_file.filename))
            excel_file.save(file2_path)
            file2_paths.append(file2_path)
        
        socketio.emit('progress_update', {
            'percent': 30, 
            'message': f'Đang so sánh mã sản phẩm giữa file HPT và {len(file2_paths)} file khác...'
        })
        
        # Gọi hàm so sánh từ module product_comparison
        from app.product_comparison import compare_product_codes
        output_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Truyền danh sách file2_paths vào hàm so sánh
        report_path = compare_product_codes(file1_path, file2_paths, output_dir, categorize_results)
        
        if report_path and isinstance(report_path, str) and os.path.exists(report_path):
            socketio.emit('progress_update', {
                'percent': 100, 
                'message': 'Đã hoàn tất so sánh mã sản phẩm!'
            })
            
            # Tạo đường dẫn tải xuống
            report_filename = os.path.basename(report_path)
            download_url = url_for('main.download_file', filename=f'reports/{report_filename}')
            
            success_message = f"Đã so sánh xong mã sản phẩm từ file HPT với {len(file2_paths)} file khác!"
            if categorize_results:
                success_message += " Kết quả đã được phân loại theo danh mục sản phẩm."
            
            return render_template('index.html', 
                                 success_message=success_message,
                                 download_url=download_url)
        else:
            error_msg = "Có lỗi xảy ra khi so sánh mã sản phẩm"
            if isinstance(report_path, dict) and 'error' in report_path:
                error_msg = f"Lỗi: {report_path.get('error', '')}"
            return render_template('index.html', error=error_msg)
            
    except Exception as e:
        print(f"Lỗi: {str(e)}")
        print(traceback.format_exc())
        socketio.emit('progress_update', {'percent': 0, 'message': f'Đã xảy ra lỗi: {str(e)}'})
        return render_template('index.html', error=f"Lỗi khi so sánh mã sản phẩm: {str(e)}")

# Hàm trích xuất mã sản phẩm từ URL hoặc nội dung trang
def extract_product_code_from_url(url):
    try:
        # Pattern để trích xuất mã sản phẩm từ URL
        # 1. Pattern cho URL BAA.vn: ... /bo-chuyen-doi-nhiet-do-autonics-kt-502h0_60329
        # 2. Pattern cho URL HaiphongTech: ... /bo-dieu-khien-nhiet-do-e5cc-rx2asm-800

        # Cố gắng trích xuất mã từ URL BAA.vn dạng URL sản phẩm
        baa_pattern = r'(?:san-pham\/.*?)(?:autonics|omron|ls|mitsubishi|fuji|idec|keyence|optex-fa|optex|panasonic|hanyoung|honeywell|koino|ckd|cikachi|chint|mean-well|weintek|eaton|lc|lsis|azbil|riko|coel|power|siemens|schneider|delta|hager|tele|taian|tend|socomec|leuze|sick|takex|heyi|trans|fotek|anly|winstar|nikkon|apator|bulgin|finder|benedict|protek|sanyu|nowox|anly|abb|ginice|wattstopper|bristoleye|hubbell|contrinex|elco|turck|banner|wenglor|rockwell|omron-movisens|st|oez|hfr|himel|yaskawa|futek|metz-connect|contactor|relay|circuit-protector|autonics-counter-timer|autonics-digital-panel-meter|panasonic-measurement).*?[_-]([a-zA-Z0-9\-]+?)(?:_|\-|$)'
        match = re.search(baa_pattern, url, re.IGNORECASE)
        if match:
            product_code = match.group(1)
            # Chuẩn hóa mã sản phẩm
            return standardize_product_code(product_code)
        
        # Cố gắng trích xuất mã từ URL HaiphongTech dạng URL sản phẩm
        hpt_pattern = r'(?:.*?\/)((?:[a-zA-Z0-9]+)(?:-[a-zA-Z0-9]+)+)(?:-(?:gia)|$)'
        match = re.search(hpt_pattern, url, re.IGNORECASE)
        if match:
            product_code = match.group(1)
            # Chuẩn hóa mã sản phẩm
            return standardize_product_code(product_code)
        
        # Nếu không thể trích xuất từ URL, thử tải nội dung trang và trích xuất từ HTML
        from app.crawler import get_html_content
        
        html_content = get_html_content(url)
        if not html_content:
            return None
        
        # Tạo đối tượng BeautifulSoup
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Tìm kiếm mã sản phẩm trong các thẻ meta, h1, div, span, etc.
        # 1. Thử tìm trong thẻ có class "product__symbol__value" (BAA.vn)
        product_symbol = soup.select_one('.product__symbol__value, .product-code, .sku, [itemprop="sku"], .product_code')
        if product_symbol and product_symbol.text.strip():
            product_code = product_symbol.text.strip()
            return standardize_product_code(product_code)
        
        # 2. Thử tìm trong tiêu đề sản phẩm
        product_title = soup.select_one('h1.product__name, h1.product-title, h1.product-name, h1')
        if product_title:
            title_text = product_title.text.strip()
            # Tìm mã sản phẩm trong tiêu đề - giả định mã có dạng XX-XXXX hoặc XXXXX
            code_pattern = r'([A-Z0-9\-]{3,}(?:-[A-Z0-9\-]+)+)'
            code_matches = re.findall(code_pattern, title_text)
            if code_matches:
                for code in code_matches:
                    # Kiểm tra xem code có hợp lệ không (ví dụ: không phải là các từ thông thường)
                    if not re.search(r'^(THE|AND|FOR|WITH|FROM)$', code, re.IGNORECASE):
                        return standardize_product_code(code)
        
        return None
        
    except Exception as e:
        print(f"Lỗi khi trích xuất mã sản phẩm từ URL {url}: {str(e)}")
        return None

# Hàm chuẩn hóa mã sản phẩm
def standardize_product_code(code):
    if not code:
        return None
    
    # Chuyển về chữ hoa
    std_code = code.upper()
    
    # Xóa các ký tự đặc biệt ở đầu và cuối
    std_code = std_code.strip('-_.,;:\'"\t ')
    
    # Xóa các text không cần thiết
    unwanted_texts = ['AUTONICS', 'OMRON', 'MITSUBISHI', 'PANASONIC', 'LS', 'GIÁ', 'PRICE', 'BÁO GIÁ', 'MUA', 'MUA HÀNG']
    for text in unwanted_texts:
        std_code = std_code.replace(text, '')
    
    # Xóa khoảng trắng
    std_code = std_code.strip()
    
    # Nếu code quá ngắn, trả về None
    if len(std_code) < 3:
        return None
    
    return std_code

@main_bp.route('/download-baa-images', methods=['POST'])
def download_baa_images():
    """
    Tải ảnh sản phẩm từ BAA.vn và chuyển sang WebP kích thước gốc
    """
    try:
        if 'product_links_file' not in request.files:
            flash('Không tìm thấy file!', 'error')
            return redirect(url_for('main.index'))
        
        file = request.files['product_links_file']
        
        if file.filename == '':
            flash('Không có file nào được chọn!', 'error')
            return redirect(url_for('main.index'))
            
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            flash('Chỉ cho phép file .txt!', 'error')
            return redirect(url_for('main.index'))
        
        # Đọc nội dung file
        content = file.read().decode('utf-8')
        
        # Lấy danh sách URL sản phẩm
        product_urls = [url.strip() for url in content.strip().split('\n') if url.strip()]
        
        if not product_urls:
            flash('File không chứa URL sản phẩm nào!', 'error')
            return redirect(url_for('main.index'))
            
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 0, 
            'message': f'Bắt đầu xử lý {len(product_urls)} URL sản phẩm...'
        })
        
        # Tạo thư mục đầu ra
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'baa_images_{timestamp}')
        
        # Tải ảnh sản phẩm - Sử dụng phiên bản đã sửa lỗi để xử lý đúng đường dẫn ảnh
        results = download_baa_product_images_fixed(product_urls, output_folder)
        
        # Nén thư mục ảnh thành file ZIP
        zip_filename = f'baa_images_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        
        if utils.create_zip_from_folder(output_folder, zip_path):
            # Tạo URL tải xuống
            download_url = url_for('main.download_file', filename=zip_filename)
            
            # Tạo URL xem ảnh
            view_images_url = url_for('main.view_baa_images', folder=f'baa_images_{timestamp}')
            
            # Thông báo thành công
            success_message = f"Đã tải {results['success']}/{results['total']} ảnh sản phẩm"
            if results['failed'] > 0:
                success_message += f", {results['failed']} thất bại"
            success_message += f" (Tổng số ảnh: {len(results['image_paths'])})"
            
            flash(success_message, 'success')
            
            # Gửi thông báo hoàn thành
            socketio.emit('progress_update', {
                'percent': 100,
                'message': 'Hoàn thành!'
            })
            
            return render_template('index.html', download_url=download_url, view_images_url=view_images_url)
        else:
            flash('Lỗi khi tạo file ZIP!', 'error')
            return redirect(url_for('main.index'))
        
    except Exception as e:
        error_message = str(e)
        print(f"Lỗi khi tải ảnh BAA: {error_message}")
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/view-baa-images/<folder>')
def view_baa_images(folder):
    """Hiển thị danh sách các ảnh BAA đã tải xuống"""
    try:
        # Đường dẫn đến thư mục chứa ảnh
        images_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], folder)
        
        if not os.path.exists(images_folder) or not os.path.isdir(images_folder):
            flash('Thư mục ảnh không tồn tại!', 'error')
            return redirect(url_for('main.index'))
        
        # Tìm tất cả các file ảnh trong thư mục
        all_images = []
        for file in os.listdir(images_folder):
            if file.endswith('.webp'):
                # Lấy mã sản phẩm từ tên file
                product_code = os.path.splitext(file)[0]
                
                # Đường dẫn tương đối đến file ảnh
                rel_path = os.path.join(folder, file)
                
                all_images.append({
                    'path': rel_path,
                    'code': product_code,
                    'url': url_for('main.view_image', image_path=rel_path)
                })
        
        # Sắp xếp ảnh theo mã sản phẩm
        all_images.sort(key=lambda x: x['code'])
        
        return render_template('view_baa_images.html', images=all_images, folder=folder)
        
    except Exception as e:
        error_message = str(e)
        flash(f'Lỗi khi xem ảnh: {error_message}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/filter-products', methods=['POST'])
@progress_tracker(name="Lọc sản phẩm", total_steps=100, verbose=True)
def filter_products(progress: TerminalProgressBar):
    """
    Lọc danh sách mã sản phẩm từ file Excel dựa trên danh sách mã cần xóa
    """
    try:
        progress.update(5, "Kiểm tra thông tin đầu vào")
        
        # Kiểm tra nếu có cả danh sách mã và file Excel
        if 'product_codes' not in request.form or 'excel_file' not in request.files:
            progress.error('Vui lòng nhập danh sách mã sản phẩm và tải lên file Excel!')
            flash('Vui lòng nhập danh sách mã sản phẩm và tải lên file Excel!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        # Lấy danh sách mã sản phẩm từ form
        product_codes_text = request.form['product_codes']
        if not product_codes_text.strip():
            progress.error('Danh sách mã sản phẩm không được để trống!')
            flash('Danh sách mã sản phẩm không được để trống!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        progress.update(10, "Đang xử lý danh sách mã sản phẩm")
        
        # Xử lý danh sách mã sản phẩm (mỗi mã trên một dòng)
        product_codes = [code.strip() for code in product_codes_text.strip().split('\n') if code.strip()]
        
        # Chuẩn hóa danh sách mã sản phẩm (loại bỏ khoảng trắng, chuyển thành chữ thường để so sánh)
        normalized_product_codes = [str(code).strip().lower() for code in product_codes]
        
        print(f"DEBUG: Đã nhận {len(product_codes)} mã sản phẩm cần lọc")
        
        progress.update(15, f"Đã xử lý {len(product_codes)} mã sản phẩm cần lọc", "Đang kiểm tra file Excel")
        
        # Kiểm tra file Excel
        excel_file = request.files['excel_file']
        if excel_file.filename == '':
            progress.error('Không có file Excel nào được chọn!')
            flash('Không có file Excel nào được chọn!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        if not allowed_file(excel_file.filename, ALLOWED_EXTENSIONS_EXCEL):
            progress.error('Chỉ chấp nhận file .xlsx hoặc .xls!')
            flash('Chỉ chấp nhận file .xlsx hoặc .xls!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        # Thông báo tiến trình (kết hợp với socketio)
        socketio.emit('progress_update', {
            'percent': 20, 
            'message': f'Đang xử lý {len(product_codes)} mã sản phẩm cần lọc...'
        })
        
        progress.update(20, "Đang lưu file Excel tạm thời", f"File: {excel_file.filename}")
        
        # Lưu file tạm thời
        temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        input_excel_path = os.path.join(temp_dir, secure_filename(excel_file.filename))
        excel_file.save(input_excel_path)
        
        print(f"DEBUG: Đã lưu file Excel tạm thời tại: {input_excel_path}")
        
        progress.update(30, "Đang đọc file Excel", "Đang phân tích cấu trúc file")
        
        # Tạo progress bar con cho việc đọc Excel
        excel_progress = create_child_progress(progress, "Đọc file Excel", 20)
        
        # Đọc file Excel với pandas
        df = None
        try:
            # Thử đọc file với nhiều sheet khác nhau
            xls = pd.ExcelFile(input_excel_path)
            sheet_names = xls.sheet_names
            print(f"DEBUG: Các sheet trong file Excel: {sheet_names}")
            
            excel_progress.update(5, f"Tìm thấy {len(sheet_names)} sheets")
            
            if len(sheet_names) > 0:
                # Nếu file có sheet "Tổng hợp sản phẩm", sử dụng sheet đó
                if "Tổng hợp sản phẩm" in sheet_names:
                    df = pd.read_excel(input_excel_path, sheet_name="Tổng hợp sản phẩm")
                    print(f"DEBUG: Đọc dữ liệu từ sheet 'Tổng hợp sản phẩm'")
                else:
                    # Sử dụng sheet đầu tiên
                    df = pd.read_excel(input_excel_path, sheet_name=sheet_names[0])
                    print(f"DEBUG: Đọc dữ liệu từ sheet '{sheet_names[0]}'")
            else:
                df = pd.read_excel(input_excel_path)
            
            excel_progress.update(15, f"Đã đọc {len(df)} dòng dữ liệu")
            
            # Chuyển tất cả các cột có kiểu object sang string để tránh lỗi NaN
            for col in df.select_dtypes(include=['object']).columns:
                df[col] = df[col].astype(str)
            
            excel_progress.complete("success", f"Đọc Excel thành công", {
                "Sheets": len(sheet_names),
                "Dòng dữ liệu": len(df),
                "Cột": len(df.columns)
            })
            
            socketio.emit('progress_update', {
                'percent': 50, 
                'message': f'Đã đọc file Excel với {len(df)} dòng. Đang lọc dữ liệu...'
            })
            
            # Log thông tin về cấu trúc của DataFrame để debug
            print(f"DEBUG: Cấu trúc DataFrame: {df.shape}, Cột: {df.columns.tolist()}")
            
        except Exception as e:
            excel_progress.error(f'Lỗi khi đọc file Excel: {str(e)}')
            error_message = f'Lỗi khi đọc file Excel: {str(e)}'
            print(f"DEBUG: {error_message}")
            traceback.print_exc()
            progress.error(error_message, traceback.format_exc())
            flash(error_message, 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        # Kiểm tra nếu DataFrame trống hoặc không có dữ liệu
        if df is None or df.empty:
            progress.error('File Excel không chứa dữ liệu!')
            flash('File Excel không chứa dữ liệu!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        # Kiểm tra nếu DataFrame không có ít nhất 2 cột (cần cột B)
        if df.shape[1] < 2:
            progress.error('File Excel phải có ít nhất 2 cột (cột B chứa mã sản phẩm)!')
            flash('File Excel phải có ít nhất 2 cột (cột B chứa mã sản phẩm)!', 'error')
            return redirect(url_for('main.index', _anchor='filter-products-tab'))
        
        progress.update(60, "Đang xác định cột chứa mã sản phẩm", f"Phân tích {len(df.columns)} cột")
        
        # Xác định cột chứa mã sản phẩm
        product_code_column = 1  # Mặc định là cột B (index 1)
        
        # Kiểm tra nếu có cột tên là "Mã sản phẩm"
        if "Mã sản phẩm" in df.columns:
            product_code_column = df.columns.get_loc("Mã sản phẩm")
            print(f"DEBUG: Đã tìm thấy cột 'Mã sản phẩm' ở vị trí {product_code_column}")
        
        # Lấy cột chứa mã sản phẩm
        df_product_codes = df.iloc[:, product_code_column].astype(str).str.strip()
        
        # Chuẩn hóa mã sản phẩm từ file Excel (loại bỏ khoảng trắng, chuyển thành chữ thường để so sánh)
        normalized_df_codes = df_product_codes.str.lower()
        
        progress.update(70, "Đang lọc sản phẩm", "Tìm các mã cần xóa trong Excel")
        
        # Tạo progress bar con cho việc lọc
        filter_progress = create_child_progress(progress, "Lọc dữ liệu", 20)
        
        # Tạo mask để lọc các hàng có mã sản phẩm nằm trong danh sách cần xóa
        rows_to_remove = normalized_df_codes.isin(normalized_product_codes)
        
        # Tạo DataFrame chứa các dòng sẽ bị xóa
        removed_df = df[rows_to_remove].copy()
        
        # Đếm số dòng sẽ bị xóa
        removed_count = rows_to_remove.sum()
        print(f"DEBUG: Đã tìm thấy {removed_count} mã sản phẩm cần xóa")
        
        filter_progress.update(10, f"Tìm thấy {removed_count} mã cần xóa")
        
        # In một số mã sản phẩm đã tìm thấy để debug
        if removed_count > 0:
            matched_codes = df[rows_to_remove].iloc[:, product_code_column].tolist()
            print(f"DEBUG: Một số mã sản phẩm đã tìm thấy (tối đa 5): {matched_codes[:5]}")
        
        socketio.emit('progress_update', {
            'percent': 80, 
            'message': f'Đã tìm thấy {removed_count} mã sản phẩm cần xóa. Đang tạo báo cáo...'
        })
        
        # Tạo DataFrame mới không chứa các hàng đã lọc
        filtered_df = df[~rows_to_remove]
        
        filter_progress.update(20, f"Đã lọc thành {len(filtered_df)} dòng")
        
        # Thêm cột "Mã sản phẩm đã lọc" vào vị trí cột thứ 3 (index 2)
        if filtered_df.shape[1] <= 2:
            # Nếu chỉ có 2 cột hoặc ít hơn, thêm cột mới
            filtered_df.insert(2, 'Mã sản phẩm đã lọc', '')
        else:
            # Đổi tên cột thứ 3 thành "Mã sản phẩm đã lọc" và xóa giá trị hiện có
            filtered_df.rename(columns={filtered_df.columns[2]: 'Mã sản phẩm đã lọc'}, inplace=True)
            filtered_df.iloc[:, 2] = ''
        
        # Thêm danh sách mã sản phẩm đã lọc vào cột C chỉ cho dòng đầu tiên
        if not filtered_df.empty:
            filtered_df.iloc[0, 2] = ', '.join([str(code).strip() for code in product_codes])
        
        filter_progress.complete("success", "Lọc dữ liệu hoàn tất", {
            "Dòng ban đầu": len(df),
            "Dòng đã xóa": removed_count,
            "Dòng còn lại": len(filtered_df)
        })
        
        progress.update(90, "Đang tạo file Excel kết quả", "Lưu dữ liệu đã lọc")
        
        # Tạo tên file output
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_filename = f"filtered_products_{timestamp}.xlsx"
        output_path = os.path.join(current_app.config['UPLOAD_FOLDER'], output_filename)
        
        print(f"DEBUG: Đang tạo file Excel: {output_path}")
        
        # Lưu kết quả vào file Excel mới với các sheet riêng
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            # Sheet chính chứa dữ liệu đã lọc
            filtered_df.to_excel(writer, sheet_name='Dữ liệu đã lọc', index=False)
            print(f"DEBUG: Đã tạo sheet 'Dữ liệu đã lọc' với {len(filtered_df)} dòng")
            
            # Sheet thứ hai chứa dữ liệu đã xóa
            if not removed_df.empty:
                removed_df.to_excel(writer, sheet_name='Mã đã xóa', index=False)
                print(f"DEBUG: Đã tạo sheet 'Mã đã xóa' với {len(removed_df)} dòng")
            
            # Tạo sheet tổng hợp
            summary_data = {
                'Thông tin': [
                    'Tổng số dòng ban đầu', 
                    'Số dòng đã xóa', 
                    'Số dòng còn lại',
                    'Số mã sản phẩm cần lọc',
                    'Số mã sản phẩm đã tìm thấy'
                ],
                'Số lượng': [
                    len(df),
                    removed_count,
                    len(filtered_df),
                    len(product_codes),
                    removed_count
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Tổng hợp', index=False)
            print(f"DEBUG: Đã tạo sheet 'Tổng hợp'")
            
            # Format các sheet
            workbook = writer.book
            for sheet in writer.sheets.values():
                sheet.set_column('A:Z', 18)
        
        socketio.emit('progress_update', {
            'percent': 100, 
            'message': 'Hoàn thành lọc dữ liệu!'
        })
        
        # Tạo URL để tải xuống file kết quả
        download_url = url_for('main.download_file', filename=output_filename)
        
        # Lưu trữ session để chuyển về đúng tab sau khi tải
        session['active_tab'] = 'filter-products-tab'
        
        # Thông báo kết quả
        success_message = f"Đã lọc thành công! Đã xóa {removed_count}/{len(df)} dòng dữ liệu. Danh sách mã sản phẩm đã lọc được lưu vào cột C."
        return render_template('index.html', 
                              download_url=download_url, 
                              success_message=success_message, 
                              active_tab='filter-products-tab')
        
    except Exception as e:
        error_message = str(e)
        print(f"DEBUG: Lỗi khi lọc sản phẩm: {error_message}")
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index', _anchor='filter-products-tab'))

@main_bp.route('/extract-prices', methods=['POST'])
def extract_prices():
    """
    Trích xuất giá sản phẩm từ danh sách các URL sản phẩm
    """
    try:
        if 'product_links_file' not in request.files:
            flash('Không tìm thấy file!', 'error')
            return redirect(url_for('main.index'))
        
        file = request.files['product_links_file']
        
        if file.filename == '':
            flash('Không có file nào được chọn!', 'error')
            return redirect(url_for('main.index'))
            
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            flash('Chỉ cho phép file .txt!', 'error')
            return redirect(url_for('main.index'))
        
        # Đọc nội dung file
        content = file.read().decode('utf-8')
        
        # Lấy danh sách URL sản phẩm
        product_urls = [url.strip() for url in content.strip().split('\n') if url.strip()]
        
        if not product_urls:
            flash('File không chứa URL sản phẩm nào!', 'error')
            return redirect(url_for('main.index'))
            
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 0, 
            'message': f'Bắt đầu trích xuất giá cho {len(product_urls)} URL sản phẩm...'
        })
        
        # Trích xuất thông tin sản phẩm, bao gồm giá
        product_data = []
        required_fields = ['STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Giá', 'URL']
        
        for index, url in enumerate(product_urls, 1):
            socketio.emit('progress_update', {
                'percent': int((index / len(product_urls)) * 70), 
                'message': f'Đang trích xuất giá từ URL {index}/{len(product_urls)}...'
            })
            
            try:
                product_info = extract_product_info(url, required_fields=required_fields, index=index)
                product_data.append(product_info)
            except Exception as e:
                print(f"Lỗi khi trích xuất URL {url}: {str(e)}")
                # Thêm dữ liệu tối thiểu nếu có lỗi
                product_data.append({
                    'STT': index,
                    'Mã sản phẩm': '',
                    'Tên sản phẩm': '',
                    'Giá': 'Lỗi: ' + str(e),
                    'URL': url
                })
        
        # Tạo DataFrame từ dữ liệu đã thu thập
        df = pd.DataFrame(product_data)
        
        # Tạo tên file output
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_filename = f"product_prices_{timestamp}.xlsx"
        output_path = os.path.join(current_app.config['UPLOAD_FOLDER'], output_filename)
        
        # Lưu kết quả vào file Excel
        socketio.emit('progress_update', {
            'percent': 80, 
            'message': 'Đang tạo file Excel...'
        })
        
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Giá sản phẩm', index=False)
            
            # Format sheet
            workbook = writer.book
            worksheet = writer.sheets['Giá sản phẩm']
            
            # Định dạng cột
            worksheet.set_column('A:A', 5)   # STT
            worksheet.set_column('B:B', 20)  # Mã sản phẩm
            worksheet.set_column('C:C', 40)  # Tên sản phẩm
            worksheet.set_column('D:D', 20)  # Giá
            worksheet.set_column('E:E', 50)  # URL
            
            # Tạo định dạng
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'align': 'center',
                'border': 1,
                'bg_color': '#D7E4BC'
            })
            
            # Áp dụng định dạng cho tiêu đề
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
        
        socketio.emit('progress_update', {
            'percent': 100, 
            'message': 'Hoàn thành trích xuất giá sản phẩm!'
        })
        
        # Tạo URL để tải xuống file kết quả
        download_url = url_for('main.download_file', filename=output_filename)
        
        # Thông báo kết quả
        success_message = f"Đã trích xuất giá cho {len(product_data)} sản phẩm thành công!"
        return render_template('index.html', download_url=download_url, success_message=success_message)
        
    except Exception as e:
        error_message = str(e)
        print(f"Lỗi khi trích xuất giá sản phẩm: {error_message}")
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index')) 

@main_bp.route('/extract-only-prices', methods=['POST'])
def extract_only_prices():
    """
    Chỉ trích xuất giá sản phẩm từ danh sách các URL sản phẩm
    """
    try:
        if 'product_links_file' not in request.files:
            flash('Không tìm thấy file!', 'error')
            return redirect(url_for('main.index'))
        
        file = request.files['product_links_file']
        
        if file.filename == '':
            flash('Không có file nào được chọn!', 'error')
            return redirect(url_for('main.index'))
            
        if not allowed_file(file.filename, ALLOWED_EXTENSIONS_TXT):
            flash('Chỉ cho phép file .txt!', 'error')
            return redirect(url_for('main.index'))
        
        # Đọc nội dung file
        content = file.read().decode('utf-8')
        
        # Lấy danh sách URL sản phẩm
        product_urls = [url.strip() for url in content.strip().split('\n') if url.strip()]
        
        if not product_urls:
            flash('File không chứa URL sản phẩm nào!', 'error')
            return redirect(url_for('main.index'))
            
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {
            'percent': 0, 
            'message': f'Bắt đầu trích xuất giá cho {len(product_urls)} URL sản phẩm...'
        })
        
        # Chỉ trích xuất giá và mã sản phẩm
        product_data = []
        
        for index, url in enumerate(product_urls, 1):
            socketio.emit('progress_update', {
                'percent': int((index / len(product_urls)) * 70), 
                'message': f'Đang trích xuất giá từ URL {index}/{len(product_urls)}...'
            })
            
            try:
                # Sử dụng hàm chuyên biệt để chỉ lấy mã và giá
                product_info = extract_product_price(url, index=index)
                product_data.append(product_info)
            except Exception as e:
                print(f"Lỗi khi trích xuất URL {url}: {str(e)}")
                # Thêm dữ liệu tối thiểu nếu có lỗi
                product_data.append({
                    'STT': index,
                    'URL': url,
                    'Mã sản phẩm': '',
                    'Giá': 'Lỗi: ' + str(e)
                })
        
        # Tạo DataFrame từ dữ liệu đã thu thập
        df = pd.DataFrame(product_data)
        
        # Tạo tên file output
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_filename = f"product_only_prices_{timestamp}.xlsx"
        output_path = os.path.join(current_app.config['UPLOAD_FOLDER'], output_filename)
        
        # Lưu kết quả vào file Excel
        socketio.emit('progress_update', {
            'percent': 80, 
            'message': 'Đang tạo file Excel...'
        })
        
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Giá sản phẩm', index=False)
            
            # Format sheet
            workbook = writer.book
            worksheet = writer.sheets['Giá sản phẩm']
            
            # Định dạng cột
            worksheet.set_column('A:A', 5)   # STT
            worksheet.set_column('B:B', 50)  # URL
            worksheet.set_column('C:C', 20)  # Mã sản phẩm
            worksheet.set_column('D:D', 20)  # Giá
            
            # Tạo định dạng
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'align': 'center',
                'border': 1,
                'bg_color': '#D7E4BC'
            })
            
            # Áp dụng định dạng cho tiêu đề
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
        
        socketio.emit('progress_update', {
            'percent': 100, 
            'message': 'Hoàn thành trích xuất giá sản phẩm!'
        })
        
        # Tạo URL để tải xuống file kết quả
        download_url = url_for('main.download_file', filename=output_filename)
        
        # Thông báo kết quả
        success_message = f"Đã trích xuất giá cho {len(product_data)} sản phẩm thành công!"
        return render_template('index.html', download_url=download_url, success_message=success_message)
        
    except Exception as e:
        error_message = str(e)
        print(f"Lỗi khi trích xuất giá sản phẩm: {error_message}")
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/extract-category-links', methods=['POST'])
def extract_category_links_separate():
    """
    Trích xuất liên kết sản phẩm từ nhiều danh mục và sắp xếp thành thư mục riêng
    """
    try:
        # Kiểm tra xem có dữ liệu danh mục không
        if 'category_urls' not in request.form:
            flash('Vui lòng nhập danh sách URL danh mục!', 'error')
            return redirect(url_for('main.index'))
            
        # Lấy danh sách URL danh mục từ form
        category_urls_text = request.form['category_urls']
        if not category_urls_text.strip():
            flash('Danh sách URL danh mục không được để trống!', 'error')
            return redirect(url_for('main.index'))
        
        # Tách thành danh sách URL, bỏ qua dòng trống
        raw_urls = category_urls_text.strip().split('\n')
        urls = [url.strip() for url in raw_urls if url.strip()]
        
        # Lọc các URL hợp lệ
        valid_urls = []
        invalid_urls = []
        
        # Gửi thông báo bắt đầu
        socketio.emit('progress_update', {'percent': 0, 'message': 'Đang kiểm tra URL danh mục...'})
        
        # Kiểm tra các URL
        for url in urls:
            if utils.is_valid_url(url) and is_category_url(url):
                valid_urls.append(url)
            else:
                invalid_urls.append(url)
        
        if not valid_urls:
            flash('Không có URL danh mục hợp lệ!', 'error')
            return redirect(url_for('main.index'))
            
        # Gửi thông báo cập nhật
        socketio.emit('progress_update', {'percent': 5, 'message': f'Đã tìm thấy {len(valid_urls)} URL danh mục hợp lệ'})
        
        # Tạo thư mục chính để lưu kết quả
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        result_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], f'category_products_{timestamp}')
        os.makedirs(result_dir, exist_ok=True)
        
        # Tạo tệp txt để lưu tất cả các liên kết
        all_products_file = os.path.join(result_dir, 'all_product_links.txt')
        all_product_links = []
        
        # Xử lý từng URL danh mục riêng biệt
        category_info = []  # Lưu thông tin về mỗi danh mục
        
        for i, category_url in enumerate(valid_urls):
            try:
                progress = 5 + int((i / len(valid_urls)) * 85)
                socketio.emit('progress_update', {
                    'percent': progress, 
                    'message': f'Đang xử lý danh mục {i+1}/{len(valid_urls)}: {category_url}'
                })
                
                # Trích xuất tên danh mục từ URL
                parsed_url = urlparse(category_url)
                url_path = parsed_url.path
                
                # Lấy phần cuối của URL làm tên danh mục
                category_name = url_path.strip('/').split('/')[-1]
                # Loại bỏ phần ID số từ tên danh mục nếu có
                category_name = re.sub(r'_\d+$', '', category_name)
                
                # Tạo thư mục cho danh mục này
                category_dir = os.path.join(result_dir, category_name)
                os.makedirs(category_dir, exist_ok=True)
                
                # Thu thập liên kết sản phẩm từ danh mục này
                category_products = extract_category_links([category_url])
                
                if category_products:
                    # Lưu các liên kết sản phẩm vào file txt riêng của danh mục
                    category_file = os.path.join(category_dir, f'{category_name}_links.txt')
                    with open(category_file, 'w', encoding='utf-8') as f:
                        for link in category_products:
                            f.write(link + '\n')
                            
                    # Thêm vào danh sách tất cả sản phẩm
                    all_product_links.extend(category_products)
                    
                    # Thêm thông tin danh mục vào danh sách
                    category_info.append({
                        'Tên danh mục': category_name,
                        'URL danh mục': category_url,
                        'Số sản phẩm': len(category_products)
                    })
                else:
                    print(f"Không tìm thấy sản phẩm nào trong danh mục: {category_url}")
                    
            except Exception as e:
                error_message = str(e)
                print(f"Lỗi khi xử lý danh mục {category_url}: {error_message}")
                traceback.print_exc()
        
        # Lưu tất cả liên kết vào file chung
        with open(all_products_file, 'w', encoding='utf-8') as f:
            for link in all_product_links:
                f.write(link + '\n')
        
        # Tạo file Excel báo cáo về các danh mục
        report_file = os.path.join(result_dir, 'category_report.xlsx')
        df = pd.DataFrame(category_info)
        
        if not df.empty:
            df.to_excel(report_file, index=False)
        
        # Nén thư mục kết quả thành file ZIP
        zip_filename = f'category_products_{timestamp}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)
        
        # Tạo file ZIP từ thư mục
        if utils.create_zip_from_folder(result_dir, zip_path):
            # Gửi thông báo hoàn thành
            socketio.emit('progress_update', {
                'percent': 100, 
                'message': f'Đã hoàn thành thu thập {len(all_product_links)} liên kết sản phẩm từ {len(valid_urls)} danh mục!'
            })
            
            # Tạo URL để tải xuống file zip
            download_url = url_for('main.download_file', filename=zip_filename)
            
            # Thông báo thành công
            success_message = f"Đã thu thập thành công {len(all_product_links)} liên kết sản phẩm từ {len(valid_urls)} danh mục."
            
            return render_template('index.html', download_url=download_url, success_message=success_message)
        else:
            flash('Lỗi khi tạo file ZIP!', 'error')
            return redirect(url_for('main.index'))
            
    except Exception as e:
        error_message = str(e)
        # In chi tiết lỗi
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/scrape-category-products', methods=['POST'])
def scrape_category_products():
    """
    Xử lý yêu cầu cào dữ liệu từ nhiều danh mục baa.vn
    """
    try:
        # Lấy danh sách URL danh mục từ form
        category_urls_text = request.form.get('category_urls', '')
        if not category_urls_text:
            return jsonify({'error': 'Vui lòng nhập danh sách URL danh mục'}), 400

        # Tách các URL thành danh sách
        category_urls = [url.strip() for url in category_urls_text.split('\n') if url.strip()]
        if not category_urls:
            return jsonify({'error': 'Không tìm thấy URL danh mục hợp lệ'}), 400

        # Tạo thư mục kết quả
        result_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'category_results')
        os.makedirs(result_dir, exist_ok=True)

        # Số luồng tối đa cho việc xử lý danh mục
        max_workers = min(8, len(category_urls))  # Tối đa 8 luồng cho danh mục

        # Danh sách lưu thông tin kết quả của từng danh mục
        category_info = []

        def process_category(category_url, index):
            try:
                # Tạo thư mục cho danh mục này
                category_name = extract_category_name(category_url)
                category_dir = os.path.join(result_dir, category_name)
                os.makedirs(category_dir, exist_ok=True)

                # Tạo thư mục cho ảnh sản phẩm
                category_images_dir = os.path.join(category_dir, 'images')
                os.makedirs(category_images_dir, exist_ok=True)

                # Thu thập liên kết sản phẩm từ danh mục này
                socketio.emit('progress_update', {
                    'percent': category_progress_base + 5, 
                    'message': f'Đang thu thập liên kết sản phẩm từ danh mục: {category_name}'
                })

                category_products = extract_product_urls(category_url)

                if not category_products:
                    return {
                        'Tên danh mục': category_name,
                        'URL danh mục': category_url,
                        'Số sản phẩm': 0,
                        'Số sản phẩm có thông tin': 0,
                        'Số ảnh đã tải': 0,
                        'Thành công': False,
                        'Lỗi': 'Không tìm thấy sản phẩm nào trong danh mục'
                    }

                # Lưu các liên kết sản phẩm vào file txt riêng của danh mục
                category_file = os.path.join(category_dir, f'{category_name}_links.txt')
                with open(category_file, 'w', encoding='utf-8') as f:
                    for link in category_products:
                        f.write(link + '\n')

                # Thu thập thông tin từ các sản phẩm trong danh mục với đa luồng
                socketio.emit('progress_update', {
                    'percent': category_progress_base + 10, 
                    'message': f'Đang thu thập thông tin {len(category_products)} sản phẩm từ danh mục: {category_name}'
                })

                # Các trường cần thu thập
                required_fields = ['STT', 'Mã sản phẩm', 'Tên sản phẩm', 'Giá', 'Tổng quan', 'URL']

                # Tạo file Excel template tạm thời
                excel_temp_path = os.path.join(category_dir, f'{category_name}_template.xlsx')
                
                # Tạo template Excel đơn giản
                wb = openpyxl.Workbook()
                ws = wb.active
                for col_idx, field in enumerate(required_fields, 1):
                    ws.cell(row=1, column=col_idx).value = field
                wb.save(excel_temp_path)

                # Xử lý đa luồng cho việc cào thông tin sản phẩm
                product_max_workers = min(10, len(category_products))  # Tối đa 10 luồng cho sản phẩm

                # Chia thành các batch nhỏ để xử lý và báo cáo tiến độ
                batch_size = max(1, len(category_products) // 10)  # Chia thành khoảng 10 batch
                product_batches = [category_products[i:i + batch_size] for i in range(0, len(category_products), batch_size)]

                all_results = []
                for batch_idx, batch in enumerate(product_batches):
                    # Cập nhật tiến độ cho batch này
                    batch_progress = category_progress_base + 10 + int((batch_idx / len(product_batches)) * 25)
                    socketio.emit('progress_update', {
                        'percent': batch_progress, 
                        'message': f'Đang cào batch {batch_idx+1}/{len(product_batches)} ({len(batch)} sản phẩm) từ danh mục: {category_name}'
                    })

                    # Sử dụng ThreadPoolExecutor cho từng batch
                    with concurrent.futures.ThreadPoolExecutor(max_workers=product_max_workers) as executor:
                        # Tạo tác vụ cho từng URL sản phẩm
                        future_to_url = {executor.submit(extract_product_info, url, required_fields, i + batch_idx * batch_size): url 
                                        for i, url in enumerate(batch)}

                        # Thu thập kết quả từ các tác vụ khi hoàn thành
                        for future in concurrent.futures.as_completed(future_to_url):
                            url = future_to_url[future]
                            try:
                                data = future.result()
                                if data:
                                    all_results.append(data)
                            except Exception as exc:
                                print(f'Lỗi khi cào dữ liệu từ {url}: {exc}')

                # Gộp tất cả kết quả và lưu vào Excel
                excel_result = os.path.join(category_dir, f'{category_name}_products.xlsx')

                if all_results:
                    # Tạo DataFrame từ danh sách thông tin sản phẩm
                    df_products = pd.DataFrame(all_results)

                    # Đảm bảo có đủ các cột cần thiết
                    for field in required_fields:
                        if field not in df_products.columns:
                            df_products[field] = ""

                    # Chỉ giữ lại các cột theo thứ tự
                    available_fields = [field for field in required_fields if field in df_products.columns]
                    df_products = df_products[available_fields]

                    # Sắp xếp lại theo STT
                    if 'STT' in df_products.columns:
                        df_products = df_products.sort_values('STT')

                    # Lưu vào file Excel
                    df_products.to_excel(excel_result, index=False, engine='openpyxl')

                # Tải ảnh sản phẩm từ baa.vn
                socketio.emit('progress_update', {
                    'percent': category_progress_base + 40, 
                    'message': f'Đang tải ảnh {len(category_products)} sản phẩm từ danh mục: {category_name}'
                })

                # Sử dụng đa luồng để tải ảnh
                image_results = download_baa_product_images_fixed(category_products, category_images_dir)

                # Kết quả xử lý danh mục
                return {
                    'Tên danh mục': category_name,
                    'URL danh mục': category_url,
                    'Số sản phẩm': len(category_products),
                    'Số sản phẩm có thông tin': len(all_results),
                    'Số ảnh đã tải': image_results['success'],
                    'Thành công': True,
                    'Lỗi': None
                }

            except Exception as e:
                error_message = str(e)
                print(f"Lỗi khi xử lý danh mục {category_url}: {error_message}")
                traceback.print_exc()
                return {
                    'Tên danh mục': category_name if 'category_name' in locals() else "Unknown",
                    'URL danh mục': category_url,
                    'Số sản phẩm': len(category_products) if 'category_products' in locals() else 0,
                    'Số sản phẩm có thông tin': len(all_results) if 'all_results' in locals() else 0,
                    'Số ảnh đã tải': image_results['success'] if 'image_results' in locals() else 0,
                    'Thành công': False,
                    'Lỗi': error_message
                }

        # Sử dụng ThreadPoolExecutor để xử lý các danh mục đồng thời
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Tạo các futures cho từng URL danh mục
            futures = [executor.submit(process_category, url, i) for i, url in enumerate(category_urls)]

            # Thu thập kết quả
            for future in concurrent.futures.as_completed(futures):
                try:
                    result = future.result()
                    if result:
                        category_info.append(result)
                except Exception as exc:
                    print(f'Lỗi từ future: {exc}')

        # Tạo file ZIP chứa tất cả kết quả
        zip_filename = f'category_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        zip_path = os.path.join(current_app.config['UPLOAD_FOLDER'], zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(result_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, result_dir)
                    zipf.write(file_path, arcname)

        # Tạo báo cáo tổng hợp
        report_data = []
        total_products = 0
        total_info = 0
        total_images = 0
        successful_categories = 0

        for info in category_info:
            report_data.append({
                'Tên danh mục': info['Tên danh mục'],
                'URL danh mục': info['URL danh mục'],
                'Số sản phẩm': info['Số sản phẩm'],
                'Số sản phẩm có thông tin': info['Số sản phẩm có thông tin'],
                'Số ảnh đã tải': info['Số ảnh đã tải'],
                'Trạng thái': 'Thành công' if info['Thành công'] else 'Thất bại',
                'Lỗi': info['Lỗi'] if info['Lỗi'] else ''
            })

            if info['Thành công']:
                total_products += info['Số sản phẩm']
                total_info += info['Số sản phẩm có thông tin']
                total_images += info['Số ảnh đã tải']
                successful_categories += 1

        # Thêm dòng tổng kết
        report_data.append({
            'Tên danh mục': 'TỔNG KẾT',
            'URL danh mục': '',
            'Số sản phẩm': total_products,
            'Số sản phẩm có thông tin': total_info,
            'Số ảnh đã tải': total_images,
            'Trạng thái': f'Thành công: {successful_categories}/{len(category_urls)} danh mục',
            'Lỗi': ''
        })

        # Lưu báo cáo vào file Excel
        report_df = pd.DataFrame(report_data)
        report_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'category_report.xlsx')
        report_df.to_excel(report_path, index=False, engine='openpyxl')

        # Thêm file báo cáo vào ZIP
        with zipfile.ZipFile(zip_path, 'a', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(report_path, 'category_report.xlsx')

        # Xóa thư mục kết quả tạm thời
        shutil.rmtree(result_dir)

        return jsonify({
            'success': True,
            'message': f'Đã cào dữ liệu từ {len(category_urls)} danh mục, thu thập được {total_info}/{total_products} sản phẩm và {total_images} ảnh',
            'download_url': url_for('main.download_file', filename=zip_filename)
        })

    except Exception as e:
        error_message = str(e)
        print(f"Lỗi khi xử lý yêu cầu: {error_message}")
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xử lý yêu cầu: {error_message}'}), 500

@main_bp.route('/download-category-baa-images', methods=['POST'])
def download_category_baa_images():
    """Tải ảnh sản phẩm từ nhiều danh mục BAA.vn"""
    try:
        # Lấy danh sách URL danh mục từ form
        category_urls_text = request.form.get('category_urls', '')
        if not category_urls_text.strip():
            flash('Vui lòng nhập ít nhất một URL danh mục', 'error')
            return redirect(url_for('main.index'))
        
        # Tách danh sách URL thành các dòng riêng biệt
        category_urls = [url.strip() for url in category_urls_text.splitlines() if url.strip()]
        
        # Kiểm tra tính hợp lệ của URL
        valid_urls = []
        for url in category_urls:
            if url.startswith(('http://', 'https://')):
                valid_urls.append(url)
            else:
                flash(f'URL không hợp lệ: {url}', 'warning')
        
        if not valid_urls:
            flash('Không có URL danh mục hợp lệ nào', 'error')
            return redirect(url_for('main.index'))
        
        # Tạo thư mục lưu ảnh với timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], f'baa_category_images_{timestamp}')
        os.makedirs(output_dir, exist_ok=True)
        
        # Tạo file báo cáo Excel để theo dõi tiến trình
        report_file = os.path.join(output_dir, 'tong_hop.xlsx')
        
        # Khởi tạo biến lưu trữ kết quả
        all_results = {
            'total_categories': len(valid_urls),
            'total_products': 0,
            'total_success': 0,
            'total_failed': 0,
            'categories': []
        }
        
        # Import các module
        from app.crawler import extract_product_urls, download_baa_product_images_fixed
        from urllib.parse import urlparse
        import pandas as pd
        import openpyxl
        import traceback
        
        # Xử lý từng danh mục
        for index, category_url in enumerate(valid_urls):
            try:
                # Cập nhật tiến trình
                progress = int((index / len(valid_urls)) * 90)
                socketio.emit('progress_update', {
                    'percent': progress,
                    'message': f'Đang xử lý danh mục {index+1}/{len(valid_urls)}: {category_url}'
                })
                
                # Trích xuất tên danh mục từ URL
                category_name = extract_category_name(category_url)
                category_dir = os.path.join(output_dir, category_name)
                os.makedirs(category_dir, exist_ok=True)
                
                # Trích xuất URL sản phẩm từ danh mục
                product_urls = extract_product_urls(category_url)
                
                # Lưu danh sách URL sản phẩm
                urls_file = os.path.join(category_dir, 'product_urls.txt')
                with open(urls_file, 'w', encoding='utf-8') as f:
                    for url in product_urls:
                        f.write(f"{url}\n")
                
                # Tải ảnh sản phẩm
                results = download_baa_product_images_fixed(product_urls, category_dir)
                
                # Cập nhật kết quả
                category_result = {
                    'name': category_name,
                    'url': category_url,
                    'total_products': len(product_urls),
                    'success': results['success'],
                    'failed': results['failed'],
                    'image_paths': results['image_paths']
                }
                
                all_results['categories'].append(category_result)
                all_results['total_products'] += len(product_urls)
                all_results['total_success'] += results['success']
                all_results['total_failed'] += results['failed']
                
                # Lưu báo cáo cho danh mục
                category_report_file = os.path.join(category_dir, 'bao_cao.xlsx')
                create_image_report(results['report_data'], category_report_file)
                
                socketio.emit('progress_update', {
                    'percent': progress + 3,
                    'message': f'Đã tải {results["success"]}/{len(product_urls)} ảnh từ danh mục {category_name}'
                })
                
            except Exception as e:
                print(f"Lỗi khi xử lý danh mục {category_url}: {str(e)}")
                traceback.print_exc()
                all_results['categories'].append({
                    'name': extract_category_name(category_url),
                    'url': category_url,
                    'error': str(e),
                    'total_products': 0,
                    'success': 0,
                    'failed': 0,
                    'image_paths': []
                })
        
        # Tạo báo cáo tổng hợp
        create_category_images_report(all_results, report_file)
        
        # Nén kết quả
        zip_file = output_dir + '.zip'
        utils.create_zip_from_folder(output_dir, zip_file)
        
        # Xóa thư mục tạm sau khi nén
        # shutil.rmtree(output_dir)
        
        socketio.emit('progress_update', {
            'percent': 100,
            'message': 'Hoàn tất tải ảnh từ các danh mục. Đang chuẩn bị tải xuống...'
        })
        
        # Chuyển hướng để tải xuống
        return redirect(url_for('main.download_file', filename=os.path.basename(zip_file)))
        
    except Exception as e:
        error_message = str(e)
        traceback.print_exc()
        flash(f'Lỗi khi tải ảnh từ danh mục: {error_message}', 'error')
        return redirect(url_for('main.index'))

def extract_category_name(url):
    """Trích xuất tên danh mục từ URL"""
    try:
        # Loại bỏ protocol và tên miền
        path = urlparse(url).path
        
        # Loại bỏ các phần không cần thiết
        parts = [p for p in path.split('/') if p and p not in ['vn', 'Category', 'tag']]
        
        if parts:
            # Lấy phần cuối cùng của đường dẫn
            last_part = parts[-1]
            
            # Xử lý trường hợp có tham số ở cuối URL
            if '?' in last_part:
                last_part = last_part.split('?')[0]
            
            # Loại bỏ ID ở cuối URL nếu có
            if '_' in last_part and last_part.split('_')[-1].isdigit():
                last_part = '_'.join(last_part.split('_')[:-1])
            
            # Chuyển đổi dấu gạch ngang thành dấu gạch dưới
            cleaned_name = last_part.replace('-', '_')
            
            return cleaned_name
        
        # Nếu không trích xuất được, sử dụng timestamp
        return datetime.now().strftime('category_%Y%m%d_%H%M%S')
        
    except Exception:
        # Nếu có lỗi, sử dụng timestamp
        return datetime.now().strftime('category_%Y%m%d_%H%M%S')

def create_image_report(report_data, output_file):
    """Tạo báo cáo Excel về việc tải ảnh"""
    try:
        # Tạo DataFrame từ dữ liệu báo cáo
        df = pd.DataFrame(report_data)
        
        # Lưu vào file Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Báo cáo tải ảnh', index=False)
            
            # Tự động điều chỉnh chiều rộng các cột
            worksheet = writer.sheets['Báo cáo tải ảnh']
            for i, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[worksheet.cell(row=1, column=i+1).column_letter].width = max_length
                
        return True
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo: {str(e)}")
        return False

def create_category_images_report(results, output_file):
    """Tạo báo cáo tổng hợp về việc tải ảnh từ các danh mục"""
    try:
        # Tạo workbook
        wb = openpyxl.Workbook()
        
        # Tạo sheet tổng quan
        ws_overview = wb.active
        ws_overview.title = "Tổng quan"
        
        # Thêm thông tin tổng quan
        ws_overview.append(["Báo cáo tải ảnh từ nhiều danh mục"])
        ws_overview.append(["Thời gian tạo:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_overview.append([])
        ws_overview.append(["Tổng số danh mục:", results['total_categories']])
        ws_overview.append(["Tổng số sản phẩm:", results['total_products']])
        ws_overview.append(["Số ảnh tải thành công:", results['total_success']])
        ws_overview.append(["Số ảnh tải thất bại:", results['total_failed']])
        
        # Tạo sheet chi tiết danh mục
        ws_details = wb.create_sheet("Chi tiết danh mục")
        ws_details.append(["STT", "Tên danh mục", "URL danh mục", "Số sản phẩm", "Thành công", "Thất bại", "Tỷ lệ thành công"])
        
        # Thêm thông tin chi tiết từng danh mục
        for i, category in enumerate(results['categories']):
            success_rate = 0
            if category.get('total_products', 0) > 0:
                success_rate = (category.get('success', 0) / category.get('total_products', 0)) * 100
                
            ws_details.append([
                i + 1,
                category.get('name', 'N/A'),
                category.get('url', 'N/A'),
                category.get('total_products', 0),
                category.get('success', 0),
                category.get('failed', 0),
                f"{success_rate:.2f}%"
            ])
        
        # Điều chỉnh độ rộng cột
        for worksheet in [ws_overview, ws_details]:
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column].width = max_length + 2
        
        # Lưu workbook
        wb.save(output_file)
        return True
        
    except Exception as e:
        print(f"Lỗi khi tạo báo cáo tổng hợp: {str(e)}")
        return False

@main_bp.route('/crawl-codienhaiau', methods=['POST'])
def crawl_codienhaiau():
    try:
        category_urls = request.form.get('category_urls', '').strip()
        
        if not category_urls:
            flash('Vui lòng nhập ít nhất một URL danh mục.', 'error')
            return redirect(url_for('main.index'))
        
        # Tạo crawler và xử lý danh mục
        crawler = CategoryCrawler(socketio, upload_folder=current_app.config['UPLOAD_FOLDER'])
        success, message, zip_path = crawler.process_codienhaiau_categories(category_urls)
        
        if success:
            filename = os.path.basename(zip_path) if zip_path else session.get('last_download', '')
            if filename:
                # Lưu thông báo thành công và URL tải xuống
                flash(message, 'success')
                return redirect(url_for('main.download_file', filename=filename))
            else:
                flash('Đã xử lý thành công nhưng không tìm thấy file để tải xuống.', 'warning')
                return redirect(url_for('main.index'))
        else:
            flash(f'Lỗi: {message}', 'error')
            return redirect(url_for('main.index'))
        
    except Exception as e:
        traceback.print_exc()
        flash(f'Lỗi không xác định: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/crawl-baa', methods=['POST'])
@progress_tracker(name="Cào dữ liệu BAA.vn", total_steps=100, verbose=True)
def crawl_baa(progress: TerminalProgressBar):
    try:
        progress.update(5, "Đang kiểm tra thông tin đầu vào")
        
        # Lấy thông tin từ form
        category_urls = request.form.get('category_urls', '').strip()
        if not category_urls:
            progress.error('Vui lòng nhập ít nhất một URL danh mục hoặc sản phẩm.')
            flash('Vui lòng nhập ít nhất một URL danh mục hoặc sản phẩm.', 'error')
            return redirect(url_for('main.index'))
        
        # Lấy tham số tùy chọn
        max_workers = int(request.form.get('max_workers', 8))
        max_retries = int(request.form.get('max_retries', 3))
        
        # Giới hạn giá trị hợp lệ
        max_workers = min(max(1, max_workers), 16)  # Từ 1-16 luồng
        max_retries = min(max(1, max_retries), 5)   # Từ 1-5 lần thử lại
        
        progress.update(10, "Đang chuẩn bị crawler", f"Max workers: {max_workers}, Max retries: {max_retries}")
        
        # Log thông tin
        print(f"Bắt đầu cào dữ liệu với {max_workers} luồng, {max_retries} lần thử lại")
        
        url_list = [u.strip() for u in category_urls.splitlines() if u.strip()]
        
        progress.update(15, f"Đã phân tích {len(url_list)} URLs", "Đang khởi tạo BAA crawler")
        
        # Tạo progress bar con cho crawler
        crawler_progress = create_child_progress(progress, "BAA Product Crawler", 70)
        
        crawler = BaaProductCrawler(output_root=current_app.config['UPLOAD_FOLDER'], 
                                   max_workers=max_workers, 
                                   max_retries=max_retries)
        
        crawler_progress.update(5, "Crawler đã sẵn sàng", "Bắt đầu cào dữ liệu...")
        
        # Inject progress vào crawler để theo dõi tiến trình
        products, result_dir = crawler.crawl_products(url_list)
        
        crawler_progress.complete("success", f"Cào dữ liệu hoàn tất", {
            "Sản phẩm đã cào": len(products),
            "Thư mục kết quả": result_dir
        })
        
        progress.update(85, f"Đã cào {len(products)} sản phẩm", "Đang nén kết quả thành ZIP")
        
        # Nén kết quả và trả về link download
        zip_progress = create_child_progress(progress, "Tạo file ZIP", 10)
        
        zip_path = result_dir + '.zip'
        if os.path.exists(zip_path):
            zip_progress.complete("success", "File ZIP đã tồn tại")
            
            progress.update(95, "Đang tạo thông tin tải xuống")
            
            download_url = url_for('main.download_file', filename=os.path.basename(zip_path))
            flash(f'Đã cào xong dữ liệu {len(products)} sản phẩm. <a href="{download_url}" class="btn btn-primary mt-2">Tải xuống kết quả</a>', 'success')
            
            # Tạo dữ liệu để hiển thị trên giao diện
            zip_size = os.path.getsize(zip_path) / (1024 * 1024)  # Kích thước MB
            products_count = len(products)
            stats = {
                'Tổng sản phẩm': products_count,
                'Thời gian xử lý': f"{os.path.basename(result_dir).split('_')[-1]} giây",
                'Kích thước file': f"{zip_size:.2f} MB"
            }
            
            # Hoàn thành progress với thống kê chi tiết
            progress.complete("success", "Cào dữ liệu BAA.vn hoàn tất", {
                "URLs đầu vào": len(url_list),
                "Sản phẩm cào được": products_count,
                "Kích thước ZIP": f"{zip_size:.2f} MB",
                "Max workers": max_workers,
                "Max retries": max_retries,
                "File ZIP": os.path.basename(zip_path)
            })
            
            # Trả về trang kết quả với đường dẫn tải xuống và thống kê
            return render_template('crawler_result.html', 
                                  download_url=download_url,
                                  zip_filename=os.path.basename(zip_path),
                                  stats=stats,
                                  products_count=products_count)
        else:
            zip_progress.error("Không tạo được file ZIP")
            progress.warning(f'Đã cào xong dữ liệu {len(products)} sản phẩm nhưng không tạo được file ZIP.')
            flash(f'Đã cào xong dữ liệu {len(products)} sản phẩm nhưng không tạo được file ZIP.', 'warning')
            return redirect(url_for('main.index'))
    
    except Exception as e:
        traceback.print_exc()
        progress.error(f'Lỗi: {str(e)}', traceback.format_exc())
        flash(f'Lỗi: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/scrap-category-products', methods=['POST'])
def scrap_category_products():
    try:
        category_urls = request.form.get('category_urls', '').strip()
        
        if not category_urls:
            return jsonify({'status': 'error', 'message': 'Vui lòng nhập ít nhất một URL danh mục.'})
        
        # Tạo crawler và xử lý danh mục
        crawler = CategoryCrawler(socketio, upload_folder=current_app.config['UPLOAD_FOLDER'])
        success, message = crawler.process_category_urls(category_urls)
        
        if success:
            filename = session.get('last_download', '')
            if filename:
                return jsonify({
                    'status': 'success',
                    'message': message,
                    'download': filename
                })
        
        return jsonify({'status': 'error', 'message': message})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'Lỗi không xác định: {str(e)}'})

@main_bp.route('/filter-product-types', methods=['POST'])
def filter_product_types():
    """
    Lọc sản phẩm theo các loại được chọn (tiệm cận, sợi quang, quang điện, áp suất, v.v.)
    """
    try:
        # Kiểm tra nếu có file Excel
        if 'excel_file' not in request.files:
            flash('Vui lòng tải lên file Excel!', 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        # Lấy danh sách loại sản phẩm được chọn
        selected_types = request.form.getlist('product_types')
        print(f"DEBUG: Các loại sản phẩm được chọn: {selected_types}")
        
        if not selected_types:
            flash('Vui lòng chọn ít nhất một loại sản phẩm!', 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        # Lấy file Excel
        excel_file = request.files['excel_file']
        if excel_file.filename == '':
            flash('Không có file Excel nào được chọn!', 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        if not allowed_file(excel_file.filename, ALLOWED_EXTENSIONS_EXCEL):
            flash('Chỉ chấp nhận file .xlsx hoặc .xls!', 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        # Thông báo tiến trình
        socketio.emit('progress_update', {
            'percent': 10, 
            'message': f'Đang xử lý file Excel và tìm kiếm {len(selected_types)} loại sản phẩm...'
        })
        
        # Các từ khóa cho mỗi loại sản phẩm
        product_type_keywords = {
            'tiem_can': ['tiệm cận', 'tiếp cận', 'proximity', 'PR', 'PS', 'PRL', 'PRCM', 'PR12', 'PR18', 'PR30'],
            'soi_quang': ['sợi quang', 'fiber optic', 'BF', 'BFL', 'BFX', 'BFF', 'fiber sensor'],
            'quang_dien': ['quang điện', 'photoelectric', 'BEN', 'BJ', 'BYD', 'BR', 'BRP', 'BH', 'BL', 'BX', 'BY', 'photo sensor'],
            'vung_kv': ['vùng', 'khu vực', 'area', 'BA', 'BAR', 'area sensor'],
            'ap_suat': ['áp suất', 'pressure', 'PSA', 'PSB', 'PS', 'pressure sensor'],
            'luu_luong': ['lưu lượng', 'flow', 'FS', 'FSA', 'flow sensor'],
            'xy_lanh': ['xy lanh', 'cylinder', 'CY', 'CP', 'cylinder sensor'],
            'dien_dung': ['điện dung', 'capacitive', 'CR', 'CRL', 'capacitive sensor']
        }
        
        # Lưu file tạm thời
        temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        input_excel_path = os.path.join(temp_dir, secure_filename(excel_file.filename))
        excel_file.save(input_excel_path)
        
        print(f"DEBUG: Đã lưu file Excel tạm thời tại: {input_excel_path}")
        
        # Đọc file Excel với pandas
        df = None
        try:
            # Thử đọc file với nhiều sheet khác nhau
            xls = pd.ExcelFile(input_excel_path)
            sheet_names = xls.sheet_names
            print(f"DEBUG: Các sheet trong file Excel: {sheet_names}")
            
            if len(sheet_names) > 0:
                # Nếu file có sheet "Tổng hợp sản phẩm", sử dụng sheet đó
                if "Tổng hợp sản phẩm" in sheet_names:
                    df = pd.read_excel(input_excel_path, sheet_name="Tổng hợp sản phẩm")
                    print(f"DEBUG: Đọc dữ liệu từ sheet 'Tổng hợp sản phẩm'")
                else:
                    # Sử dụng sheet đầu tiên
                    df = pd.read_excel(input_excel_path, sheet_name=sheet_names[0])
                    print(f"DEBUG: Đọc dữ liệu từ sheet '{sheet_names[0]}'")
            else:
                df = pd.read_excel(input_excel_path)
            
            # In thông tin cấu trúc DataFrame
            print(f"DEBUG: Thông tin DataFrame: {df.shape}, Columns: {df.columns.tolist()}")
            
            # Chuyển tất cả các cột có kiểu object sang string để tránh lỗi NaN
            for col in df.select_dtypes(include=['object']).columns:
                df[col] = df[col].astype(str)
            
            socketio.emit('progress_update', {
                'percent': 30, 
                'message': f'Đã đọc file Excel với {len(df)} dòng. Đang lọc dữ liệu...'
            })
        except Exception as e:
            error_message = f'Lỗi khi đọc file Excel: {str(e)}'
            print(f"DEBUG: {error_message}")
            traceback.print_exc()
            flash(error_message, 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        # Kiểm tra nếu DataFrame trống hoặc không có dữ liệu
        if df is None or df.empty:
            flash('File Excel không chứa dữ liệu!', 'error')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        # Xác định các cột chứa thông tin sản phẩm
        product_name_col = None
        product_code_col = None
        product_desc_col = None
        
        # Tìm các cột có thể chứa tên sản phẩm, mã sản phẩm hoặc mô tả
        for i, col in enumerate(df.columns):
            col_lower = str(col).lower()
            if 'tên' in col_lower or 'sản phẩm' in col_lower or 'name' in col_lower:
                product_name_col = i
                print(f"DEBUG: Tìm thấy cột tên sản phẩm: {col} (index: {i})")
            elif 'mã' in col_lower or 'code' in col_lower:
                product_code_col = i
                print(f"DEBUG: Tìm thấy cột mã sản phẩm: {col} (index: {i})")
            elif 'mô tả' in col_lower or 'description' in col_lower or 'desc' in col_lower:
                product_desc_col = i
                print(f"DEBUG: Tìm thấy cột mô tả: {col} (index: {i})")
        
        # Nếu không tìm thấy cột, sử dụng cột mặc định
        if product_name_col is None:
            product_name_col = 2  # Giả sử cột C chứa tên sản phẩm
            print(f"DEBUG: Không tìm thấy cột tên sản phẩm, sử dụng cột mặc định: {product_name_col}")
        if product_code_col is None:
            product_code_col = 1  # Giả sử cột B chứa mã sản phẩm
            print(f"DEBUG: Không tìm thấy cột mã sản phẩm, sử dụng cột mặc định: {product_code_col}")
        if product_desc_col is None:
            # Cố gắng tìm cột có vẻ như chứa mô tả (thường có nhiều chữ)
            for i, col in enumerate(df.columns):
                if i != product_name_col and i != product_code_col:
                    # Kiểm tra một số dòng đầu tiên để xem cột nào có nhiều chữ nhất
                    sample = df.iloc[:min(10, len(df)), i].astype(str)
                    if sample.str.len().mean() > 20:  # Nếu trung bình độ dài > 20 ký tự
                        product_desc_col = i
                        print(f"DEBUG: Đoán cột mô tả: {col} (index: {i})")
                        break
        
        print(f"DEBUG: Sử dụng cột tên: {product_name_col}, cột mã: {product_code_col}, cột mô tả: {product_desc_col}")
        
        # Tạo các DataFrame cho từng loại sản phẩm
        filtered_dfs = {}
        for product_type in selected_types:
            filtered_dfs[product_type] = pd.DataFrame(columns=df.columns)
        
        # Lọc sản phẩm dựa trên từ khóa
        for index, row in df.iterrows():
            try:
                # Lấy giá trị của các cột, bảo vệ từ lỗi index ngoài phạm vi và xử lý NaN
                product_name = str(row.iloc[product_name_col] if product_name_col < len(row) else "").lower()
                product_code = str(row.iloc[product_code_col] if product_code_col < len(row) else "").lower()
                product_desc = ""
                if product_desc_col is not None and product_desc_col < len(row):
                    product_desc = str(row.iloc[product_desc_col]).lower()
                
                combined_text = f"{product_name} {product_code} {product_desc}"
                
                # Debug thông tin của một số dòng mẫu
                if index < 5 or index % 1000 == 0:
                    print(f"DEBUG: Dòng {index} - Mã: {product_code}, Tên: {product_name[:50]}{'...' if len(product_name) > 50 else ''}")
                
                for product_type in selected_types:
                    if product_type in product_type_keywords:
                        for keyword in product_type_keywords[product_type]:
                            if keyword.lower() in combined_text:
                                # Sử dụng concat thay vì append (đã deprecated trong pandas mới)
                                row_df = pd.DataFrame([row.values], columns=df.columns)
                                filtered_dfs[product_type] = pd.concat([filtered_dfs[product_type], row_df], ignore_index=True)
                                # In thông tin debug nếu tìm thấy kết quả
                                if index < 5 or index % 1000 == 0:
                                    print(f"DEBUG: Dòng {index} khớp với loại '{product_type}' (từ khóa: '{keyword}')")
                                break
            except Exception as e:
                print(f"DEBUG: Lỗi khi xử lý dòng {index}: {str(e)}")
                continue
        
        # Kiểm tra xem có sản phẩm nào được lọc không
        total_filtered = sum(len(df_type) for df_type in filtered_dfs.values())
        print(f"DEBUG: Tổng số sản phẩm đã lọc: {total_filtered}")
        
        if total_filtered == 0:
            flash('Không tìm thấy sản phẩm nào thuộc các loại đã chọn!', 'warning')
            return redirect(url_for('main.index', _anchor='filter-product-types-tab'))
        
        socketio.emit('progress_update', {
            'percent': 60, 
            'message': 'Đã lọc dữ liệu. Đang tạo báo cáo...'
        })
        
        # Tạo tên file output
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_filename = f"filtered_product_types_{timestamp}.xlsx"
        output_path = os.path.join(current_app.config['UPLOAD_FOLDER'], output_filename)
        
        # Tổng hợp số lượng sản phẩm mỗi loại
        summary_data = {
            'Loại sản phẩm': [],
            'Số lượng': []
        }
        
        print(f"DEBUG: Đang tạo file Excel: {output_path}")
        
        # Lưu kết quả vào file Excel với sheet cho mỗi loại sản phẩm
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            # Lưu một sheet tổng hợp với tất cả sản phẩm đã lọc
            all_filtered = pd.DataFrame()
            
            for product_type in selected_types:
                if product_type in filtered_dfs and not filtered_dfs[product_type].empty:
                    # Xóa các hàng trùng lặp
                    filtered_dfs[product_type] = filtered_dfs[product_type].drop_duplicates()
                    
                    # Tên hiển thị cho mỗi loại
                    type_display_name = {
                        'tiem_can': 'Tiệm cận',
                        'soi_quang': 'Sợi quang',
                        'quang_dien': 'Quang điện',
                        'vung_kv': 'Vùng (khu vực)',
                        'ap_suat': 'Áp suất',
                        'luu_luong': 'Lưu lượng',
                        'xy_lanh': 'Xy lanh',
                        'dien_dung': 'Điện dung'
                    }
                    
                    display_name = type_display_name.get(product_type, product_type)
                    sheet_name = display_name[:31]  # Giới hạn độ dài sheet name
                    
                    # Thêm vào dữ liệu tổng hợp
                    all_filtered = pd.concat([all_filtered, filtered_dfs[product_type]], ignore_index=True)
                    
                    # Thêm vào dữ liệu tổng kết
                    summary_data['Loại sản phẩm'].append(display_name)
                    summary_data['Số lượng'].append(len(filtered_dfs[product_type]))
                    
                    # Lưu vào sheet riêng
                    filtered_dfs[product_type].to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"DEBUG: Đã tạo sheet '{sheet_name}' với {len(filtered_dfs[product_type])} dòng")
            
            # Xóa các hàng trùng lặp trong sheet tổng hợp
            if not all_filtered.empty:
                all_filtered = all_filtered.drop_duplicates()
                all_filtered.to_excel(writer, sheet_name='Tất cả', index=False)
                print(f"DEBUG: Đã tạo sheet 'Tất cả' với {len(all_filtered)} dòng")
            
            # Tạo sheet tổng kết
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Tổng kết', index=False)
            print(f"DEBUG: Đã tạo sheet 'Tổng kết'")
            
            # Format các sheet
            workbook = writer.book
            for sheet in writer.sheets.values():
                sheet.set_column('A:Z', 18)
        
        socketio.emit('progress_update', {
            'percent': 100, 
            'message': 'Hoàn thành lọc dữ liệu!'
        })
        
        # Tạo URL để tải xuống file kết quả
        download_url = url_for('main.download_file', filename=output_filename)
        
        # Thông báo kết quả
        success_message = f"Đã lọc thành công {total_filtered} sản phẩm thuộc {len(selected_types)} loại!"
        
        # Lưu trữ session để chuyển về đúng tab sau khi tải
        session['active_tab'] = 'filter-product-types-tab'
        
        return render_template('index.html', 
                               title='CrawlBot - Công cụ thu thập dữ liệu sản phẩm',
                               download_url=download_url, 
                               success_message=success_message,
                               active_tab='filter-product-types-tab')
        
    except Exception as e:
        error_message = str(e)
        print(f"DEBUG: Lỗi khi lọc sản phẩm theo loại: {error_message}")
        traceback.print_exc()
        flash(f'Lỗi: {error_message}', 'error')
        return redirect(url_for('main.index', _anchor='filter-product-types-tab'))

@main_bp.route('/categorize-products', methods=['POST'])
def categorize_products():
    """
    Phân loại sản phẩm theo danh mục từ file Excel/CSV
    """
    try:
        if 'product_file' not in request.files:
            return render_template('index.html', error="Không tìm thấy file dữ liệu sản phẩm", active_tab="categorize-products-tab")
        
        product_file = request.files['product_file']
        
        if product_file.filename == '':
            return render_template('index.html', error="Không có file nào được chọn", active_tab="categorize-products-tab")
        
        # Kiểm tra định dạng file
        file_ext = os.path.splitext(product_file.filename)[1].lower()
        if file_ext not in ['.xlsx', '.xls', '.csv']:
            return render_template('index.html', error="File phải có định dạng .xlsx, .xls hoặc .csv", active_tab="categorize-products-tab")
        
        # Lưu file tạm
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_ext)
        temp_file_path = temp_file.name
        temp_file.close()
        product_file.save(temp_file_path)
        
        # Lấy thư mục đầu ra từ form
        output_dir = request.form.get('output_dir', '').strip()
        if not output_dir:
            output_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'categorized')
            
        # Đảm bảo thư mục đầu ra tồn tại
        os.makedirs(output_dir, exist_ok=True)
        
        # Khởi tạo và chạy trình phân loại sản phẩm
        socketio.emit('progress_update', {'percent': 10, 'message': 'Đang đọc dữ liệu sản phẩm...'})
        categorizer = ProductCategorizer()
        
        socketio.emit('progress_update', {'percent': 30, 'message': 'Đang phân loại sản phẩm theo danh mục...'})
        result = categorizer.categorize_and_export(temp_file_path, output_dir)
        
        # Xóa file tạm
        os.unlink(temp_file_path)
        
        # Kiểm tra kết quả
        if isinstance(result, dict) and 'error' in result:
            return render_template('index.html', error=f"Lỗi khi phân loại sản phẩm: {result['error']}", active_tab="categorize-products-tab")
        
        # Tạo đường dẫn download
        output_filename = os.path.basename(result)
        download_url = url_for('main.download_file', filename=f"categorized/{output_filename}")
        
        socketio.emit('progress_update', {'percent': 100, 'message': 'Hoàn thành phân loại sản phẩm!'})
        
        success_message = f"Đã phân loại sản phẩm thành công theo danh mục và tạo file Excel {output_filename}"
        return render_template('index.html', success_message=success_message, download_url=download_url, active_tab="categorize-products-tab")
        
    except Exception as e:
        error_msg = f"Lỗi khi phân loại sản phẩm: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        return render_template('index.html', error=error_msg, active_tab="categorize-products-tab")

@main_bp.route('/upscale-image', methods=['POST'])
def upscale_image():
    """
    Nâng cao chất lượng ảnh sử dụng Real-ESRGAN
    """
    try:
        if 'input_image' not in request.files:
            flash('Không tìm thấy file ảnh!', 'error')
            return redirect(url_for('main.index', active_tab='image-upscale-tab'))
            
        file = request.files['input_image']
        
        if file.filename == '':
            flash('Không có file nào được chọn!', 'error')
            return redirect(url_for('main.index', active_tab='image-upscale-tab'))
            
        if not file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
            flash('Chỉ chấp nhận file ảnh định dạng JPG, JPEG, PNG hoặc WebP!', 'error')
            return redirect(url_for('main.index', active_tab='image-upscale-tab'))
        
        # Sử dụng kích thước cố định 600x600
        target_size = (600, 600)
        
        # Lưu file tạm
        temp_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(input_path)
        
        # Gửi thông báo tiến trình
        socketio.emit('progress_update', {
            'percent': 10, 
            'message': 'Đang chuẩn bị xử lý ảnh...',
            'detail': f'Tệp: {file.filename}, Kích thước đích: {target_size[0]}x{target_size[1]} pixel'
        })
        
        # Thêm module resize
        from app.resize import ImageResizer
        
        # Khởi tạo ImageResizer
        resizer = ImageResizer()
        
        # Kiểm tra xem Real-ESRGAN có khả dụng không
        if resizer.realesrgan_available:
            # Gửi thông báo tiến trình
            socketio.emit('progress_update', {
                'percent': 20, 
                'message': 'Đang nâng cao chất lượng ảnh bằng AI (Real-ESRGAN)...',
                'detail': 'Quá trình này có thể mất vài phút tùy thuộc vào kích thước ảnh'
            })
        else:
            # Gửi thông báo tiến trình
            socketio.emit('progress_update', {
                'percent': 20, 
                'message': 'Đang nâng cao chất lượng ảnh...',
                'detail': 'Sử dụng phương pháp thay thế vì Real-ESRGAN không khả dụng'
            })
        
        # Tạo tên file đầu ra
        output_filename = os.path.splitext(secure_filename(file.filename))[0] + f"_enhanced_600x600.webp"
        output_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'upscaled_images')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        
        socketio.emit('progress_update', {
            'percent': 30,
            'message': 'Đang xử lý ảnh...',
            'detail': 'Đang thực hiện nâng cao chất lượng ảnh với kích thước cố định 600x600'
        })
        
        # Thực hiện upscale ảnh với kích thước cố định
        try:
            # Thử với chất lượng cao nhất
            result_path = resizer.upscale_image(input_path, output_path, target_size=target_size)
            
            socketio.emit('progress_update', {
                'percent': 90,
                'message': 'Đã xử lý ảnh thành công!',
                'detail': 'Đang kiểm tra chất lượng ảnh đầu ra'
            })
            
            # Kiểm tra kích thước file đầu ra để đảm bảo nó được tạo đúng cách
            if not os.path.exists(result_path) or os.path.getsize(result_path) < 1024:
                # Nếu file đầu ra quá nhỏ hoặc không tồn tại, thử lại với phương pháp thay thế
                socketio.emit('progress_update', {
                    'percent': 60,
                    'message': 'Đang thử lại với phương pháp thay thế...',
                    'detail': 'Phương pháp AI gặp vấn đề, đang sử dụng phương pháp truyền thống'
                })
                # Đảm bảo Real-ESRGAN không được sử dụng cho lần thử lại
                resizer.realesrgan_available = False
                result_path = resizer.upscale_image(input_path, output_path, target_size=target_size)
        except Exception as upscale_error:
            # Log lỗi
            current_app.logger.error(f"Lỗi khi upscale ảnh: {str(upscale_error)}")
            traceback.print_exc()
            
            # Thử lại với phương pháp thay thế nếu Real-ESRGAN thất bại
            socketio.emit('progress_update', {
                'percent': 40,
                'message': 'Gặp lỗi khi xử lý bằng AI, đang thử phương pháp thay thế...',
                'detail': f'Lỗi: {str(upscale_error)}'
            })
            
            # Đảm bảo Real-ESRGAN không được sử dụng cho lần thử lại
            resizer.realesrgan_available = False
            result_path = resizer.upscale_image(input_path, output_path, target_size=target_size)
        
        # Xóa thư mục tạm
        shutil.rmtree(temp_dir)
        
        # Tạo URL cho file kết quả
        result_url = url_for('main.download_upscaled_image', filename=output_filename)
        
        # Gửi thông báo hoàn thành
        socketio.emit('progress_update', {
            'percent': 100,
            'message': 'Đã hoàn thành nâng cao chất lượng ảnh!',
            'detail': f'Tệp kết quả: {output_filename} (kích thước 600x600)'
        })
        
        # Kiểm tra xem đã dùng Real-ESRGAN hay chưa
        if resizer.realesrgan_available:
            success_message = f'Đã nâng cao chất lượng ảnh thành công với AI Real-ESRGAN (kích thước 600x600 pixel)!'
        else:
            success_message = f'Đã nâng cao chất lượng ảnh thành công với kích thước 600x600 pixel!'
        
        # Trả về trang kết quả
        return render_template('index.html', 
                               active_tab='image-upscale-tab', 
                               success_message=success_message,
                               download_url=result_url)
                               
    except Exception as e:
        error_message = str(e)
        traceback.print_exc()
        flash(f'Lỗi khi xử lý ảnh: {error_message}', 'error')
        return redirect(url_for('main.index', active_tab='image-upscale-tab'))

@main_bp.route('/download-upscaled-image/<filename>')
def download_upscaled_image(filename):
    """
    Tải xuống ảnh đã được upscale
    """
    upscaled_images_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'upscaled_images')
    return send_from_directory(upscaled_images_dir, filename, as_attachment=True)

@main_bp.route('/debug-extract-products', methods=['POST'])
def debug_extract_products():
    """Route để debug việc trích xuất sản phẩm từ URL"""
    try:
        # Lấy URL từ request
        data = request.get_json()
        if not data or 'url' not in data:
            return jsonify({'error': 'Thiếu URL trong request'}), 400
        
        url = data['url'].strip()
        
        if not url:
            return jsonify({'error': 'URL không hợp lệ'}), 400
        
        # Import hàm debug
        from app.crawler import debug_extract_products_from_url
        
        print(f"Debug extract products từ URL: {url}")
        
        # Gọi hàm debug
        product_links = debug_extract_products_from_url(url)
        
        return jsonify({
            'success': True,
            'url': url,
            'product_count': len(product_links),
            'product_links': product_links[:20],  # Giới hạn 20 URL để tránh response quá lớn
            'message': f'Tìm thấy {len(product_links)} URL sản phẩm'
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Lỗi debug extract products: {str(e)}")
        print(error_details)
        
        return jsonify({
            'error': f'Lỗi khi debug: {str(e)}',
            'details': error_details
        }), 500

@main_bp.route('/download-baa-result/<filename>')
def download_baa_result(filename):
    """Tải xuống file kết quả từ BAA crawler với thông tin chi tiết"""
    try:
        # Lấy đường dẫn file từ thư mục output_baa
        baa_output_dir = os.path.join(os.getcwd(), "output_baa")
        file_path = os.path.join(baa_output_dir, filename)
        
        # Kiểm tra file tồn tại
        if not os.path.exists(file_path):
            flash('File không tồn tại hoặc đã bị xóa', 'error')
            return redirect(url_for('main.index'))
        
        return send_file(file_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Lỗi khi tải file: {str(e)}")
        flash(f'Lỗi khi tải file: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/crawl-baa-old', methods=['POST'])
def crawl_baa_old():
    try:
        # Lấy thông tin từ form
        category_urls = request.form.get('category_urls', '').strip()
        if not category_urls:
            flash('Vui lòng nhập ít nhất một URL danh mục hoặc sản phẩm.', 'error')
            return redirect(url_for('main.index'))
        
        # Lấy tham số tùy chọn
        max_workers = int(request.form.get('max_workers', 8))
        max_retries = int(request.form.get('max_retries', 3))
        
        # Giới hạn giá trị hợp lệ
        max_workers = min(max(1, max_workers), 16)  # Từ 1-16 luồng
        max_retries = min(max(1, max_retries), 5)   # Từ 1-5 lần thử lại
        
        # Log thông tin
        print(f"Bắt đầu cào dữ liệu với {max_workers} luồng, {max_retries} lần thử lại")
        
        url_list = [u.strip() for u in category_urls.splitlines() if u.strip()]
        crawler = BaaProductCrawler(output_root=current_app.config['UPLOAD_FOLDER'], 
                                   max_workers=max_workers, 
                                   max_retries=max_retries)
        products, result_dir = crawler.crawl_products(url_list)
        
        # Nén kết quả và trả về link download
        zip_path = result_dir + '.zip'
        if os.path.exists(zip_path):
            download_url = url_for('main.download_file', filename=os.path.basename(zip_path))
            flash(f'Đã cào xong dữ liệu {len(products)} sản phẩm. <a href="{download_url}" class="btn btn-primary mt-2">Tải xuống kết quả</a>', 'success')
            
            # Tạo dữ liệu để hiển thị trên giao diện
            zip_size = os.path.getsize(zip_path) / (1024 * 1024)  # Kích thước MB
            products_count = len(products)
            stats = {
                'Tổng sản phẩm': products_count,
                'Thời gian xử lý': f"{os.path.basename(result_dir).split('_')[-1]} giây",
                'Kích thước file': f"{zip_size:.2f} MB"
            }
            
            # Trả về trang kết quả với đường dẫn tải xuống và thống kê
            return render_template('crawler_result.html', 
                                  download_url=download_url,
                                  zip_filename=os.path.basename(zip_path),
                                  stats=stats,
                                  products_count=products_count)
        else:
            flash(f'Đã cào xong dữ liệu {len(products)} sản phẩm nhưng không tạo được file ZIP.', 'warning')
            return redirect(url_for('main.index'))
    
    except Exception as e:
        traceback.print_exc()
        flash(f'Lỗi: {str(e)}', 'error')
        return redirect(url_for('main.index'))